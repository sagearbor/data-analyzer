"""
Data Quality Analyzer - Fixed Clean Design
Professional interface with all issues resolved
"""

import streamlit as st
import pandas as pd
import json
import base64
import io
import asyncio
import csv
from typing import Dict, Any, Optional
from datetime import datetime

# Import custom modules
from demo_dictionaries import DEMO_DICTIONARIES, get_demo_dictionary
from mermaid_renderer import render_mermaid

# Import navigation menu
from streamlit_option_menu import option_menu

# Configure Streamlit page
st.set_page_config(
    page_title="Data Quality Analyzer",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Modern, clean CSS with fixes
st.markdown("""
<style>
    /* Hide Streamlit default elements */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}
    .stDeployButton {display: none;}

    /* Remove Streamlit's default padding */
    .block-container {
        padding-top: 2rem !important;
        padding-bottom: 0 !important;
    }

    /* Modern color scheme */
    :root {
        --primary-color: #2563eb;
        --primary-hover: #1d4ed8;
        --success-color: #10b981;
        --warning-color: #f59e0b;
        --danger-color: #ef4444;
        --bg-color: #ffffff;
        --bg-secondary: #f8f9fa;
        --navbar-bg: #1e293b;
        --text-primary: #111827;
        --text-secondary: #6b7280;
        --border-color: #e5e7eb;
        --shadow-sm: 0 1px 2px 0 rgba(0, 0, 0, 0.05);
        --shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1);
    }

    /* Clean background */
    .stApp {
        background: var(--bg-color);
    }

    /* Slim fixed navbar */
    .navbar {
        position: fixed;
        top: 0;
        left: 0;
        right: 0;
        height: 48px;
        background: var(--navbar-bg);
        z-index: 1000;
        box-shadow: var(--shadow);
        display: flex;
        align-items: center;
        padding: 0 20px;
    }

    /* Content offset for fixed navbar */
    .main-content {
        margin-top: 60px;
    }

    /* Navigation styling override */
    [data-testid="stHorizontalBlock"] > div:first-child {
        background: var(--navbar-bg) !important;
        padding: 0 !important;
        height: 48px !important;
        position: fixed !important;
        top: 0 !important;
        left: 0 !important;
        right: 0 !important;
        z-index: 1000 !important;
        margin-bottom: 0 !important;
        border-bottom: none !important;
    }

    /* Adjust navigation menu container */
    div[data-baseweb="base-provider"] {
        margin-top: 0 !important;
    }

    .st-emotion-cache-1rtdyuf {
        margin-top: 60px;
    }

    /* Modern cards */
    .card {
        background: white;
        border: 1px solid var(--border-color);
        border-radius: 8px;
        padding: 20px;
        margin-bottom: 16px;
        box-shadow: var(--shadow-sm);
    }

    /* Primary button styling */
    .stButton > button {
        background-color: var(--primary-color);
        color: white;
        border: none;
        border-radius: 6px;
        padding: 10px 24px;
        font-weight: 500;
        font-size: 16px;
        transition: all 0.2s ease;
        box-shadow: var(--shadow-sm);
    }

    .stButton > button:hover:not(:disabled) {
        background-color: var(--primary-hover);
        box-shadow: var(--shadow);
        transform: translateY(-1px);
    }

    .stButton > button:disabled {
        background-color: #d1d5db;
        cursor: not-allowed;
        transform: none;
    }

    /* File uploader styling */
    .stFileUploader {
        border: 2px dashed var(--border-color);
        border-radius: 8px;
        padding: 20px;
        background: var(--bg-secondary);
        transition: all 0.2s ease;
        text-align: center;
        min-height: 140px;
    }

    .stFileUploader:hover {
        border-color: var(--primary-color);
        background: #f0f9ff;
    }

    /* File uploader with data loaded */
    .file-loaded {
        border-color: var(--success-color);
        background: #d1fae5;
    }

    /* Clean typography */
    h1 {
        color: var(--text-primary);
        font-size: 28px;
        font-weight: 700;
        margin-bottom: 4px;
        letter-spacing: -0.025em;
    }

    h2 {
        color: var(--text-primary);
        font-size: 20px;
        font-weight: 600;
        margin-bottom: 12px;
    }

    h3 {
        color: var(--text-primary);
        font-size: 18px;
        font-weight: 600;
        margin-bottom: 8px;
    }

    p {
        color: var(--text-secondary);
        line-height: 1.5;
    }

    /* Sidebar styling */
    .css-1d391kg {
        background: var(--bg-secondary);
        border-right: 1px solid var(--border-color);
        padding-top: 60px !important;
    }

    /* Remove weird rectangles */
    .st-emotion-cache-16idsys p {
        margin: 0;
    }

    /* Compact layout */
    .row-widget.stHorizontal > div {
        padding: 0 8px;
    }

    /* Success message styling */
    .stSuccess {
        background-color: #d1fae5;
        color: #065f46;
        border: 1px solid #a7f3d0;
        padding: 8px 12px;
        border-radius: 6px;
    }

    /* Tab styling */
    .stTabs [data-baseweb="tab-list"] {
        background-color: transparent;
        gap: 24px;
        border-bottom: 1px solid var(--border-color);
    }

    .stTabs [data-baseweb="tab"] {
        background-color: transparent;
        border: none;
        color: var(--text-secondary);
        font-weight: 500;
        padding: 8px 4px;
        border-bottom: 2px solid transparent;
    }

    .stTabs [aria-selected="true"] {
        color: var(--primary-color);
        border-bottom-color: var(--primary-color);
    }

    /* Mobile responsiveness */
    @media (max-width: 768px) {
        h1 {
            font-size: 24px;
        }

        .stFileUploader {
            min-height: 100px;
            padding: 16px;
        }
    }
</style>
""", unsafe_allow_html=True)


class MCPClient:
    """MCP client for data analysis with proper validation"""

    def __init__(self):
        pass

    def _simulate_mcp_call(self, request_data: Dict, debug: bool = False) -> Dict:
        """Simulate MCP server call with proper validation"""
        tool = request_data.get("tool")
        args = request_data.get("arguments", {})

        if tool == "analyze_data":
            try:
                data_content = args.get("data_content", "")
                file_format = args.get("file_format", "csv")
                schema = args.get("schema", {})
                rules = args.get("rules", {})
                min_rows = args.get("min_rows", 1)

                # Parse data based on format
                df = None
                if file_format.lower() == "csv":
                    df = pd.read_csv(io.StringIO(data_content))
                elif file_format.lower() == "json":
                    json_data = json.loads(data_content)
                    if isinstance(json_data, list):
                        df = pd.json_normalize(json_data)
                    elif isinstance(json_data, dict):
                        first_key = list(json_data.keys())[0] if json_data else None
                        if first_key and isinstance(json_data[first_key], list):
                            df = pd.json_normalize(json_data[first_key])
                        else:
                            df = pd.json_normalize([json_data])
                elif file_format.lower() in ["xlsx", "xls"]:
                    excel_bytes = base64.b64decode(data_content)
                    df = pd.read_excel(io.BytesIO(excel_bytes))
                elif file_format.lower() == "parquet":
                    parquet_bytes = base64.b64decode(data_content)
                    df = pd.read_parquet(io.BytesIO(parquet_bytes))
                else:
                    return {"error": f"Unsupported format: {file_format}"}

                if df is None or df.empty:
                    return {"error": "No data could be loaded"}

                # Analysis with proper validation
                issues = []
                row_count = len(df)

                # Check minimum rows
                if row_count < min_rows:
                    issues.append({
                        "type": "row_count",
                        "severity": "error",
                        "message": f"Row count ({row_count}) is less than minimum required ({min_rows})"
                    })

                # Check for invalid numeric values
                for col in df.columns:
                    if df[col].dtype == 'object':
                        # Check if column should be numeric
                        numeric_pattern = False

                        # Check if most values look numeric
                        non_null_values = df[col].dropna()
                        if len(non_null_values) > 0:
                            numeric_count = 0
                            for val in non_null_values:
                                try:
                                    float(str(val))
                                    numeric_count += 1
                                except:
                                    pass

                            # If more than 50% look numeric, it should probably be numeric
                            if numeric_count / len(non_null_values) > 0.5:
                                numeric_pattern = True

                        # Check for specific invalid values
                        for idx, val in df[col].items():
                            if pd.notna(val):
                                val_str = str(val).lower()
                                # Check for common invalid values
                                if val_str in ['invalid', 'error', 'n/a', 'null', 'none', 'invalid-date']:
                                    issues.append({
                                        "type": "invalid_value",
                                        "severity": "error",
                                        "column": col,
                                        "row": idx,
                                        "value": val,
                                        "message": f"Invalid value '{val}' found in column '{col}' at row {idx}"
                                    })
                                elif numeric_pattern and val_str not in ['0', '1', '2', '3', '4', '5', '6', '7', '8', '9']:
                                    # Check if this should be numeric but isn't
                                    try:
                                        float(val_str)
                                    except:
                                        issues.append({
                                            "type": "type_error",
                                            "severity": "error",
                                            "column": col,
                                            "row": idx,
                                            "value": val,
                                            "message": f"Non-numeric value '{val}' in numeric column '{col}' at row {idx}"
                                        })

                    # Check for out-of-range values
                    elif df[col].dtype in ['int64', 'float64']:
                        # Check for unrealistic values
                        if 'age' in col.lower():
                            invalid_ages = df[(df[col] < 0) | (df[col] > 150)]
                            for idx, row in invalid_ages.iterrows():
                                issues.append({
                                    "type": "range_violation",
                                    "severity": "error",
                                    "column": col,
                                    "row": idx,
                                    "value": row[col],
                                    "message": f"Invalid age value {row[col]} at row {idx}"
                                })

                # Data type validation with schema
                if schema:
                    for col, expected_type in schema.items():
                        if col in df.columns:
                            actual_type = str(df[col].dtype)
                            type_match = False

                            if expected_type == "int" and "int" in actual_type:
                                type_match = True
                            elif expected_type == "float" and ("float" in actual_type or "int" in actual_type):
                                type_match = True
                            elif expected_type == "str" and "object" in actual_type:
                                type_match = True
                            elif expected_type == "bool" and "bool" in actual_type:
                                type_match = True
                            elif expected_type == "datetime" and "datetime" in actual_type:
                                type_match = True

                            if not type_match:
                                issues.append({
                                    "type": "type_mismatch",
                                    "severity": "warning",
                                    "column": col,
                                    "message": f"Column '{col}' has type '{actual_type}' but expected '{expected_type}'"
                                })

                # Value range validation with rules
                if rules:
                    for col, col_rules in rules.items():
                        if col in df.columns and df[col].dtype in ['int64', 'float64']:
                            if "min" in col_rules:
                                min_val = col_rules["min"]
                                violations = df[df[col] < min_val]
                                for idx, row in violations.iterrows():
                                    issues.append({
                                        "type": "range_violation",
                                        "severity": "error",
                                        "column": col,
                                        "row": idx,
                                        "value": row[col],
                                        "message": f"Value {row[col]} below minimum {min_val} in column '{col}' at row {idx}"
                                    })

                            if "max" in col_rules:
                                max_val = col_rules["max"]
                                violations = df[df[col] > max_val]
                                for idx, row in violations.iterrows():
                                    issues.append({
                                        "type": "range_violation",
                                        "severity": "error",
                                        "column": col,
                                        "row": idx,
                                        "value": row[col],
                                        "message": f"Value {row[col]} above maximum {max_val} in column '{col}' at row {idx}"
                                    })

                # Check for missing values
                missing_counts = df.isnull().sum()
                for col, count in missing_counts.items():
                    if count > 0:
                        issues.append({
                            "type": "missing_values",
                            "severity": "warning",
                            "column": col,
                            "count": int(count),
                            "message": f"Column '{col}' has {count} missing values"
                        })

                # Summary statistics
                summary = {
                    "row_count": row_count,
                    "column_count": len(df.columns),
                    "columns": df.columns.tolist(),
                    "dtypes": {col: str(dtype) for col, dtype in df.dtypes.items()},
                    "missing_values": missing_counts.to_dict(),
                    "issues_count": len(issues),
                    "issues": issues,
                    "data": df
                }

                return {
                    "success": True,
                    "summary": summary,
                    "preview": df.head(10).to_dict('records') if not df.empty else []
                }

            except Exception as e:
                return {"error": str(e)}

        elif tool == "parse_data_dictionary":
            try:
                dictionary_content = args.get("dictionary_content", "")
                schema = {}
                rules = {}

                lines = dictionary_content.strip().split('\n')
                if ',' in lines[0] or '\t' in lines[0]:
                    reader = csv.DictReader(io.StringIO(dictionary_content))
                    for row in reader:
                        field_name = row.get('field_name') or row.get('column') or row.get('name')
                        if field_name:
                            field_type = row.get('type') or row.get('data_type') or 'str'
                            schema[field_name] = field_type

                            field_rules = {}
                            if row.get('min'):
                                try:
                                    field_rules['min'] = float(row['min'])
                                except:
                                    pass
                            if row.get('max'):
                                try:
                                    field_rules['max'] = float(row['max'])
                                except:
                                    pass
                            if field_rules:
                                rules[field_name] = field_rules

                return {
                    "success": True,
                    "schema": schema,
                    "rules": rules
                }
            except Exception as e:
                return {"error": str(e)}

        return {"error": "Unknown tool"}

    async def analyze_data(self, data_content: str, file_format: str = "csv",
                          schema: Dict = None, rules: Dict = None,
                          min_rows: int = 1, debug: bool = False):
        """Analyze data with optional schema and rules"""
        request_data = {
            "tool": "analyze_data",
            "arguments": {
                "data_content": data_content,
                "file_format": file_format,
                "schema": schema or {},
                "rules": rules or {},
                "min_rows": min_rows
            }
        }
        return self._simulate_mcp_call(request_data, debug)

    async def parse_data_dictionary(self, dictionary_content: str,
                                   format_hint: str = "auto",
                                   use_cache: bool = True,
                                   debug: bool = False) -> Dict:
        """Parse data dictionary"""
        request_data = {
            "tool": "parse_data_dictionary",
            "arguments": {
                "dictionary_content": dictionary_content,
                "format_hint": format_hint,
                "use_cache": use_cache
            }
        }
        return self._simulate_mcp_call(request_data, debug)


def main():
    """Main application entry point"""

    # Initialize session state
    if 'analysis_complete' not in st.session_state:
        st.session_state.analysis_complete = False
    if 'analysis_results' not in st.session_state:
        st.session_state.analysis_results = None
    if 'uploaded_file' not in st.session_state:
        st.session_state.uploaded_file = None
    if 'uploaded_file_name' not in st.session_state:
        st.session_state.uploaded_file_name = None
    if 'dictionary_file' not in st.session_state:
        st.session_state.dictionary_file = None
    if 'dictionary_file_name' not in st.session_state:
        st.session_state.dictionary_file_name = None

    # Create slim navigation bar
    selected = option_menu(
        menu_title=None,
        options=["Analyze", "About"],
        icons=["graph-up", "info-circle"],
        menu_icon=None,
        default_index=0,
        orientation="horizontal",
        styles={
            "container": {
                "padding": "0px",
                "background-color": "#1e293b",
                "margin": "0px",
                "position": "fixed",
                "top": "0",
                "width": "100%",
                "z-index": "1000",
                "height": "48px"
            },
            "icon": {"color": "#94a3b8", "font-size": "16px"},
            "nav-link": {
                "font-size": "15px",
                "text-align": "center",
                "margin": "0px",
                "padding": "12px 20px",
                "color": "#e2e8f0",
                "--hover-color": "#334155",
                "height": "48px"
            },
            "nav-link-selected": {
                "background-color": "#3b82f6",
                "color": "white",
                "font-weight": "500"
            }
        }
    )

    # Add spacing after navbar
    st.markdown("<div style='height: 60px;'></div>", unsafe_allow_html=True)

    # Route to pages
    if selected == "Analyze":
        analyze_page()
    else:
        about_page()


def analyze_page():
    """Main analysis page with three-element layout"""

    st.title("Data Quality Analyzer")
    st.markdown("<p style='color: #6b7280; margin-top: -16px; margin-bottom: 24px;'>Upload your data, optionally add validation rules, and analyze</p>", unsafe_allow_html=True)

    # Main three-element upload area
    col1, col2, col3 = st.columns([1.2, 1.2, 1])

    with col1:
        st.markdown("#### 📁 Upload Data File")

        # Show loaded file info if demo data is selected
        if st.session_state.uploaded_file_name:
            st.success(f"✓ Loaded: {st.session_state.uploaded_file_name}")

        uploaded_file = st.file_uploader(
            "Drag and drop or browse",
            type=['csv', 'json', 'xlsx', 'xls', 'parquet'],
            key="main_data_upload",
            label_visibility="collapsed",
            help="CSV, JSON, Excel, or Parquet files"
        )

        if uploaded_file:
            file_size = len(uploaded_file.read())
            uploaded_file.seek(0)
            st.session_state.uploaded_file = uploaded_file
            st.session_state.uploaded_file_name = uploaded_file.name
            st.success(f"✓ {uploaded_file.name} ({file_size:,} bytes)")

    with col2:
        st.markdown("#### 📚 Upload Data Dictionary")
        st.markdown("<small style='color: #6b7280;'>Optional - defines validation rules</small>", unsafe_allow_html=True)

        # Show loaded dictionary info if demo is selected
        if st.session_state.dictionary_file_name:
            st.success(f"✓ Loaded: {st.session_state.dictionary_file_name}")

        dict_file = st.file_uploader(
            "Drag and drop or browse",
            type=['csv', 'json', 'txt'],
            key="dict_upload",
            label_visibility="collapsed",
            help="CSV or JSON dictionary file"
        )

        if dict_file:
            st.session_state.dictionary_file = dict_file
            st.session_state.dictionary_file_name = dict_file.name
            st.success(f"✓ {dict_file.name}")

    with col3:
        st.markdown("#### ⚡ Analyze")
        st.markdown("<div style='height: 24px;'></div>", unsafe_allow_html=True)

        # Enable analyze button only if data is uploaded (dictionary is optional)
        analyze_enabled = st.session_state.uploaded_file is not None

        if st.button("🚀 Run Analysis",
                    disabled=not analyze_enabled,
                    type="primary",
                    use_container_width=True,
                    help="Upload data file to enable analysis"):
            run_analysis()

    # Sidebar for additional options
    with st.sidebar:
        st.markdown("### Quick Actions")

        # Demo data section
        with st.expander("📂 Load Demo Data", expanded=False):
            st.markdown("**Sample Datasets**")
            demo_data_choice = st.selectbox(
                "Choose demo data",
                ["None", "CSV - Western Names", "CSV - Asian Names",
                 "JSON - Mixed Names", "CSV - Clinical Trial"],
                key="demo_data_select"
            )

            if demo_data_choice != "None":
                if demo_data_choice == "CSV - Western Names":
                    demo_data = """name,age,email,country
John Smith,25,john@email.com,USA
Jane Doe,30,jane@email.com,Canada
Bob Wilson,invalid,bob@email,UK"""
                elif demo_data_choice == "CSV - Asian Names":
                    demo_data = """name,age,city,join_date
李明,28,Beijing,2023-01-15
田中太郎,35,Tokyo,2023-02-20
김철수,42,Seoul,invalid-date"""
                elif demo_data_choice == "JSON - Mixed Names":
                    demo_data = json.dumps([
                        {"name": "Alice Smith", "age": 30, "score": 85.5},
                        {"name": "王小明", "age": 25, "score": 92.0},
                        {"name": "Carlos García", "age": "invalid", "score": 88.0}
                    ])
                elif demo_data_choice == "CSV - Clinical Trial":
                    demo_data = """patient_id,age,treatment,outcome,date
P001,45,Drug_A,Improved,2023-01-15
P002,52,Drug_B,No_Change,2023-01-20
P003,999,Drug_A,Improved,invalid"""

                st.session_state.uploaded_file = demo_data
                st.session_state.uploaded_file_name = f"Demo: {demo_data_choice}"
                st.success(f"✅ Loaded {demo_data_choice}")
                st.rerun()

            st.markdown("**Sample Dictionaries**")
            demo_dict_choice = st.selectbox(
                "Choose demo dictionary",
                ["None"] + list(DEMO_DICTIONARIES.keys()),
                key="demo_dict_select"
            )

            if demo_dict_choice != "None":
                demo_dict = get_demo_dictionary(demo_dict_choice)
                st.session_state.dictionary_file = demo_dict
                st.session_state.dictionary_file_name = f"Demo: {demo_dict_choice}"
                st.success(f"✅ Loaded {demo_dict_choice}")
                st.rerun()

        # Results section (shown after analysis)
        if st.session_state.analysis_complete:
            st.markdown("### 📊 Results Summary")
            show_results_sidebar()

        # Export section (shown after analysis)
        if st.session_state.analysis_complete:
            st.markdown("### 💾 Export Options")
            show_export_sidebar()

    # Main content area for results
    if st.session_state.analysis_complete:
        st.markdown("---")
        show_results_main()


def run_analysis():
    """Execute the data analysis with proper validation"""
    with st.spinner("Analyzing your data..."):
        progress_bar = st.progress(0)
        status_text = st.empty()

        # Analysis stages
        stages = [
            (0.25, "Loading data..."),
            (0.50, "Checking data quality..."),
            (0.75, "Validating against rules..."),
            (1.0, "Generating report...")
        ]

        client = MCPClient()

        # Prepare data
        if isinstance(st.session_state.uploaded_file, str):
            data_content = st.session_state.uploaded_file
            file_format = "json" if data_content.startswith('[') or data_content.startswith('{') else "csv"
        else:
            file_ext = st.session_state.uploaded_file.name.split('.')[-1].lower()
            file_format = file_ext

            if file_ext in ['xlsx', 'xls', 'parquet']:
                file_bytes = st.session_state.uploaded_file.read()
                data_content = base64.b64encode(file_bytes).decode('utf-8')
                st.session_state.uploaded_file.seek(0)
            else:
                data_content = st.session_state.uploaded_file.read().decode('utf-8')
                st.session_state.uploaded_file.seek(0)

        # Parse dictionary if provided (optional)
        schema = {}
        rules = {}
        if st.session_state.dictionary_file:
            if isinstance(st.session_state.dictionary_file, str):
                dict_content = st.session_state.dictionary_file
            else:
                dict_content = st.session_state.dictionary_file.read().decode('utf-8')
                st.session_state.dictionary_file.seek(0)

            dict_result = asyncio.run(
                client.parse_data_dictionary(dict_content)
            )

            if dict_result.get("success"):
                schema = dict_result.get("schema", {})
                rules = dict_result.get("rules", {})

        # Update progress
        for progress, message in stages:
            status_text.text(message)
            progress_bar.progress(progress)
            asyncio.run(asyncio.sleep(0.2))

        # Run analysis
        result = asyncio.run(
            client.analyze_data(
                data_content=data_content,
                file_format=file_format,
                schema=schema,
                rules=rules,
                min_rows=1
            )
        )

        if result.get("success"):
            st.session_state.analysis_complete = True
            st.session_state.analysis_results = result
            progress_bar.empty()
            status_text.empty()
            st.success("✅ Analysis complete!")
            st.rerun()
        else:
            progress_bar.empty()
            status_text.empty()
            st.error(f"❌ Analysis failed: {result.get('error')}")


def show_results_main():
    """Display main results area"""
    results = st.session_state.analysis_results
    summary = results.get("summary", {})

    st.markdown("## 📊 Analysis Results")

    # Metrics row
    col1, col2, col3, col4 = st.columns(4)

    with col1:
        st.metric("Total Rows", f"{summary.get('row_count', 0):,}")

    with col2:
        st.metric("Columns", summary.get('column_count', 0))

    with col3:
        issues_count = summary.get('issues_count', 0)
        st.metric("Issues Found", issues_count,
                 delta="✓ Clean" if issues_count == 0 else f"{issues_count} issues",
                 delta_color="normal" if issues_count == 0 else "inverse")

    with col4:
        missing = sum(summary.get('missing_values', {}).values())
        st.metric("Missing Values", missing)

    # Issues details
    if summary.get('issues'):
        st.markdown("### ⚠️ Data Quality Issues")

        # Group issues by type
        errors = [i for i in summary['issues'] if i.get('severity') == 'error']
        warnings = [i for i in summary['issues'] if i.get('severity') == 'warning']

        if errors:
            st.markdown("#### 🔴 Errors")
            for error in errors[:10]:  # Show first 10 errors
                st.error(f"**{error.get('type', 'Error')}**: {error.get('message', 'Unknown error')}")
            if len(errors) > 10:
                st.warning(f"... and {len(errors) - 10} more errors")

        if warnings:
            st.markdown("#### 🟡 Warnings")
            for warning in warnings[:5]:  # Show first 5 warnings
                st.warning(f"**{warning.get('type', 'Warning')}**: {warning.get('message', 'Unknown warning')}")
            if len(warnings) > 5:
                st.info(f"... and {len(warnings) - 5} more warnings")

    # Data preview
    if results.get('preview'):
        st.markdown("### 📋 Data Preview")
        df_preview = pd.DataFrame(results['preview'])
        st.dataframe(df_preview, use_container_width=True, height=300)


def show_results_sidebar():
    """Show results summary in sidebar"""
    results = st.session_state.analysis_results
    summary = results.get("summary", {})

    errors = len([i for i in summary.get('issues', []) if i.get('severity') == 'error'])
    warnings = len([i for i in summary.get('issues', []) if i.get('severity') == 'warning'])

    st.info(f"""
    **Quick Stats**
    - Rows: {summary.get('row_count', 0):,}
    - Columns: {summary.get('column_count', 0)}
    - Errors: {errors}
    - Warnings: {warnings}
    """)


def show_export_sidebar():
    """Show export options in sidebar"""
    results = st.session_state.analysis_results
    summary = results.get("summary", {})

    # Excel export
    if st.button("📊 Export to Excel", use_container_width=True):
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Data Analysis"

            # Get data
            df = summary.get('data', pd.DataFrame(results.get('preview', [])))

            # Write headers
            for col_idx, col_name in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = Font(bold=True)

            # Write data with error highlighting
            error_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")

            for row_idx, row_data in df.iterrows():
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx + 2, column=col_idx, value=value)

                    # Check if this cell has errors
                    for issue in summary.get('issues', []):
                        if (issue.get('row') == row_idx and
                            issue.get('column') == df.columns[col_idx - 1]):
                            cell.fill = error_fill

            # Save
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)

            st.download_button(
                label="⬇️ Download Excel",
                data=excel_buffer.getvalue(),
                file_name=f"analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except ImportError:
            st.error("Install openpyxl: pip install openpyxl")

    # CSV export
    if st.button("📄 Export to CSV", use_container_width=True):
        df = summary.get('data', pd.DataFrame(results.get('preview', [])))
        csv_buffer = io.StringIO()
        df.to_csv(csv_buffer, index=False)

        st.download_button(
            label="⬇️ Download CSV",
            data=csv_buffer.getvalue(),
            file_name=f"analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            mime="text/csv"
        )


def about_page():
    """About page with information"""
    st.title("About Data Quality Analyzer")

    col1, col2, col3 = st.columns([1, 2, 1])

    with col2:
        st.markdown("""
        ### Professional Data Validation Tool

        Data Quality Analyzer helps data professionals validate and clean their datasets efficiently.
        It provides instant feedback on data quality issues with detailed error reporting.

        #### ✨ Key Features
        - Support for CSV, JSON, Excel, and Parquet files
        - Automatic detection of invalid values like "invalid", "error", "n/a"
        - Data type validation and range checking
        - Optional data dictionary for custom rules
        - Export results with error highlighting

        #### 🚀 How to Use
        1. **Upload your data file** - Required for analysis
        2. **Optionally add a data dictionary** - Define custom validation rules
        3. **Click Analyze** - Get instant quality report
        4. **Export results** - Download with errors highlighted

        #### 📊 What We Check
        - Invalid values (text in numeric fields)
        - Out-of-range values (e.g., age > 150)
        - Missing or null values
        - Data type mismatches
        - Custom rule violations

        #### 🛠️ Technology
        - Built with Streamlit and Python
        - Uses Pandas for data processing
        - Model Context Protocol (MCP) for analysis
        - Deployable on Azure Container Apps

        ---

        *Version 1.1.0 | © 2024 Data Quality Analyzer*
        """)


if __name__ == "__main__":
    main()