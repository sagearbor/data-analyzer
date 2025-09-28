"""
Data Quality Analyzer - Improved UI with Navigation
Professional, minimalist interface with navigation bar and better visual design
"""

import streamlit as st
import pandas as pd
import json
import base64
import io
import asyncio
import csv
from typing import Dict, Any, Optional
from datetime import datetime

# Import custom modules
from demo_dictionaries import DEMO_DICTIONARIES, get_demo_dictionary
from mermaid_renderer import render_mermaid

# Import navigation menu
from streamlit_option_menu import option_menu

# Configure Streamlit page
st.set_page_config(
    page_title="Data Quality Analyzer",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# Enhanced CSS for professional design
st.markdown("""
<style>
    /* Modern gradient background */
    .stApp {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
    }

    /* Custom navbar styling */
    .nav-container {
        background: rgba(255, 255, 255, 0.95);
        border-radius: 10px;
        padding: 10px;
        margin-bottom: 20px;
        box-shadow: 0 2px 10px rgba(0,0,0,0.1);
    }

    /* Professional card styling */
    .main-card {
        background: white;
        border-radius: 12px;
        padding: 2rem;
        box-shadow: 0 4px 6px rgba(0,0,0,0.1);
        margin-bottom: 1.5rem;
        transition: transform 0.3s ease;
    }

    .main-card:hover {
        transform: translateY(-2px);
        box-shadow: 0 6px 12px rgba(0,0,0,0.15);
    }

    /* Button styling */
    .stButton > button {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        border: none;
        padding: 0.75rem 2rem;
        font-size: 1.1rem;
        font-weight: 500;
        border-radius: 8px;
        transition: all 0.3s ease;
        width: 100%;
    }

    .stButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 4px 12px rgba(102, 126, 234, 0.4);
    }

    /* File uploader styling */
    .stFileUploader {
        border: 2px dashed #667eea;
        border-radius: 12px;
        padding: 2rem;
        background: rgba(102, 126, 234, 0.05);
        transition: all 0.3s ease;
    }

    .stFileUploader:hover {
        border-color: #764ba2;
        background: rgba(102, 126, 234, 0.1);
    }

    /* Headers */
    h1 {
        color: white;
        text-align: center;
        font-weight: 600;
        text-shadow: 2px 2px 4px rgba(0,0,0,0.2);
        margin-bottom: 0.5rem;
    }

    h2 {
        color: #2c3e50;
        font-weight: 500;
        border-bottom: 2px solid #667eea;
        padding-bottom: 0.5rem;
        margin-bottom: 1rem;
    }

    h3 {
        color: #34495e;
        font-weight: 500;
    }

    .subtitle {
        text-align: center;
        color: rgba(255, 255, 255, 0.9);
        font-size: 1.2rem;
        margin-bottom: 2rem;
    }

    /* Success/Error/Warning messages */
    .success-msg {
        background: linear-gradient(135deg, #84fab0 0%, #8fd3f4 100%);
        color: #0c5d35;
        padding: 1rem;
        border-radius: 8px;
        margin: 1rem 0;
        font-weight: 500;
    }

    .error-msg {
        background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%);
        color: white;
        padding: 1rem;
        border-radius: 8px;
        margin: 1rem 0;
        font-weight: 500;
    }

    .warning-msg {
        background: linear-gradient(135deg, #fa709a 0%, #fee140 100%);
        color: #5a3001;
        padding: 1rem;
        border-radius: 8px;
        margin: 1rem 0;
        font-weight: 500;
    }

    /* Progress indicator */
    .progress-indicator {
        background: rgba(255, 255, 255, 0.9);
        border-radius: 8px;
        padding: 1rem;
        text-align: center;
        margin: 1rem 0;
    }

    /* Tab styling */
    .stTabs [data-baseweb="tab-list"] {
        background: white;
        border-radius: 8px;
        padding: 0.5rem;
    }

    .stTabs [data-baseweb="tab"] {
        font-weight: 500;
        color: #667eea;
    }

    .stTabs [aria-selected="true"] {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white !important;
    }

    /* Metric cards */
    .metric-card {
        background: white;
        border-radius: 8px;
        padding: 1.5rem;
        text-align: center;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
    }

    .metric-value {
        font-size: 2rem;
        font-weight: 700;
        color: #667eea;
    }

    .metric-label {
        color: #7f8c8d;
        font-size: 0.9rem;
        text-transform: uppercase;
        letter-spacing: 1px;
    }

    /* Hide Streamlit branding */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}

    /* Expander styling */
    .streamlit-expanderHeader {
        background: rgba(102, 126, 234, 0.1);
        border-radius: 8px;
        font-weight: 500;
    }
</style>
""", unsafe_allow_html=True)


class MCPClient:
    """Simplified MCP client for data analysis"""

    def __init__(self):
        pass

    def _simulate_mcp_call(self, request_data: Dict, debug: bool = False) -> Dict:
        """Simulate MCP server call for development"""
        tool = request_data.get("tool")
        args = request_data.get("arguments", {})

        if tool == "analyze_data":
            try:
                data_content = args.get("data_content", "")
                file_format = args.get("file_format", "csv")
                schema = args.get("schema", {})
                rules = args.get("rules", {})
                min_rows = args.get("min_rows", 1)

                # Parse data based on format
                df = None
                if file_format.lower() == "csv":
                    df = pd.read_csv(io.StringIO(data_content))
                elif file_format.lower() == "json":
                    json_data = json.loads(data_content)
                    if isinstance(json_data, list):
                        df = pd.json_normalize(json_data)
                    elif isinstance(json_data, dict):
                        first_key = list(json_data.keys())[0] if json_data else None
                        if first_key and isinstance(json_data[first_key], list):
                            df = pd.json_normalize(json_data[first_key])
                        else:
                            df = pd.json_normalize([json_data])
                elif file_format.lower() in ["xlsx", "xls"]:
                    excel_bytes = base64.b64decode(data_content)
                    df = pd.read_excel(io.BytesIO(excel_bytes))
                elif file_format.lower() == "parquet":
                    parquet_bytes = base64.b64decode(data_content)
                    df = pd.read_parquet(io.BytesIO(parquet_bytes))
                else:
                    return {"error": f"Unsupported format: {file_format}"}

                if df is None or df.empty:
                    return {"error": "No data could be loaded"}

                # Basic analysis
                issues = []
                row_count = len(df)

                # Check minimum rows
                if row_count < min_rows:
                    issues.append({
                        "type": "row_count",
                        "severity": "error",
                        "message": f"Row count ({row_count}) is less than minimum required ({min_rows})"
                    })

                # Data type validation
                if schema:
                    for col, expected_type in schema.items():
                        if col in df.columns:
                            actual_type = str(df[col].dtype)
                            type_match = False

                            if expected_type == "int" and "int" in actual_type:
                                type_match = True
                            elif expected_type == "float" and "float" in actual_type:
                                type_match = True
                            elif expected_type == "str" and "object" in actual_type:
                                type_match = True
                            elif expected_type == "bool" and "bool" in actual_type:
                                type_match = True
                            elif expected_type == "datetime" and "datetime" in actual_type:
                                type_match = True

                            if not type_match:
                                issues.append({
                                    "type": "type_mismatch",
                                    "severity": "warning",
                                    "column": col,
                                    "message": f"Column '{col}' has type '{actual_type}' but expected '{expected_type}'"
                                })

                # Value range validation
                if rules:
                    for col, col_rules in rules.items():
                        if col in df.columns:
                            if "min" in col_rules:
                                min_val = col_rules["min"]
                                violations = df[df[col] < min_val][col].tolist()
                                if violations:
                                    issues.append({
                                        "type": "range_violation",
                                        "severity": "error",
                                        "column": col,
                                        "message": f"Column '{col}' has {len(violations)} values below minimum {min_val}",
                                        "rows": df[df[col] < min_val].index.tolist()
                                    })

                            if "max" in col_rules:
                                max_val = col_rules["max"]
                                violations = df[df[col] > max_val][col].tolist()
                                if violations:
                                    issues.append({
                                        "type": "range_violation",
                                        "severity": "error",
                                        "column": col,
                                        "message": f"Column '{col}' has {len(violations)} values above maximum {max_val}",
                                        "rows": df[df[col] > max_val].index.tolist()
                                    })

                # Calculate summary statistics
                summary = {
                    "row_count": row_count,
                    "column_count": len(df.columns),
                    "columns": df.columns.tolist(),
                    "dtypes": {col: str(dtype) for col, dtype in df.dtypes.items()},
                    "missing_values": df.isnull().sum().to_dict(),
                    "issues_count": len(issues),
                    "issues": issues,
                    "data": df  # Store dataframe for export
                }

                return {
                    "success": True,
                    "summary": summary,
                    "preview": df.head(10).to_dict('records') if not df.empty else []
                }

            except Exception as e:
                return {"error": str(e)}

        elif tool == "parse_data_dictionary":
            # [Dictionary parsing logic - same as before]
            try:
                dictionary_content = args.get("dictionary_content", "")
                schema = {}
                rules = {}

                # Parse dictionary logic (simplified)
                lines = dictionary_content.strip().split('\n')
                if ',' in lines[0] or '\t' in lines[0]:
                    reader = csv.DictReader(io.StringIO(dictionary_content))
                    for row in reader:
                        field_name = row.get('field_name') or row.get('column') or row.get('name')
                        if field_name:
                            field_type = row.get('type') or row.get('data_type') or 'str'
                            schema[field_name] = field_type

                            field_rules = {}
                            if row.get('min'):
                                try:
                                    field_rules['min'] = float(row['min'])
                                except:
                                    pass
                            if row.get('max'):
                                try:
                                    field_rules['max'] = float(row['max'])
                                except:
                                    pass
                            if field_rules:
                                rules[field_name] = field_rules

                return {
                    "success": True,
                    "schema": schema,
                    "rules": rules
                }
            except Exception as e:
                return {"error": str(e)}

        return {"error": "Unknown tool"}

    async def analyze_data(self, data_content: str, file_format: str = "csv",
                          schema: Dict = None, rules: Dict = None,
                          min_rows: int = 1, debug: bool = False):
        """Analyze data with optional schema and rules"""
        request_data = {
            "tool": "analyze_data",
            "arguments": {
                "data_content": data_content,
                "file_format": file_format,
                "schema": schema or {},
                "rules": rules or {},
                "min_rows": min_rows
            }
        }
        return self._simulate_mcp_call(request_data, debug)

    async def parse_data_dictionary(self, dictionary_content: str,
                                   format_hint: str = "auto",
                                   use_cache: bool = True,
                                   debug: bool = False) -> Dict:
        """Parse data dictionary"""
        request_data = {
            "tool": "parse_data_dictionary",
            "arguments": {
                "dictionary_content": dictionary_content,
                "format_hint": format_hint,
                "use_cache": use_cache
            }
        }
        return self._simulate_mcp_call(request_data, debug)


def main():
    """Main application with navigation"""

    # Initialize session state
    if 'analysis_complete' not in st.session_state:
        st.session_state.analysis_complete = False
    if 'analysis_results' not in st.session_state:
        st.session_state.analysis_results = None
    if 'uploaded_file' not in st.session_state:
        st.session_state.uploaded_file = None
    if 'dictionary_file' not in st.session_state:
        st.session_state.dictionary_file = None

    # Create horizontal navigation menu with custom styling
    selected = option_menu(
        menu_title=None,
        options=["🏠 Home", "📊 Analysis", "ℹ️ About", "❓ Help"],
        icons=["house-fill", "graph-up", "info-circle-fill", "question-circle-fill"],
        menu_icon="cast",
        default_index=0,
        orientation="horizontal",
        styles={
            "container": {
                "padding": "5px",
                "background-color": "rgba(255, 255, 255, 0.95)",
                "border-radius": "10px",
                "margin-bottom": "20px",
                "box-shadow": "0 2px 10px rgba(0,0,0,0.1)"
            },
            "icon": {"color": "#667eea", "font-size": "20px"},
            "nav-link": {
                "font-size": "16px",
                "text-align": "center",
                "margin": "0px",
                "padding": "10px 20px",
                "border-radius": "8px",
                "--hover-color": "rgba(102, 126, 234, 0.1)",
            },
            "nav-link-selected": {
                "background": "linear-gradient(135deg, #667eea 0%, #764ba2 100%)",
                "color": "white",
                "font-weight": "500"
            },
        }
    )

    # Route to different pages
    if selected == "🏠 Home":
        home_page()
    elif selected == "📊 Analysis":
        analysis_page()
    elif selected == "ℹ️ About":
        about_page()
    elif selected == "❓ Help":
        help_page()


def home_page():
    """Landing page with quick start"""

    st.markdown("<h1>📊 Data Quality Analyzer</h1>", unsafe_allow_html=True)
    st.markdown("<p class='subtitle'>Professional Data Validation Made Simple</p>", unsafe_allow_html=True)

    # Hero section
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.markdown("""
        <div class='main-card' style='text-align: center; padding: 3rem;'>
            <h2 style='border: none;'>Welcome to Data Quality Analyzer</h2>
            <p style='font-size: 1.1rem; color: #7f8c8d; margin: 1.5rem 0;'>
                Validate your data with confidence. Upload files, apply validation rules,
                and get instant quality reports with error highlighting.
            </p>
            <div style='margin-top: 2rem;'>
                <h3>✨ Key Features</h3>
                <ul style='text-align: left; display: inline-block; color: #34495e;'>
                    <li>📁 Support for CSV, JSON, Excel, and Parquet files</li>
                    <li>📏 Custom validation rules and data dictionaries</li>
                    <li>🎯 Instant error detection and highlighting</li>
                    <li>📊 Export results with visual error markers</li>
                    <li>⚡ Fast processing for files up to 100MB</li>
                </ul>
            </div>
        </div>
        """, unsafe_allow_html=True)

        if st.button("🚀 Start Analyzing", help="Go to the Analysis page"):
            st.switch_page("pages/analysis.py")  # Note: This would need pages setup

    # Quick stats
    st.markdown("---")
    col1, col2, col3, col4 = st.columns(4)

    with col1:
        st.markdown("""
        <div class='metric-card'>
            <div class='metric-value'>5+</div>
            <div class='metric-label'>File Formats</div>
        </div>
        """, unsafe_allow_html=True)

    with col2:
        st.markdown("""
        <div class='metric-card'>
            <div class='metric-value'>100MB</div>
            <div class='metric-label'>Max File Size</div>
        </div>
        """, unsafe_allow_html=True)

    with col3:
        st.markdown("""
        <div class='metric-card'>
            <div class='metric-value'><1s</div>
            <div class='metric-label'>Processing Time</div>
        </div>
        """, unsafe_allow_html=True)

    with col4:
        st.markdown("""
        <div class='metric-card'>
            <div class='metric-value'>∞</div>
            <div class='metric-label'>Validations</div>
        </div>
        """, unsafe_allow_html=True)


def analysis_page():
    """Main analysis page with tabbed interface"""

    st.markdown("<h1>📊 Data Analysis</h1>", unsafe_allow_html=True)
    st.markdown("<p class='subtitle'>Upload, Validate, and Export</p>", unsafe_allow_html=True)

    # Create tabs for better organization
    tab1, tab2, tab3, tab4 = st.tabs(["📁 Upload", "🔍 Analyze", "📊 Results", "💾 Export"])

    with tab1:
        upload_section()

    with tab2:
        analysis_section()

    with tab3:
        results_section()

    with tab4:
        export_section()


def upload_section():
    """File upload section"""
    col1, col2 = st.columns(2)

    with col1:
        st.markdown("<div class='main-card'>", unsafe_allow_html=True)
        st.markdown("### 📁 Upload Data File")

        uploaded_file = st.file_uploader(
            "Choose your data file",
            type=['csv', 'json', 'xlsx', 'xls', 'parquet'],
            help="Supported formats: CSV, JSON, Excel, Parquet"
        )

        if uploaded_file:
            st.session_state.uploaded_file = uploaded_file
            file_size = len(uploaded_file.read())
            uploaded_file.seek(0)
            st.success(f"✅ Loaded: {uploaded_file.name} ({file_size:,} bytes)")

        # Demo data option
        with st.expander("📂 Use Demo Data"):
            demo_choice = st.selectbox("Select demo dataset",
                ["None", "CSV - Western", "CSV - Asian", "JSON - Mixed", "CSV - Clinical"])

            if demo_choice != "None":
                # Load demo data
                if demo_choice == "CSV - Western":
                    demo_data = """name,age,email,country
John Smith,25,john@email.com,USA
Jane Doe,30,jane@email.com,Canada
Bob Wilson,invalid,bob@email,UK"""
                elif demo_choice == "CSV - Asian":
                    demo_data = """name,age,city,join_date
李明,28,Beijing,2023-01-15
田中太郎,35,Tokyo,2023-02-20
김철수,42,Seoul,invalid-date"""
                elif demo_choice == "JSON - Mixed":
                    demo_data = json.dumps([
                        {"name": "Alice Smith", "age": 30, "score": 85.5},
                        {"name": "王小明", "age": 25, "score": 92.0},
                        {"name": "Carlos García", "age": "invalid", "score": 88.0}
                    ])
                elif demo_choice == "CSV - Clinical":
                    demo_data = """patient_id,age,treatment,outcome,date
P001,45,Drug_A,Improved,2023-01-15
P002,52,Drug_B,No_Change,2023-01-20
P003,999,Drug_A,Improved,invalid"""

                st.session_state.uploaded_file = demo_data
                st.success(f"✅ Demo data loaded: {demo_choice}")

        st.markdown("</div>", unsafe_allow_html=True)

    with col2:
        st.markdown("<div class='main-card'>", unsafe_allow_html=True)
        st.markdown("### 📚 Upload Data Dictionary")
        st.markdown("*Optional - defines validation rules*")

        dict_file = st.file_uploader(
            "Choose dictionary file",
            type=['csv', 'json', 'txt'],
            help="Define expected schema and rules",
            key="dict_uploader"
        )

        if dict_file:
            st.session_state.dictionary_file = dict_file
            st.success(f"✅ Dictionary loaded: {dict_file.name}")

        # Demo dictionary option
        with st.expander("📖 Use Demo Dictionary"):
            dict_choice = st.selectbox("Select demo dictionary",
                ["None"] + list(DEMO_DICTIONARIES.keys()))

            if dict_choice != "None":
                demo_dict = get_demo_dictionary(dict_choice)
                st.session_state.dictionary_file = demo_dict
                st.success(f"✅ Demo dictionary loaded: {dict_choice}")

        st.markdown("</div>", unsafe_allow_html=True)


def analysis_section():
    """Analysis configuration section"""
    st.markdown("<div class='main-card'>", unsafe_allow_html=True)
    st.markdown("### ⚙️ Analysis Configuration")

    col1, col2 = st.columns(2)

    with col1:
        min_rows = st.number_input("Minimum expected rows", min_value=1, value=1)
        check_duplicates = st.checkbox("Check for duplicates", value=True)
        check_nulls = st.checkbox("Check for null values", value=True)

    with col2:
        confidence_level = st.slider("Confidence level (%)", 0, 100, 95)
        sample_size = st.slider("Sample size for preview", 5, 50, 10)
        debug_mode = st.checkbox("Show debug information", value=False)

    # Advanced options in expander
    with st.expander("🔧 Advanced Options"):
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("**Schema Configuration**")
            auto_detect = st.checkbox("Auto-detect data types", value=True)
            strict_mode = st.checkbox("Strict type checking", value=False)

        with col2:
            st.markdown("**Performance Options**")
            use_cache = st.checkbox("Use caching", value=True)
            parallel = st.checkbox("Parallel processing", value=False)

    # Run analysis button
    if st.button("🚀 Run Analysis", type="primary", use_container_width=True):
        if st.session_state.uploaded_file:
            run_analysis()
        else:
            st.error("⚠️ Please upload a file first")

    st.markdown("</div>", unsafe_allow_html=True)


def run_analysis():
    """Execute the analysis"""
    with st.spinner("🔍 Analyzing your data..."):
        progress = st.progress(0)
        status = st.empty()

        # Simulate progress stages
        stages = [
            (0.2, "📖 Parsing data dictionary..."),
            (0.4, "📁 Loading data file..."),
            (0.6, "🔍 Validating schema..."),
            (0.8, "📏 Checking validation rules..."),
            (1.0, "✅ Analysis complete!")
        ]

        client = MCPClient()

        # Get data content
        if isinstance(st.session_state.uploaded_file, str):
            data_content = st.session_state.uploaded_file
            file_format = "json" if data_content.startswith('[') or data_content.startswith('{') else "csv"
        else:
            file_ext = st.session_state.uploaded_file.name.split('.')[-1].lower()
            file_format = file_ext

            if file_ext in ['xlsx', 'xls', 'parquet']:
                file_bytes = st.session_state.uploaded_file.read()
                data_content = base64.b64encode(file_bytes).decode('utf-8')
                st.session_state.uploaded_file.seek(0)
            else:
                data_content = st.session_state.uploaded_file.read().decode('utf-8')
                st.session_state.uploaded_file.seek(0)

        # Parse dictionary if provided
        schema = {}
        rules = {}
        if st.session_state.dictionary_file:
            if isinstance(st.session_state.dictionary_file, str):
                dict_content = st.session_state.dictionary_file
            else:
                dict_content = st.session_state.dictionary_file.read().decode('utf-8')
                st.session_state.dictionary_file.seek(0)

            dict_result = asyncio.run(
                client.parse_data_dictionary(dict_content)
            )

            if dict_result.get("success"):
                schema = dict_result.get("schema", {})
                rules = dict_result.get("rules", {})

        # Update progress through stages
        for progress_val, message in stages:
            status.text(message)
            progress.progress(progress_val)
            asyncio.run(asyncio.sleep(0.5))  # Simulate processing time

        # Run analysis
        result = asyncio.run(
            client.analyze_data(
                data_content=data_content,
                file_format=file_format,
                schema=schema,
                rules=rules,
                min_rows=1
            )
        )

        if result.get("success"):
            st.session_state.analysis_complete = True
            st.session_state.analysis_results = result
            st.success("✅ Analysis complete! Check the Results tab.")
        else:
            st.error(f"❌ Analysis failed: {result.get('error', 'Unknown error')}")

        progress.empty()
        status.empty()


def results_section():
    """Display analysis results"""
    if not st.session_state.analysis_complete:
        st.info("💡 Run analysis first to see results")
        return

    results = st.session_state.analysis_results
    summary = results.get("summary", {})

    # Summary dashboard
    st.markdown("<div class='main-card'>", unsafe_allow_html=True)
    st.markdown("### 📊 Analysis Summary")

    # Metrics row
    col1, col2, col3, col4 = st.columns(4)

    with col1:
        st.metric("Total Rows", f"{summary.get('row_count', 0):,}")

    with col2:
        st.metric("Total Columns", summary.get('column_count', 0))

    with col3:
        issues_count = summary.get('issues_count', 0)
        st.metric("Issues Found", issues_count,
                 delta="Good" if issues_count == 0 else f"-{issues_count}",
                 delta_color="normal" if issues_count == 0 else "inverse")

    with col4:
        missing_total = sum(summary.get('missing_values', {}).values())
        st.metric("Missing Values", missing_total)

    st.markdown("</div>", unsafe_allow_html=True)

    # Issues detail
    if summary.get('issues'):
        st.markdown("<div class='main-card'>", unsafe_allow_html=True)
        st.markdown("### ⚠️ Issues Detected")

        # Group issues by severity
        errors = [i for i in summary['issues'] if i['severity'] == 'error']
        warnings = [i for i in summary['issues'] if i['severity'] == 'warning']

        if errors:
            st.markdown("#### 🔴 Errors")
            for error in errors:
                st.error(f"**{error['type']}**: {error['message']}")
                if 'rows' in error:
                    with st.expander(f"Affected rows ({len(error['rows'])})"):
                        st.write(error['rows'][:10])  # Show first 10 rows

        if warnings:
            st.markdown("#### 🟡 Warnings")
            for warning in warnings:
                st.warning(f"**{warning['type']}**: {warning['message']}")

        st.markdown("</div>", unsafe_allow_html=True)

    # Data preview
    if results.get('preview'):
        st.markdown("<div class='main-card'>", unsafe_allow_html=True)
        st.markdown("### 👁️ Data Preview")

        df_preview = pd.DataFrame(results['preview'])
        st.dataframe(df_preview, use_container_width=True)

        st.markdown("</div>", unsafe_allow_html=True)


def export_section():
    """Export results with error highlighting"""
    if not st.session_state.analysis_complete:
        st.info("💡 Run analysis first to export results")
        return

    st.markdown("<div class='main-card'>", unsafe_allow_html=True)
    st.markdown("### 💾 Export Options")
    st.markdown("Download your data with errors highlighted for easy review")

    results = st.session_state.analysis_results
    summary = results.get("summary", {})

    col1, col2 = st.columns(2)

    with col1:
        st.markdown("#### Excel Export")
        st.markdown("Download Excel file with:")
        st.markdown("- 🔴 Red cells for errors")
        st.markdown("- 🟡 Yellow cells for warnings")
        st.markdown("- 📝 Comments with issue details")

        if st.button("📊 Generate Excel Report", use_container_width=True):
            try:
                import openpyxl
                from openpyxl.styles import PatternFill, Font
                from openpyxl.comments import Comment

                # Create workbook
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Data Analysis"

                # Get dataframe
                if 'data' in summary:
                    df = summary['data']
                else:
                    # Recreate from preview if needed
                    df = pd.DataFrame(results.get('preview', []))

                # Write headers
                for col_idx, col_name in enumerate(df.columns, 1):
                    cell = ws.cell(row=1, column=col_idx, value=col_name)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")

                # Write data with error highlighting
                error_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                warning_fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")

                for row_idx, row_data in df.iterrows():
                    for col_idx, (col_name, value) in enumerate(row_data.items(), 1):
                        cell = ws.cell(row=row_idx + 2, column=col_idx, value=value)

                        # Check if this cell has issues
                        for issue in summary.get('issues', []):
                            if 'rows' in issue and row_idx in issue['rows']:
                                if 'column' in issue and issue['column'] == col_name:
                                    if issue['severity'] == 'error':
                                        cell.fill = error_fill
                                        cell.comment = Comment(issue['message'], "Data Analyzer")
                                    elif issue['severity'] == 'warning':
                                        cell.fill = warning_fill
                                        cell.comment = Comment(issue['message'], "Data Analyzer")

                # Save to bytes
                excel_buffer = io.BytesIO()
                wb.save(excel_buffer)
                excel_buffer.seek(0)

                st.download_button(
                    label="⬇️ Download Excel Report",
                    data=excel_buffer.getvalue(),
                    file_name=f"data_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            except ImportError:
                st.error("Please install openpyxl: pip install openpyxl")

    with col2:
        st.markdown("#### CSV Export")
        st.markdown("Download CSV file with:")
        st.markdown("- 📝 Error markers in cells")
        st.markdown("- 📊 Separate error report")

        if st.button("📄 Generate CSV Report", use_container_width=True):
            # Generate CSV with error markers
            if 'data' in summary:
                df = summary['data'].copy()
            else:
                df = pd.DataFrame(results.get('preview', [])).copy()

            # Add error markers
            for issue in summary.get('issues', []):
                if 'rows' in issue and 'column' in issue:
                    for row_idx in issue['rows']:
                        if row_idx < len(df):
                            current_val = df.at[row_idx, issue['column']]
                            df.at[row_idx, issue['column']] = f"{current_val} [ERROR]"

            csv_buffer = io.StringIO()
            df.to_csv(csv_buffer, index=False)

            st.download_button(
                label="⬇️ Download CSV Report",
                data=csv_buffer.getvalue(),
                file_name=f"data_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                mime="text/csv"
            )

    st.markdown("</div>", unsafe_allow_html=True)


def about_page():
    """About page"""
    st.markdown("<h1>ℹ️ About Data Quality Analyzer</h1>", unsafe_allow_html=True)

    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.markdown("""
        <div class='main-card'>
            <h2 style='text-align: center;'>Professional Data Validation Tool</h2>

            <p style='text-align: justify; margin: 1.5rem 0;'>
            Data Quality Analyzer is a powerful yet simple tool designed to help data professionals
            validate and clean their datasets efficiently. Built with modern technologies and
            best practices, it provides instant feedback on data quality issues.
            </p>

            <h3>🎯 Our Mission</h3>
            <p>
            To make data validation accessible, fast, and reliable for everyone - from data scientists
            to business analysts, from small startups to large enterprises.
            </p>

            <h3>🛠️ Technology Stack</h3>
            <ul>
                <li><strong>Frontend:</strong> Streamlit with custom CSS</li>
                <li><strong>Backend:</strong> Python with Pandas</li>
                <li><strong>Analysis:</strong> MCP (Model Context Protocol)</li>
                <li><strong>Deployment:</strong> Azure Container Apps</li>
            </ul>

            <h3>✨ Features</h3>
            <ul>
                <li>Multi-format support (CSV, JSON, Excel, Parquet)</li>
                <li>Custom validation rules and data dictionaries</li>
                <li>Real-time error detection and highlighting</li>
                <li>Export with visual error markers</li>
                <li>Cloud-ready deployment</li>
            </ul>

            <h3>📞 Contact</h3>
            <p>
            For questions, suggestions, or support, please reach out through our GitHub repository.
            </p>

            <div style='text-align: center; margin-top: 2rem;'>
                <p style='color: #7f8c8d;'>
                    Version 1.0.0 | © 2024 Data Quality Analyzer
                </p>
            </div>
        </div>
        """, unsafe_allow_html=True)


def help_page():
    """Help page"""
    st.markdown("<h1>❓ Help & Documentation</h1>", unsafe_allow_html=True)

    st.markdown("<div class='main-card'>", unsafe_allow_html=True)

    # FAQ section
    st.markdown("### 🤔 Frequently Asked Questions")

    with st.expander("How do I upload a file?"):
        st.markdown("""
        1. Go to the **Analysis** page
        2. Click on the **Upload** tab
        3. Use the file uploader to select your data file
        4. Supported formats: CSV, JSON, Excel (xlsx/xls), Parquet
        """)

    with st.expander("What is a data dictionary?"):
        st.markdown("""
        A data dictionary defines the expected structure of your data:
        - Column names and data types
        - Validation rules (min/max values, allowed values)
        - It helps the analyzer understand what to validate
        """)

    with st.expander("How do I export results?"):
        st.markdown("""
        1. Run your analysis first
        2. Go to the **Export** tab
        3. Choose your preferred format:
           - **Excel**: Includes color-coded error highlighting
           - **CSV**: Includes error markers in cells
        4. Click the download button
        """)

    with st.expander("What do the error colors mean?"):
        st.markdown("""
        - 🔴 **Red**: Critical errors that must be fixed
        - 🟡 **Yellow**: Warnings that should be reviewed
        - 🟢 **Green**: Valid data that passed all checks
        """)

    with st.expander("Can I process large files?"):
        st.markdown("""
        - Files up to 100MB can be processed directly
        - For larger files, consider:
          - Splitting into smaller chunks
          - Using the sampling option
          - Upgrading to our enterprise version
        """)

    st.markdown("### 📚 Quick Start Guide")

    st.markdown("""
    1. **Upload Your Data**
       - Navigate to the Analysis page
       - Upload your data file or use demo data
       - Optionally upload a data dictionary

    2. **Configure Analysis**
       - Set minimum row requirements
       - Choose validation options
       - Configure advanced settings if needed

    3. **Run Analysis**
       - Click "Run Analysis" button
       - Wait for processing to complete
       - Review the progress indicators

    4. **Review Results**
       - Check the summary dashboard
       - Review detected issues
       - Preview your data

    5. **Export Results**
       - Choose export format
       - Download file with error highlighting
       - Share with your team
    """)

    st.markdown("### 🎯 Best Practices")

    col1, col2 = st.columns(2)

    with col1:
        st.markdown("""
        **Data Preparation:**
        - Ensure consistent formatting
        - Use standard date formats
        - Remove special characters from headers
        - Check file encoding (UTF-8 preferred)
        """)

    with col2:
        st.markdown("""
        **Validation Rules:**
        - Start with basic type checking
        - Add range validations for numeric data
        - Use allowed values for categorical data
        - Test with sample data first
        """)

    st.markdown("</div>", unsafe_allow_html=True)

    # Support section
    st.markdown("<div class='main-card'>", unsafe_allow_html=True)
    st.markdown("### 💬 Need More Help?")
    st.markdown("""
    - 📧 Email: support@dataqualityanalyzer.com
    - 📘 Documentation: [docs.dataqualityanalyzer.com](https://docs.dataqualityanalyzer.com)
    - 🐛 Report Issues: [GitHub Issues](https://github.com/dataqualityanalyzer/issues)
    - 💡 Feature Requests: [GitHub Discussions](https://github.com/dataqualityanalyzer/discussions)
    """)
    st.markdown("</div>", unsafe_allow_html=True)


if __name__ == "__main__":
    main()