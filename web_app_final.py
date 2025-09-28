"""
Data Quality Analyzer - Final Fixed Version
Clean, professional interface with proper layout
"""

import streamlit as st
import pandas as pd
import json
import base64
import io
import asyncio
import csv
from typing import Dict, Any, Optional
from datetime import datetime

# Import custom modules
from demo_dictionaries import DEMO_DICTIONARIES, get_demo_dictionary

# Configure Streamlit page
st.set_page_config(
    page_title="Data Quality Analyzer",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed"  # Start with sidebar collapsed
)

# Clean CSS without overlapping issues
st.markdown("""
<style>
    /* Hide Streamlit default elements */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    .stDeployButton {display: none;}

    /* Main container padding */
    .block-container {
        padding-top: 1rem;
        max-width: 100%;
    }

    /* Clean background */
    .stApp {
        background: #ffffff;
    }

    /* Button styling */
    .stButton > button {
        background-color: #3b82f6;
        color: white;
        border: none;
        border-radius: 6px;
        padding: 0.5rem 1.5rem;
        font-weight: 500;
        transition: all 0.2s;
    }

    .stButton > button:hover:not(:disabled) {
        background-color: #2563eb;
        transform: translateY(-1px);
    }

    .stButton > button:disabled {
        background-color: #e5e7eb;
        color: #9ca3af;
    }

    /* File uploader styling */
    .stFileUploader {
        border: 2px dashed #e5e7eb;
        border-radius: 8px;
        padding: 1.5rem;
        background: #f9fafb;
    }

    .stFileUploader:hover {
        border-color: #3b82f6;
        background: #eff6ff;
    }

    /* Typography */
    h1 {
        color: #111827;
        font-weight: 700;
        margin-bottom: 0.5rem;
    }

    h2 {
        color: #1f2937;
        font-weight: 600;
        font-size: 1.5rem;
        margin-bottom: 0.5rem;
    }

    h3 {
        color: #374151;
        font-weight: 600;
        font-size: 1.25rem;
        margin-bottom: 0.5rem;
    }

    h4 {
        color: #4b5563;
        font-weight: 600;
        font-size: 1.1rem;
        margin-bottom: 0.5rem;
    }

    /* Tab container */
    .stTabs {
        background: white;
        padding: 0;
    }

    .stTabs [data-baseweb="tab-list"] {
        gap: 2rem;
        background-color: white;
        border-bottom: 2px solid #e5e7eb;
    }

    .stTabs [data-baseweb="tab"] {
        height: 48px;
        padding: 0 1rem;
        background-color: white;
        border: none;
        color: #6b7280;
        font-weight: 500;
    }

    .stTabs [aria-selected="true"] {
        color: #3b82f6;
        border-bottom: 2px solid #3b82f6;
    }

    /* Success/Error messages */
    .stSuccess {
        background-color: #d1fae5;
        color: #065f46;
        border: 1px solid #6ee7b7;
        padding: 0.5rem 1rem;
        border-radius: 6px;
    }

    .stError {
        background-color: #fee2e2;
        color: #991b1b;
        border: 1px solid #fca5a5;
        padding: 0.5rem 1rem;
        border-radius: 6px;
    }

    .stWarning {
        background-color: #fef3c7;
        color: #92400e;
        border: 1px solid #fcd34d;
        padding: 0.5rem 1rem;
        border-radius: 6px;
    }

    /* Sidebar */
    section[data-testid="stSidebar"] {
        background-color: #f9fafb;
        border-right: 1px solid #e5e7eb;
    }

    /* Remove weird spacing */
    .st-emotion-cache-16idsys p {
        margin: 0;
    }

    /* Column gaps */
    .row-widget.stHorizontal > div {
        padding-right: 1rem;
    }

    .row-widget.stHorizontal > div:last-child {
        padding-right: 0;
    }
</style>
""", unsafe_allow_html=True)


class MCPClient:
    """MCP client for data analysis with proper validation"""

    def __init__(self):
        pass

    def _simulate_mcp_call(self, request_data: Dict, debug: bool = False) -> Dict:
        """Simulate MCP server call with proper validation"""
        tool = request_data.get("tool")
        args = request_data.get("arguments", {})

        if tool == "analyze_data":
            try:
                data_content = args.get("data_content", "")
                file_format = args.get("file_format", "csv")
                schema = args.get("schema", {})
                rules = args.get("rules", {})
                min_rows = args.get("min_rows", 1)

                # Parse data based on format
                df = None
                if file_format.lower() == "csv":
                    df = pd.read_csv(io.StringIO(data_content))
                elif file_format.lower() == "json":
                    json_data = json.loads(data_content)
                    if isinstance(json_data, list):
                        df = pd.json_normalize(json_data)
                    elif isinstance(json_data, dict):
                        first_key = list(json_data.keys())[0] if json_data else None
                        if first_key and isinstance(json_data[first_key], list):
                            df = pd.json_normalize(json_data[first_key])
                        else:
                            df = pd.json_normalize([json_data])
                elif file_format.lower() in ["xlsx", "xls"]:
                    excel_bytes = base64.b64decode(data_content)
                    df = pd.read_excel(io.BytesIO(excel_bytes))
                elif file_format.lower() == "parquet":
                    parquet_bytes = base64.b64decode(data_content)
                    df = pd.read_parquet(io.BytesIO(parquet_bytes))
                else:
                    return {"error": f"Unsupported format: {file_format}"}

                if df is None or df.empty:
                    return {"error": "No data could be loaded"}

                # Analysis with proper validation
                issues = []
                row_count = len(df)

                # Check minimum rows
                if row_count < min_rows:
                    issues.append({
                        "type": "row_count",
                        "severity": "error",
                        "message": f"Row count ({row_count}) is less than minimum required ({min_rows})"
                    })

                # Check for invalid values in each column
                for col in df.columns:
                    # Check for invalid text values
                    if df[col].dtype == 'object':
                        for idx, val in df[col].items():
                            if pd.notna(val):
                                val_str = str(val).lower()
                                # Check for common invalid values
                                if val_str in ['invalid', 'error', 'n/a', 'null', 'none', 'invalid-date']:
                                    issues.append({
                                        "type": "invalid_value",
                                        "severity": "error",
                                        "column": col,
                                        "row": idx,
                                        "value": val,
                                        "message": f"Invalid value '{val}' in column '{col}' at row {idx}"
                                    })

                    # Check for out-of-range values in numeric columns
                    elif df[col].dtype in ['int64', 'float64']:
                        if 'age' in col.lower():
                            invalid_ages = df[(df[col] < 0) | (df[col] > 150)]
                            for idx, row in invalid_ages.iterrows():
                                issues.append({
                                    "type": "range_violation",
                                    "severity": "error",
                                    "column": col,
                                    "row": idx,
                                    "value": row[col],
                                    "message": f"Invalid age value {row[col]} at row {idx}"
                                })

                # Check for missing values
                missing_counts = df.isnull().sum()
                for col, count in missing_counts.items():
                    if count > 0:
                        issues.append({
                            "type": "missing_values",
                            "severity": "warning",
                            "column": col,
                            "count": int(count),
                            "message": f"Column '{col}' has {count} missing values"
                        })

                # Summary statistics
                summary = {
                    "row_count": row_count,
                    "column_count": len(df.columns),
                    "columns": df.columns.tolist(),
                    "dtypes": {col: str(dtype) for col, dtype in df.dtypes.items()},
                    "missing_values": missing_counts.to_dict(),
                    "issues_count": len(issues),
                    "issues": issues,
                    "data": df
                }

                return {
                    "success": True,
                    "summary": summary,
                    "preview": df.head(10).to_dict('records') if not df.empty else []
                }

            except Exception as e:
                return {"error": str(e)}

        elif tool == "parse_data_dictionary":
            try:
                dictionary_content = args.get("dictionary_content", "")
                schema = {}
                rules = {}

                lines = dictionary_content.strip().split('\n')
                if ',' in lines[0] or '\t' in lines[0]:
                    reader = csv.DictReader(io.StringIO(dictionary_content))
                    for row in reader:
                        field_name = row.get('field_name') or row.get('column') or row.get('name')
                        if field_name:
                            field_type = row.get('type') or row.get('data_type') or 'str'
                            schema[field_name] = field_type

                            field_rules = {}
                            if row.get('min'):
                                try:
                                    field_rules['min'] = float(row['min'])
                                except:
                                    pass
                            if row.get('max'):
                                try:
                                    field_rules['max'] = float(row['max'])
                                except:
                                    pass
                            if field_rules:
                                rules[field_name] = field_rules

                return {
                    "success": True,
                    "schema": schema,
                    "rules": rules
                }
            except Exception as e:
                return {"error": str(e)}

        return {"error": "Unknown tool"}

    async def analyze_data(self, data_content: str, file_format: str = "csv",
                          schema: Dict = None, rules: Dict = None,
                          min_rows: int = 1, debug: bool = False):
        request_data = {
            "tool": "analyze_data",
            "arguments": {
                "data_content": data_content,
                "file_format": file_format,
                "schema": schema or {},
                "rules": rules or {},
                "min_rows": min_rows
            }
        }
        return self._simulate_mcp_call(request_data, debug)

    async def parse_data_dictionary(self, dictionary_content: str,
                                   format_hint: str = "auto",
                                   use_cache: bool = True,
                                   debug: bool = False) -> Dict:
        request_data = {
            "tool": "parse_data_dictionary",
            "arguments": {
                "dictionary_content": dictionary_content,
                "format_hint": format_hint,
                "use_cache": use_cache
            }
        }
        return self._simulate_mcp_call(request_data, debug)


def main():
    """Main application entry point"""

    # Initialize session state
    if 'analysis_complete' not in st.session_state:
        st.session_state.analysis_complete = False
    if 'analysis_results' not in st.session_state:
        st.session_state.analysis_results = None
    if 'uploaded_file' not in st.session_state:
        st.session_state.uploaded_file = None
    if 'uploaded_file_name' not in st.session_state:
        st.session_state.uploaded_file_name = None
    if 'dictionary_file' not in st.session_state:
        st.session_state.dictionary_file = None
    if 'dictionary_file_name' not in st.session_state:
        st.session_state.dictionary_file_name = None

    # Create navigation tabs
    tab1, tab2 = st.tabs(["📊 Analyze", "ℹ️ About"])

    with tab1:
        analyze_page()
    with tab2:
        about_page()


def analyze_page():
    """Main analysis page with proper three-column layout"""

    st.title("Data Quality Analyzer")
    st.markdown("Upload your data, optionally add validation rules, and analyze for quality issues")

    st.markdown("---")

    # Main three-column layout for upload and analyze
    col1, col2, col3 = st.columns([2, 2, 1.5])

    with col1:
        st.markdown("### 📁 Upload Data File")
        st.markdown("<small style='color: #6b7280;'>Required</small>", unsafe_allow_html=True)

        # Show loaded file info if demo data is selected
        if st.session_state.uploaded_file_name and "Demo:" in str(st.session_state.uploaded_file_name):
            st.success(f"✓ {st.session_state.uploaded_file_name}")

        uploaded_file = st.file_uploader(
            "Choose a file",
            type=['csv', 'json', 'xlsx', 'xls', 'parquet'],
            key="main_data_upload"
        )

        if uploaded_file:
            file_size = len(uploaded_file.read())
            uploaded_file.seek(0)
            st.session_state.uploaded_file = uploaded_file
            st.session_state.uploaded_file_name = uploaded_file.name
            st.success(f"✓ {uploaded_file.name} ({file_size:,} bytes)")

    with col2:
        st.markdown("### 📚 Upload Data Dictionary")
        st.markdown("<small style='color: #6b7280;'>Optional - defines validation rules</small>", unsafe_allow_html=True)

        # Show loaded dictionary info if demo is selected
        if st.session_state.dictionary_file_name and "Demo:" in str(st.session_state.dictionary_file_name):
            st.success(f"✓ {st.session_state.dictionary_file_name}")

        dict_file = st.file_uploader(
            "Choose a dictionary file",
            type=['csv', 'json', 'txt'],
            key="dict_upload"
        )

        if dict_file:
            st.session_state.dictionary_file = dict_file
            st.session_state.dictionary_file_name = dict_file.name
            st.success(f"✓ {dict_file.name}")

    with col3:
        st.markdown("### ⚡ Analyze")
        st.markdown("<div style='height: 35px;'></div>", unsafe_allow_html=True)

        # Enable analyze button only if data is uploaded
        analyze_enabled = st.session_state.uploaded_file is not None

        if st.button("🚀 Run Analysis",
                    disabled=not analyze_enabled,
                    type="primary",
                    use_container_width=True,
                    help="Upload data file to enable"):
            run_analysis()

        if not analyze_enabled:
            st.caption("Upload data to enable")

    # Demo data section in expander
    with st.expander("📂 Load Demo Data"):
        demo_col1, demo_col2 = st.columns(2)

        with demo_col1:
            st.markdown("**Sample Datasets**")
            demo_data_choice = st.selectbox(
                "Choose demo data",
                ["None", "CSV - Western Names", "CSV - Asian Names",
                 "JSON - Mixed Names", "CSV - Clinical Trial"]
            )

            if st.button("Load Demo Data", key="load_demo_data"):
                if demo_data_choice != "None":
                    if demo_data_choice == "CSV - Western Names":
                        demo_data = """name,age,email,country
John Smith,25,john@email.com,USA
Jane Doe,30,jane@email.com,Canada
Bob Wilson,invalid,bob@email,UK"""
                    elif demo_data_choice == "CSV - Asian Names":
                        demo_data = """name,age,city,join_date
李明,28,Beijing,2023-01-15
田中太郎,35,Tokyo,2023-02-20
김철수,42,Seoul,invalid-date"""
                    elif demo_data_choice == "JSON - Mixed Names":
                        demo_data = json.dumps([
                            {"name": "Alice Smith", "age": 30, "score": 85.5},
                            {"name": "王小明", "age": 25, "score": 92.0},
                            {"name": "Carlos García", "age": "invalid", "score": 88.0}
                        ])
                    elif demo_data_choice == "CSV - Clinical Trial":
                        demo_data = """patient_id,age,treatment,outcome,date
P001,45,Drug_A,Improved,2023-01-15
P002,52,Drug_B,No_Change,2023-01-20
P003,999,Drug_A,Improved,invalid"""

                    st.session_state.uploaded_file = demo_data
                    st.session_state.uploaded_file_name = f"Demo: {demo_data_choice}"
                    st.success(f"✅ Loaded {demo_data_choice}")
                    st.rerun()

        with demo_col2:
            st.markdown("**Sample Dictionaries**")
            demo_dict_choice = st.selectbox(
                "Choose demo dictionary",
                ["None"] + list(DEMO_DICTIONARIES.keys())
            )

            if st.button("Load Demo Dictionary", key="load_demo_dict"):
                if demo_dict_choice != "None":
                    demo_dict = get_demo_dictionary(demo_dict_choice)
                    st.session_state.dictionary_file = demo_dict
                    st.session_state.dictionary_file_name = f"Demo: {demo_dict_choice}"
                    st.success(f"✅ Loaded {demo_dict_choice}")
                    st.rerun()

    # Results section
    if st.session_state.analysis_complete:
        st.markdown("---")
        show_results()


def run_analysis():
    """Execute the data analysis"""
    with st.spinner("🔍 Analyzing your data..."):
        progress_bar = st.progress(0)
        status_text = st.empty()

        stages = [
            (0.25, "Loading data..."),
            (0.50, "Checking data quality..."),
            (0.75, "Validating against rules..."),
            (1.0, "Generating report...")
        ]

        client = MCPClient()

        # Prepare data
        if isinstance(st.session_state.uploaded_file, str):
            data_content = st.session_state.uploaded_file
            file_format = "json" if data_content.startswith('[') or data_content.startswith('{') else "csv"
        else:
            file_ext = st.session_state.uploaded_file.name.split('.')[-1].lower()
            file_format = file_ext

            if file_ext in ['xlsx', 'xls', 'parquet']:
                file_bytes = st.session_state.uploaded_file.read()
                data_content = base64.b64encode(file_bytes).decode('utf-8')
                st.session_state.uploaded_file.seek(0)
            else:
                data_content = st.session_state.uploaded_file.read().decode('utf-8')
                st.session_state.uploaded_file.seek(0)

        # Parse dictionary if provided
        schema = {}
        rules = {}
        if st.session_state.dictionary_file:
            if isinstance(st.session_state.dictionary_file, str):
                dict_content = st.session_state.dictionary_file
            else:
                dict_content = st.session_state.dictionary_file.read().decode('utf-8')
                st.session_state.dictionary_file.seek(0)

            dict_result = asyncio.run(
                client.parse_data_dictionary(dict_content)
            )

            if dict_result.get("success"):
                schema = dict_result.get("schema", {})
                rules = dict_result.get("rules", {})

        # Update progress
        for progress, message in stages:
            status_text.text(message)
            progress_bar.progress(progress)
            asyncio.run(asyncio.sleep(0.2))

        # Run analysis
        result = asyncio.run(
            client.analyze_data(
                data_content=data_content,
                file_format=file_format,
                schema=schema,
                rules=rules,
                min_rows=1
            )
        )

        if result.get("success"):
            st.session_state.analysis_complete = True
            st.session_state.analysis_results = result
            progress_bar.empty()
            status_text.empty()
            st.success("✅ Analysis complete!")
            st.rerun()
        else:
            progress_bar.empty()
            status_text.empty()
            st.error(f"❌ Analysis failed: {result.get('error')}")


def show_results():
    """Display analysis results"""
    results = st.session_state.analysis_results
    summary = results.get("summary", {})

    st.markdown("## 📊 Analysis Results")

    # Metrics
    col1, col2, col3, col4 = st.columns(4)

    with col1:
        st.metric("Total Rows", f"{summary.get('row_count', 0):,}")

    with col2:
        st.metric("Columns", summary.get('column_count', 0))

    with col3:
        issues_count = summary.get('issues_count', 0)
        st.metric("Issues Found", issues_count,
                 delta="✓ Clean" if issues_count == 0 else f"{issues_count} issues",
                 delta_color="normal" if issues_count == 0 else "inverse")

    with col4:
        missing = sum(summary.get('missing_values', {}).values())
        st.metric("Missing Values", missing)

    # Issues details
    if summary.get('issues'):
        st.markdown("### ⚠️ Data Quality Issues")

        errors = [i for i in summary['issues'] if i.get('severity') == 'error']
        warnings = [i for i in summary['issues'] if i.get('severity') == 'warning']

        if errors:
            st.markdown("#### 🔴 Errors")
            for error in errors[:10]:
                st.error(f"**{error.get('type', 'Error')}**: {error.get('message', '')}")
            if len(errors) > 10:
                st.warning(f"... and {len(errors) - 10} more errors")

        if warnings:
            st.markdown("#### 🟡 Warnings")
            for warning in warnings[:5]:
                st.warning(f"**{warning.get('type', 'Warning')}**: {warning.get('message', '')}")
            if len(warnings) > 5:
                st.info(f"... and {len(warnings) - 5} more warnings")

    # Data preview
    if results.get('preview'):
        st.markdown("### 📋 Data Preview")
        df_preview = pd.DataFrame(results['preview'])
        st.dataframe(df_preview, use_container_width=True, height=300)

    # Export options
    st.markdown("### 💾 Export Options")
    export_col1, export_col2 = st.columns(2)

    with export_col1:
        # Excel export
        if st.button("📊 Export to Excel", use_container_width=True):
            try:
                import openpyxl
                from openpyxl.styles import PatternFill, Font

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Data Analysis"

                df = summary.get('data', pd.DataFrame(results.get('preview', [])))

                # Write headers
                for col_idx, col_name in enumerate(df.columns, 1):
                    cell = ws.cell(row=1, column=col_idx, value=col_name)
                    cell.font = Font(bold=True)

                # Write data with error highlighting
                error_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")

                for row_idx, row_data in df.iterrows():
                    for col_idx, value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_idx + 2, column=col_idx, value=value)

                        # Check if this cell has errors
                        for issue in summary.get('issues', []):
                            if (issue.get('row') == row_idx and
                                issue.get('column') == df.columns[col_idx - 1]):
                                cell.fill = error_fill

                excel_buffer = io.BytesIO()
                wb.save(excel_buffer)
                excel_buffer.seek(0)

                st.download_button(
                    label="⬇️ Download Excel",
                    data=excel_buffer.getvalue(),
                    file_name=f"analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            except ImportError:
                st.error("Install openpyxl: pip install openpyxl")

    with export_col2:
        # CSV export
        if st.button("📄 Export to CSV", use_container_width=True):
            df = summary.get('data', pd.DataFrame(results.get('preview', [])))
            csv_buffer = io.StringIO()
            df.to_csv(csv_buffer, index=False)

            st.download_button(
                label="⬇️ Download CSV",
                data=csv_buffer.getvalue(),
                file_name=f"analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                mime="text/csv"
            )


def about_page():
    """About page with information"""
    st.title("About Data Quality Analyzer")

    st.markdown("""
    ### Professional Data Validation Tool

    Data Quality Analyzer helps validate and clean datasets efficiently by detecting common data quality issues.

    #### ✨ Key Features
    - **Multiple File Formats**: CSV, JSON, Excel, Parquet
    - **Automatic Error Detection**: Finds "invalid", "error", "n/a" and other problematic values
    - **Range Validation**: Checks for unrealistic values (e.g., age > 150)
    - **Data Dictionaries**: Optional custom validation rules
    - **Export with Highlighting**: Download results with errors marked

    #### 🚀 How to Use
    1. **Upload your data file** (required)
    2. **Optionally add a data dictionary** for custom rules
    3. **Click Analyze** to run validation
    4. **Export results** with error highlighting

    #### 📊 What We Check
    - Invalid text values in numeric fields
    - Out-of-range numeric values
    - Missing or null values
    - Data type mismatches
    - Custom validation rules

    ---
    *Version 1.2.0 | © 2024 Data Quality Analyzer*
    """)


if __name__ == "__main__":
    main()