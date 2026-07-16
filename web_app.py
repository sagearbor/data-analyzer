"""
Data Quality Analyzer - Clean UI with Proper Layout
Fixed navbar, three-column layout, no sidebar issues
"""

import streamlit as st
import pandas as pd
import json
import base64
import io
import asyncio
import csv
import requests
from typing import Dict, Any, Optional
from datetime import datetime
import matplotlib.pyplot as plt
import matplotlib.patches as patches
import numpy as np
import pypdf
import re
import plotly.graph_objects as go
import plotly.express as px
import hashlib
import os
import tempfile
from pathlib import Path

# Import custom modules
from demo_dictionaries import DEMO_DICTIONARIES, get_demo_dictionary
from src.__version__ import __version__
# NOTE: QualityPipeline/QualityChecker (mcp_server.py) are no longer imported
# here. Rule-based quality analysis now runs centrally behind the REST API
# (POST /api/v1/analyze) so the UI, the API, and the MCP server all share one
# running backend instead of the UI running the engine in-process. See
# DataQualityAnalyzer below.
#
# DataLoader IS still imported and used locally for parsing/previewing
# uploaded Excel (.xlsx/.xls) files client-side (e.g. sheet peeking, dictionary
# ingestion) before data is handed off to DataQualityAnalyzer, which sends it
# to the API for analysis.
from mcp_server import DataLoader
# Force use of custom renderer for better compatibility
MERMAID_AVAILABLE = False
from mermaid_renderer import render_mermaid

# Import LLM parser
try:
    from src.llm_client import LLMDictionaryParser, get_available_deployments
    LLM_AVAILABLE = True
except ImportError:
    LLM_AVAILABLE = False
    print("LLM client not available. Install openai package to enable.")

# Import Named Program Cache and Logic Validation Engine
try:
    from src.program_manager import ProgramManager
    from src.logic_engine import LogicValidator, RuleExtractor, ConditionalRule
    PROGRAM_CACHE_AVAILABLE = True
except ImportError:
    PROGRAM_CACHE_AVAILABLE = False
    print("Program cache and logic validation not available.")

# Import environment banner
try:
    import envbanner
    ENVBANNER_AVAILABLE = True
except ImportError:
    ENVBANNER_AVAILABLE = False
    print("env-banner not available. Install with: pip install env-banner")

# Browser console logging helper
def log_to_browser_console(message: str, data: dict = None):
    """Inject JavaScript to log to browser console (visible in Chrome DevTools)"""
    import streamlit.components.v1 as components
    import json
    log_data = json.dumps(data) if data else "{}"
    html = f"""
    <script>
        console.log('[Data Analyzer] {message}', {log_data});
    </script>
    """
    components.html(html, height=0, width=0)


def resolve_effective_dictionary(
    demo_dict_selection: str,
    demo_dictionary_built: Optional[dict],
    dict_file_uploaded: bool,
    previous_session_dictionary: Optional[dict],
) -> Optional[dict]:
    """Decide what `st.session_state.dictionary` should be for this render pass.

    Pure function (no Streamlit calls, no CSV parsing) so it can be unit
    tested without a running Streamlit runtime. Centralizes the "which
    dictionary wins" decision that used to be implicit in the demo-dictionary
    selectbox's `if demo_dict != "None":` branch having no `else`. That
    missing `else` was the root cause of the reported bug: selecting a demo
    dictionary, analyzing, then resetting the selector back to "None" left
    the previously-loaded demo dictionary's rules sitting in
    `st.session_state.dictionary` forever, silently driving subsequent
    "Analyze Data Quality" runs.

    Precedence, evaluated in order:
    1. If a dictionary file is currently present in the file uploader
       (`dict_file_uploaded`), the uploaded dictionary wins and is preserved
       as-is (`previous_session_dictionary`) - the upload-handling code
       earlier in the same render pass is the one that actually parses the
       file and assigns `st.session_state.dictionary`, so by the time this
       function runs, `previous_session_dictionary` already reflects that
       upload. This prevents a leftover "None" demo-selector value (e.g.
       from before a file was uploaded) from clobbering an active upload.
    2. Else, if a demo dictionary is selected (`demo_dict_selection` is not
       "None"/falsy), use the freshly-built dict for that selection
       (`demo_dictionary_built`). Because this is the dict the caller just
       built fresh from the CURRENTLY selected demo entry (not whatever was
       previously in session state), switching from one demo dictionary to
       another fully replaces the old one - there is no merge/append with
       the prior selection's rules.
    3. Else (no upload, demo selector is "None"): return `None`, clearing
       any stale dictionary left over from a prior demo selection or a
       dictionary file that has since been removed from the uploader.

    Args:
        demo_dict_selection: current value of the "Or load demo dictionary:"
            selectbox, e.g. "None" or a key from DEMO_DICTIONARIES.
        demo_dictionary_built: the `{"source", "filename", "rules"}` dict
            freshly parsed this render pass for `demo_dict_selection`, or
            None if `demo_dict_selection == "None"`.
        dict_file_uploaded: True if a dictionary file is currently present
            in the dictionary file uploader widget this render pass (i.e.
            `dict_file is not None`).
        previous_session_dictionary: `st.session_state.dictionary` as it was
            before this render pass's decision. Only consulted when
            `dict_file_uploaded` is True, to preserve what the upload
            handling code already assigned.

    Returns:
        The dict that `st.session_state.dictionary` should be set to, or
        None if no dictionary should be active.
    """
    if dict_file_uploaded:
        return previous_session_dictionary

    if demo_dict_selection and demo_dict_selection != "None":
        return demo_dictionary_built

    return None


# Configure Streamlit page
st.set_page_config(
    page_title="Data Quality Analyzer",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed"  # Keep sidebar collapsed by default
)

# Initialize environment banner (bottom position to avoid navbar conflict)
if ENVBANNER_AVAILABLE:
    # Get environment and create custom message with PHI warning
    import os
    app_env = os.getenv("APP_ENV", "dev").lower()
    # Only show banner in dev/staging, hide in production
    if app_env != "prod":
        banner_text = f"{app_env.upper()} - do NOT use real data or files with PHI."
        envbanner.streamlit(position="bottom", opacity=0.9, text=banner_text)

# --- Design system -----------------------------------------------------
#
# Palette (single accent, navy anchor):
#   Navy (navbar / dark text):     #1e293b
#   Navy, deeper (dark mode bg):   #0f172a
#   Slate borders/surfaces:        #e2e8f0 / #334155
#   Accent (primary actions):      #2563eb   (hover/active: #1d4ed8)
#   Accent, soft surfaces:         #eff6ff / #3b82f6 (used sparingly - focus
#                                   rings, links, chart highlight - not a
#                                   second competing hue)
#   Body text:                     #1e293b (light) / #e2e8f0 (dark)
#   Muted text:                    #64748b (light) / #94a3b8 (dark)
#
# NAVBAR APPROACH (version-robust, and actually contains the nav):
# Older versions of this file rendered the title as raw injected HTML
# (`.app-title-bar`, position:fixed) sitting ABOVE st.tabs, with a *separate*
# st.toggle placed in normal document flow just below it - so the "navbar"
# never actually contained navigation or the toggle, just the title.
#
# This version uses `st.container(key="navbar")`, which Streamlit guarantees
# renders a DOM node carrying a stable, public CSS class `.st-key-navbar`
# (documented, public `key=` behavior - unlike `data-baseweb` internals,
# which are private implementation detail and have broken across Streamlit
# versions before, e.g. 1.28 -> 1.59). Real Streamlit widgets (title
# markdown, in-bar navigation, dark-mode toggle) are mounted inside that
# container via ordinary st.* calls - widgets cannot be mounted inside raw
# `unsafe_allow_html` markup, only inside a real container - and the
# container itself is pinned with position:fixed. Navigation inside the bar
# uses st.segmented_control (its own public testid root,
# [data-testid="stSegmentedControl"], not a data-baseweb internal) instead of
# st.tabs, because st.tabs cannot be relocated into an arbitrary container.
st.markdown("""
<style>
    /* Hide Streamlit default elements */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}
    .stDeployButton {display: none;}

    /* Hide sidebar by default */
    section[data-testid="stSidebar"] {
        display: none !important;
    }

    html, body, [class*="css"] {
        font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Helvetica, Arial, sans-serif;
    }

    /* Leave room for the fixed navbar below; generous, calm spacing */
    .block-container {
        padding-top: 4.75rem !important;
        padding-bottom: 3rem !important;
        max-width: 100% !important;
    }

    .stApp {
        background: #ffffff;
    }

    h1, h2, h3, h4 {
        color: #1e293b;
        font-weight: 600;
        letter-spacing: -0.01em;
    }
    h3 {
        white-space: nowrap;
        overflow: hidden;
        text-overflow: ellipsis;
    }
    p, span, label, li {
        color: #334155;
    }

    /* ---- Fixed navbar: st.container(key="navbar") -> .st-key-navbar ---- */
    .st-key-navbar {
        position: fixed;
        top: 0;
        left: 0;
        right: 0;
        z-index: 1000;
        background-color: #1e293b;
        padding: 0.65rem 1.5rem;
        box-shadow: 0 1px 3px rgba(0,0,0,0.15);
    }

    /* Title + version subscript (plain markdown mounted inside the bar) */
    .app-title {
        color: #f1f5f9;
        font-size: 1.15rem;
        font-weight: 600;
        line-height: 2.2rem;
        white-space: nowrap;
    }
    .app-version {
        color: #94a3b8;
        font-size: 0.68rem;
        font-weight: 500;
        margin-left: 0.4rem;
        vertical-align: sub;
    }

    /* Segmented-control navigation, restyled as light pills on the dark bar.
       Selector targets the widget's own public testid root only - no
       data-baseweb descendant selectors. */
    .st-key-navbar [data-testid="stSegmentedControl"] {
        display: flex;
        justify-content: center;
    }
    .st-key-navbar [data-testid="stSegmentedControl"] label {
        background-color: transparent !important;
        color: #cbd5e1 !important;
        border: none !important;
        font-weight: 500;
    }
    .st-key-navbar [data-testid="stSegmentedControl"] label:hover {
        background-color: #334155 !important;
    }
    .st-key-navbar [data-testid="stSegmentedControl"] label[aria-checked="true"] {
        background-color: #2563eb !important;
        color: #ffffff !important;
    }

    /* Dark-mode toggle mounted in the bar */
    .st-key-navbar [data-testid="stToggle"] label p {
        color: #cbd5e1 !important;
        font-size: 0.85rem;
    }
    .st-key-navbar div[data-testid="stVerticalBlock"] {
        gap: 0 !important;
    }
    /* Keep every direct widget wrapper inside the bar vertically centered
       and free of the default block spacing so title/nav/toggle line up. */
    .st-key-navbar [data-testid="stVerticalBlockBorderWrapper"],
    .st-key-navbar [data-testid="element-container"] {
        display: flex;
        align-items: center;
    }

    /* ---- Cards: shared surface treatment for upload / dictionary / results ----
       Applied via st.container(key="card-...") -> .st-key-card-... */
    [class*="st-key-card-"] {
        background: #f8fafc;
        border: 1px solid #e2e8f0;
        border-radius: 10px;
        padding: 1.25rem 1.5rem 1.5rem 1.5rem;
        margin-bottom: 0.75rem;
    }
    .app-section-title {
        margin-top: 0 !important;
        font-size: 1rem;
        font-weight: 600;
        color: #1e293b;
        margin-bottom: 0.75rem;
    }

    [data-testid="stFileUploaderDropzone"] {
        background: #ffffff;
        border: 1.5px dashed #cbd5e1;
        border-radius: 8px;
        transition: border-color 0.15s ease, background-color 0.15s ease;
    }
    [data-testid="stFileUploaderDropzone"]:hover {
        border-color: #2563eb;
        background: #eff6ff;
    }

    /* Primary button: well-proportioned, restrained (native button shape,
       one accent color, no gradient/glow). */
    .stButton > button[kind="primary"] {
        background-color: #2563eb;
        border: none;
        font-weight: 600;
        transition: background-color 0.15s ease;
    }
    .stButton > button[kind="primary"]:hover {
        background-color: #1d4ed8;
    }
    .stButton > button[kind="secondary"] {
        font-weight: 500;
    }

    /* Metrics: quiet surface, accent reserved for the number itself */
    [data-testid="stMetric"] {
        background: #f8fafc;
        border: 1px solid #e2e8f0;
        padding: 0.9rem 1rem;
        border-radius: 8px;
    }
    [data-testid="stMetricValue"] {
        color: #1e293b;
    }

    [data-testid="stExpander"] {
        border: 1px solid #e2e8f0;
        border-radius: 8px;
    }

    .dataframe {
        font-size: 0.9rem;
    }
    .element-container {
        margin-top: 0.3rem;
    }
</style>
""", unsafe_allow_html=True)

# --- Navbar: title + in-bar navigation + dark-mode toggle, all inside one
# real st.container(key=...) so they are visually and structurally contained
# by the fixed bar (not stacked in normal flow beneath it). ------------------
if "dark_mode" not in st.session_state:
    st.session_state.dark_mode = False

with st.container(key="navbar"):
    _nav_title_col, _nav_links_col, _nav_toggle_col = st.columns([2.2, 2.5, 1])

    with _nav_title_col:
        st.markdown(
            f'<span class="app-title">Data Quality Analyzer'
            f'<sub class="app-version">v{__version__}</sub></span>',
            unsafe_allow_html=True,
        )

    with _nav_links_col:
        nav_choice = st.segmented_control(
            "Navigation",
            ["Analyze", "About"],
            default="Analyze",
            required=True,
            key="nav_choice",
            label_visibility="collapsed",
        )

    with _nav_toggle_col:
        st.toggle("🌙", key="dark_mode", help="Toggle dark mode")

# segmented_control returns None if a user deselects the active pill by
# clicking it again - treat that the same as staying on "Analyze" so the app
# never renders a blank page.
nav_choice = nav_choice or "Analyze"

# --- Dark mode -----------------------------------------------------------
# Coherent restyle (not a bolted-on override): same card/spacing/radius
# system as light mode, just remapped color tokens. Persisted via
# st.session_state so switching pages or re-running the script doesn't reset
# it, and toggling itself never touches st.session_state.data/dictionary/
# analysis_results, so uploaded data and results survive a theme change.
if st.session_state.dark_mode:
    st.markdown("""
    <style>
        .stApp, .block-container {
            background: #0f172a !important;
        }
        .stApp, .block-container, p, span, label, li,
        div[data-testid="stMarkdownContainer"] {
            color: #e2e8f0 !important;
        }
        h1, h2, h3, h4, h5, h6 {
            color: #f1f5f9 !important;
        }
        [class*="st-key-card-"] {
            background: #1e293b !important;
            border-color: #334155 !important;
        }
        .app-section-title {
            color: #f1f5f9 !important;
        }
        [data-testid="stExpander"] {
            background: #1e293b !important;
            border: 1px solid #334155 !important;
        }
        [data-testid="stExpander"] summary {
            background: #1e293b !important;
            color: #e2e8f0 !important;
        }
        input, textarea, select,
        div[data-baseweb="select"] > div,
        div[data-baseweb="input"] > div {
            background-color: #1e293b !important;
            color: #e2e8f0 !important;
            border-color: #334155 !important;
        }
        [data-testid="stFileUploaderDropzone"] {
            background: #1e293b !important;
            border-color: #334155 !important;
        }
        [data-testid="stFileUploaderDropzone"]:hover {
            border-color: #3b82f6 !important;
            background: #1e2a3f !important;
        }
        pre, code {
            background-color: #1e293b !important;
            color: #e2e8f0 !important;
        }
        [data-testid="stMetric"] {
            background: #1e293b !important;
            border: 1px solid #334155 !important;
        }
        [data-testid="stMetricValue"] {
            color: #f1f5f9 !important;
        }
        [data-testid="stDataFrame"] {
            background: #1e293b !important;
        }
        .stButton > button[kind="secondary"] {
            background-color: #1e293b !important;
            color: #e2e8f0 !important;
            border-color: #334155 !important;
        }
    </style>
    """, unsafe_allow_html=True)
# --- End design system -----------------------------------------------------

class DataQualityAnalyzer:
    """
    HTTP client for the centralized data-analyzer REST API.

    Historically this class ran mcp_server.QualityPipeline in-process. It now
    POSTs to the API's POST /api/v1/analyze endpoint instead, so there is a
    single running backend for rule-based quality checks shared by the API,
    the MCP server, and this UI (Option A: centralize the runtime).

    Conditional logic validation (extracted from a dictionary's 'fields' via
    RuleExtractor/LogicValidator) is NOT part of /api/v1/analyze - that
    endpoint is intentionally rule-based only (schema/range/allowed-value
    checks, no LLM). Logic validation stays local here and its violations are
    merged into the API's issues/summary after the HTTP call returns.

    Configuration (environment variables):
        DATA_ANALYZER_API_URL: Base URL of the data-analyzer REST API.
            Defaults to "http://localhost:8000" for local dev (matches
            api_server.py's default API_PORT). For deployment, set this to
            the API Container App's internal DNS FQDN so UI -> API traffic
            stays inside the Azure Container Apps environment's VNET rather
            than round-tripping over the public internet.
        DATA_ANALYZER_API_KEY: Value sent as the X-API-Key header. Must match
            the API app's own DATA_ANALYZER_API_KEY.
    """

    def __init__(self):
        self.api_base_url = os.getenv("DATA_ANALYZER_API_URL", "http://localhost:8000").rstrip("/")
        self.api_key = os.getenv("DATA_ANALYZER_API_KEY", "")

    async def analyze_data_quality(self, data: pd.DataFrame, dictionary: Optional[Dict] = None, source_format: Optional[str] = None) -> Dict[str, Any]:
        """
        Run data quality analysis via the centralized REST API.

        Args:
            data: DataFrame to analyze
            dictionary: Optional dictionary with validation rules and schema.
                       Can have 'rules' and/or 'schema' keys, or direct field definitions.
            source_format: Optional hint for how to serialize `data` for the
                       HTTP upload. One of "xlsx", "xls", or None (default:
                       CSV). When the original upload was an Excel file, pass
                       "xlsx"/"xls" so the API receives real Excel bytes with
                       the correct content-type instead of a CSV re-encoding.

        Returns:
            Dict with summary, issues, recommendations, quality_checks, and
            summary_stats - the exact shape the rest of this module (the
            dashboard rendering code) has always consumed.

        Raises:
            RuntimeError: if the API call fails (non-200 response, connection
                error, or timeout). Callers should catch this and surface it
                via st.error(...) rather than letting the UI crash.
        """
        # Parse dictionary to extract schema and rules to send to the API
        # (unchanged from the old in-process code path - only the destination
        # of schema/rules changed, from a local QualityPipeline call to an
        # HTTP request).
        schema = None
        rules = None

        if dictionary and isinstance(dictionary, dict):
            # Handle different dictionary formats
            if 'rules' in dictionary:
                dict_rules = dictionary['rules']
                # Convert web_app dictionary format to QualityChecker format
                schema = {}
                rules = {}
                for field_name, field_spec in dict_rules.items():
                    if isinstance(field_spec, dict):
                        # Extract type for schema
                        if 'type' in field_spec:
                            schema[field_name] = field_spec['type']

                        # Extract validation rules
                        field_rules = {}
                        if 'min' in field_spec and pd.notna(field_spec['min']):
                            try:
                                field_rules['min'] = float(field_spec['min'])
                            except (ValueError, TypeError):
                                pass
                        if 'max' in field_spec and pd.notna(field_spec['max']):
                            try:
                                field_rules['max'] = float(field_spec['max'])
                            except (ValueError, TypeError):
                                pass
                        if 'allowed_values' in field_spec and field_spec['allowed_values']:
                            field_rules['allowed'] = field_spec['allowed_values']

                        if field_rules:
                            rules[field_name] = field_rules

            elif 'schema' in dictionary:
                schema = dictionary.get('schema')
                rules = dictionary.get('validation_rules', {})

        # Run the rule-based engine via the API (replaces the old in-process
        # QualityPipeline(data, schema, rules).run_all_checks() call). The
        # API already returns issues/summary/recommendations/quality_checks/
        # summary_stats in this exact shape - no reshaping needed here.
        result = self._call_analyze_api(data, schema, rules, source_format=source_format)

        issues = result["issues"]
        summary = result["summary"]
        recommendations = result["recommendations"]

        # LOGIC VALIDATION - Run conditional logic checks if available.
        # This is separate from the rule-based REST engine (it's driven by
        # dictionary 'fields' extracted via RuleExtractor, not schema/rules),
        # so it stays local and its results are merged in after the API call.
        logic_violations_count = 0
        if PROGRAM_CACHE_AVAILABLE and dictionary and 'fields' in dictionary:
            try:
                # Extract conditional rules from dictionary fields
                extractor = RuleExtractor()
                conditional_rules = extractor.extract_rules_from_fields(
                    dictionary.get('fields', []),
                    format_type="REDCap"  # Default format, could be configurable
                )

                if conditional_rules:
                    # Run logic validation
                    validator = LogicValidator()
                    logic_violations = validator.validate(conditional_rules, data)

                    # Add logic violations to issues list
                    for violation in logic_violations:
                        issues.append({
                            "type": "logic_violation",
                            "severity": violation.severity,
                            "column": ', '.join(violation.affected_fields),
                            "row": violation.row_index,
                            "value": violation.actual_values,
                            "message": f"Logic rule violated: {violation.rule_description}"
                        })

                    logic_violations_count = len(logic_violations)
            except Exception as e:
                # Log error but don't fail the entire analysis
                print(f"Logic validation error: {e}")

        if logic_violations_count:
            # Logic issues were appended after the API computed its counts -
            # recompute them so the summary stays consistent with `issues`.
            summary["issues_found"] = len(issues)
            summary["critical_issues"] = sum(1 for i in issues if i.get('severity') == 'error')
            summary["warnings"] = sum(1 for i in issues if i.get('severity') == 'warning')
            recommendations.append({
                "type": "conditional_logic",
                "priority": "critical",
                "message": "Conditional logic violations detected. Review field dependencies and branching rules in data dictionary"
            })
        summary["logic_violations_count"] = logic_violations_count

        return {
            "summary": summary,
            "issues": issues,
            "recommendations": recommendations,
            "quality_checks": result["quality_checks"],
            "summary_stats": result["summary_stats"]
        }

    # Content-types matching api_server.py's /api/v1/analyze extension
    # dispatch (and the standard IANA media types for each Excel format).
    _EXCEL_CONTENT_TYPES = {
        "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "xls": "application/vnd.ms-excel",
    }

    def _call_analyze_api(self, data: pd.DataFrame, schema: Optional[Dict], rules: Optional[Dict], source_format: Optional[str] = None) -> Dict[str, Any]:
        """POST a DataFrame to the API's /api/v1/analyze endpoint and return
        the parsed JSON report.

        Args:
            data: DataFrame to analyze.
            schema: Optional column-type map, sent as a JSON form field.
            rules: Optional validation-rule map, sent as a JSON form field.
            source_format: "xlsx", "xls", or None. When set, `data` is
                re-encoded as a real Excel workbook and uploaded with the
                matching Excel content-type so the API's Excel branch
                (mcp_server.DataLoader.load_excel) handles it. Defaults to
                CSV serialization, which the API also fully supports.

        Returns:
            The parsed JSON body: {"summary", "issues", "recommendations",
            "quality_checks", "summary_stats"}.

        Raises:
            RuntimeError: on any non-200 response, connection error, or
                timeout, with a message suitable for st.error(...).
        """
        source_format = (source_format or "").lower().lstrip(".") or None

        if source_format in self._EXCEL_CONTENT_TYPES:
            excel_buffer = io.BytesIO()
            data.to_excel(excel_buffer, index=False, engine="openpyxl")
            file_bytes = excel_buffer.getvalue()
            filename = f"data.{source_format}"
            content_type = self._EXCEL_CONTENT_TYPES[source_format]
        else:
            file_bytes = data.to_csv(index=False).encode("utf-8")
            filename = "data.csv"
            content_type = "text/csv"

        files = {"data_file": (filename, file_bytes, content_type)}

        form_data = {}
        if schema:
            form_data["schema"] = json.dumps(schema, default=str)
        if rules:
            form_data["rules"] = json.dumps(rules, default=str)

        headers = {"X-API-Key": self.api_key} if self.api_key else {}

        try:
            response = requests.post(
                f"{self.api_base_url}/api/v1/analyze",
                files=files,
                data=form_data,
                headers=headers,
                timeout=60,
            )
        except requests.exceptions.ConnectionError as e:
            raise RuntimeError(
                f"Could not connect to the data-analyzer API at {self.api_base_url}. "
                f"Is it running and reachable? ({e})"
            )
        except requests.exceptions.Timeout:
            raise RuntimeError(
                f"Data-analyzer API request timed out after 60s ({self.api_base_url}/api/v1/analyze)."
            )
        except requests.exceptions.RequestException as e:
            raise RuntimeError(f"Data-analyzer API request failed: {e}")

        if response.status_code != 200:
            try:
                error_body = response.json()
                detail = error_body.get("detail") or error_body.get("error") or response.text
            except ValueError:
                detail = response.text
            raise RuntimeError(f"Data-analyzer API returned HTTP {response.status_code}: {detail}")

        try:
            return response.json()
        except ValueError as e:
            raise RuntimeError(f"Data-analyzer API returned a non-JSON response: {e}")

def load_demo_data(dataset_name: str):
    """Load demo dataset matching the dictionary options"""
    demo_data = {
        'western': pd.DataFrame({
            'employee_id': [1001, 1002, 1003, 1004, 1005],
            'first_name': ['John', 'Jane', 'Mike', 'Bob', 'Alice'],
            'last_name': ['Smith', 'Doe', 'Johnson', 'Brown', 'Wilson'],
            'age': [35, 28, 67, 45, 32],  # ERROR: 67 is outside range (max 65)
            'salary': [75000, 85000, 45000, 95000, None],  # ERROR: 45000 below min (50000), WARNING: None is missing
            'hire_date': ['2022-03-15', '2023-01-10', '2023-99-99', '2021-07-22', '2022-11-30'],  # ERROR: invalid date 2023-99-99
            'last_login_datetime': ['2023-01-15 10:30:00', '2023-02-20 14:45:00', '2023-03-10 09:15:00', '2023-04-05 16:20:00', '2023-05-12 11:00:00'],
            'bonus_percentage': [5.5, 10.0, 15.5, 8.0, 12.5],
            'department': ['Engineering', 'Marketing', 'InvalidDept', 'Sales', 'Finance'],  # ERROR: InvalidDept not in allowed values
            'is_active': [True, True, False, None, True],  # WARNING: None is missing
            'skills': ['Python;SQL', 'Marketing;Analytics', 'Java;AWS', 'Sales;CRM', 'Python;Leadership'],
            'email': ['john@company.com', 'jane@company.com', 'mike@company.com', 'bob@company.com', 'alice@company.com'],
            'phone': ['+1-555-1234', '+1-555-5678', '+1-555-9012', None, '+1-555-3456']  # WARNING: None is missing
        }),
        'clinical': pd.DataFrame({
            'patient_id': ['P001', 'P002', 'P003', 'P004', 'P005', 'P006', 'P007', 'P008'],
            'age': [45, 62, 73, 28, 155, 39, 67, 51],  # 155 is invalid age
            'gender': ['M', 'F', 'M', 'F', 'X', 'F', 'M', 'invalid'],  # X and invalid are issues
            'diagnosis_code': ['I21.0', 'J18.9', 'N18.9', 'K92.2', 'invalid', 'G20.9', 'E11.9', ''],
            'admission_date': ['2024-01-15', '2024-01-16', '2024-01-17', '2024-01-18', 'invalid-date', '2024-01-20', '2024-01-21', '2024-01-22'],
            'discharge_date': ['2024-01-20', '2024-01-22', '2024-01-25', '2024-01-19', None, '2024-01-21', '', '2024-01-23'],
            'treatment_type': ['Emergency', 'Inpatient', 'Inpatient', 'Observation', 'invalid', 'Outpatient', 'Inpatient', 'Unknown'],
            'lab_result_wbc': [7.5, 12.3, 6.8, 8.2, 50.0, 7.8, 10.5, None],  # 50.0 is out of range
            'lab_result_hemoglobin': [14.2, 12.8, 10.5, 13.5, 'invalid', 14.5, None, 13.2],
            'blood_pressure_systolic': [130, 145, 155, 110, 250, 125, 165, 135],  # 250 is too high
            'blood_pressure_diastolic': [85, 92, 95, 70, 130, 80, 98, 82],  # 130 is too high
            'temperature': [37.2, 38.5, 36.8, 36.9, 45.0, 36.7, 37.3, None],  # 45.0 is impossible
            'heart_rate': [88, 96, 78, 72, 200, 75, 90, 68],  # 200 is too high
            'follow_up_required': ['Yes', 'Yes', 'Yes', 'No', 'Maybe', 'Yes', 'Yes', None],  # Maybe is invalid
            'outcome_status': ['Improved', 'Recovered', 'Stable', 'Recovered', 'invalid', 'Stable', 'Ongoing', 'Improved'],
            'length_of_stay': [5, 6, 8, 1, 500, 1, None, 1]  # 500 days is unrealistic
        }),
        'asian': pd.DataFrame({
            'staff_id': [2001, 2002, 2003, 2004, 2005],
            'given_name': ['Akiko', 'Wei', None, 'Raj', 'Mei'],
            'family_name': ['Tanaka', 'Zhang', 'Kumar', 'invalid', 'Chen'],
            'age': [30, 21, 45, 38, 62],  # 21 below min, 62 above max
            'monthly_salary': [8500, 9200, 6000, 10500, None],  # 6000 below min
            'join_date': ['2020-06-01', 'invalid-date', '2021-03-15', '2022-09-10', '2023-01-20'],
            'dept_code': ['DEV', 'MKT', 'invalid', 'OPS', 'FIN'],
            'active_status': [1, 1, 0, None, 1],
            'work_email': ['akiko@work.com', 'wei@work.com', 'invalid', 'raj@work.com', 'mei@work.com']
        }),
        'mixed': pd.DataFrame({
            'id': [3001, 3002, 3003, 3004, 3005],
            'name_first': ['Carlos', 'Emma', 'invalid', 'Liu', None],
            'name_last': ['Rodriguez', 'invalid', 'Brown', 'Wang', 'Lee'],
            'age': [40, 24, 35, 56, 45],  # 24 below min, 56 above max
            'salary': [70000, 80000, 60000, 90000, None],  # 60000 below min, 90000 above max
            'hired': ['2022-05-01', '2023-08-15', 'invalid-date', '2021-12-01', '2023-03-10'],
            'active': [True, False, None, True, True],
            'department': ['Research', 'invalid', 'Engineering', 'Quality', 'Sales']
        }),
        # Synthetic REDCap-style clinical dataset matching the field set of the
        # "REDCap - Clinical (synthetic, with branching logic)" demo dictionary
        # in demo_dictionaries.py (16 columns spanning the demographics,
        # treatment, medical_history, and safety REDCap forms). 20 rows,
        # deliberately containing several quality issues so the demo produces
        # findings:
        #   1. age[4] = 91          -> ERROR: exceeds dictionary max (85)
        #   2. gender[7] = 'Unknown'-> ERROR: not in allowed values (Male/Female/Other)
        #   3. lab_glucose[9] = None-> WARNING: missing value where diabetic + required
        #   4. dosage_mg[9] = 750   -> ERROR: exceeds dictionary max (500)
        #   5. BRANCHING-LOGIC VIOLATION row 14 (subject TEST-014): pregnant='Yes'
        #      but gender='Male' (pregnant should only be answerable when
        #      gender='Female' per the dictionary's branching logic) -
        #      demonstrates a logic-check finding.
        'redcap_clinical': pd.DataFrame({
            'subject_id': [f'TEST-{i:03d}' for i in range(1, 21)],
            'age': [34, 45, 58, 91, 27, 62, 39, 50, 71, 44,  # row 4 (idx3) = 91 -> ERROR (max 85)
                    29, 55, 48, 36, 60, 41, 33, 52, 67, 25],
            'gender': ['Female', 'Male', 'Female', 'Male', 'Other', 'Female', 'Unknown', 'Male', 'Female', 'Male',
                       'Female', 'Male', 'Female', 'Male', 'Female', 'Male', 'Female', 'Male', 'Female', 'Male'],  # row 7 (idx6) = 'Unknown' -> ERROR
            'pregnant': ['No', None, 'Yes', None, None, 'No', None, None, 'No', None,
                         'Yes', None, 'No', 'Yes', 'No', None, 'No', None, 'No', None],  # row 14 (idx13) = 'Yes' but gender[13]='Male' -> BRANCHING-LOGIC VIOLATION
            'weeks_pregnant': [None, None, 24, None, None, None, None, None, None, None,
                               12, None, None, None, None, None, None, None, None, None],
            'due_date': [None, None, '2024-08-15', None, None, None, None, None, None, None,
                         '2024-11-02', None, None, None, None, None, None, None, None, None],
            'treatment_arm': ['Active Treatment', 'Placebo', 'Active Treatment', 'Placebo', 'Active Treatment',
                               'Placebo', 'Active Treatment', 'Placebo', 'Active Treatment', 'Placebo',
                               'Active Treatment', 'Placebo', 'Active Treatment', 'Placebo', 'Active Treatment',
                               'Placebo', 'Active Treatment', 'Placebo', 'Active Treatment', 'Placebo'],
            'dosage_mg': [120, None, 85, None, 200, None, 150, None, 750, None,  # row 9 (idx8) = 750 -> ERROR (max 500)
                          95, None, 180, None, 110, None, 60, None, 140, None],
            'placebo_type': [None, 'Tablet', None, 'Capsule', None, 'Tablet', None, 'Capsule', None, 'Tablet',
                              None, 'Capsule', None, 'Tablet', None, 'Capsule', None, 'Tablet', None, 'Capsule'],
            'diabetes': ['No', 'Yes', 'No', 'Yes', 'No', 'Yes', 'No', 'Yes', 'No', 'Yes',
                         'No', 'Yes', 'No', 'Yes', 'No', 'Yes', 'No', 'Yes', 'No', 'Yes'],
            'diabetes_type': [None, 'Type 2', None, 'Type 1', None, 'Type 2', None, 'Type 2', None, 'Type 1',
                               None, 'Type 2', None, 'Type 2', None, 'Type 1', None, 'Type 2', None, 'Type 2'],
            'insulin_dependent': [None, 'No', None, 'Yes', None, 'No', None, 'No', None, 'Yes',
                                   None, 'No', None, 'No', None, 'Yes', None, 'No', None, 'No'],
            'lab_glucose': [None, 145.0, None, 210.5, None, 130.0, None, None, None, 190.0,  # row 8 (idx7): diabetic + required but missing -> WARNING
                             None, 128.0, None, 175.0, None, 205.0, None, 122.0, None, 168.0],
            'adverse_event': ['No', 'No', 'Yes', 'No', 'No', 'Yes', 'No', 'No', 'No', 'Yes',
                               'No', 'No', 'Yes', 'No', 'No', 'No', 'Yes', 'No', 'No', 'No'],
            'ae_description': [None, None, 'Mild headache, resolved', None, None, 'Nausea after dosing', None, None, None, 'Dizziness',
                                None, None, 'Injection site rash', None, None, None, 'Fatigue', None, None, None],
            'ae_severity': [None, None, 'Mild', None, None, 'Moderate', None, None, None, 'Mild',
                             None, None, 'Mild', None, None, None, 'Moderate', None, None, None],
        })
    }
    return demo_data.get(dataset_name, demo_data['western'])

# Severity vocabulary used across the codebase (verified against api_server.py's
# _build_quality_report and src/logic_engine.py's LogicValidator/ConditionalRule):
# only "error" and "warning" are ever emitted. logic_violation issues (merged in
# by DataQualityAnalyzer.analyze_data_quality from `violation.severity`) also draw
# from that same two-value vocabulary - src/logic_engine.py never emits "critical"
# or any other string. We still map defensively: anything that isn't literally
# "warning" is treated as the red/error tier, so an unexpected future severity
# string degrades to "shown as an error" rather than silently vanishing.
_HEATMAP_WARNING_SEVERITIES = {'warning'}


def _build_issue_matrix(df: pd.DataFrame, issues: list, max_display_rows: int = 60, max_display_cols: int = 100):
    """Pure helper: turn a DataFrame + issues list into heatmap inputs.

    Separated from create_issue_heatmap() so the row/column mapping and
    severity classification logic can be unit tested without a Streamlit
    runtime (create_issue_heatmap has no return value since it renders
    directly via st.plotly_chart/st.caption).

    Returns a dict with:
        issue_matrix: np.ndarray (display_rows x display_cols) of 0/1/2
        hover_text: list of lists of hover strings, same shape
        display_rows, display_cols, row_factor, col_factor: layout ints
        unmapped_column_warnings: dict[column_name -> count] of column-level
            issues (e.g. missing_values) that have no 'row' key and so
            cannot be placed on the per-cell grid
    """
    rows, cols = len(df), len(df.columns)

    row_factor = max(1, rows // max_display_rows)
    col_factor = max(1, cols // max_display_cols)

    display_rows = min(rows, max_display_rows)
    display_cols = min(cols, max_display_cols)

    issue_matrix = np.zeros((display_rows, display_cols))
    hover_text = [['' for _ in range(display_cols)] for _ in range(display_rows)]
    unmapped_column_warnings: Dict[str, int] = {}

    for issue in issues:
        if 'row' in issue and 'column' in issue:
            try:
                col_idx = df.columns.get_loc(issue['column'])
                row_idx = issue['row']

                # Map to display coordinates
                display_row = min(row_idx // row_factor, display_rows - 1)
                display_col = min(col_idx // col_factor, display_cols - 1)

                # Set severity (2 for error/other, 1 for warning). Only the
                # "warning" string maps to the yellow tier; everything else
                # (currently just "error") maps to the red tier.
                severity_value = 1 if issue.get('severity') in _HEATMAP_WARNING_SEVERITIES else 2
                issue_matrix[display_row, display_col] = max(issue_matrix[display_row, display_col], severity_value)

                # Build hover text
                issue_info = f"<b>{issue['type'].replace('_', ' ').title()}</b><br>"
                issue_info += f"Row: {row_idx}<br>"
                issue_info += f"Column: {issue['column']}<br>"
                issue_info += f"Value: {issue.get('value', 'N/A')}<br>"
                issue_info += f"Severity: {issue.get('severity', 'unknown')}"

                if hover_text[display_row][display_col]:
                    hover_text[display_row][display_col] += "<br><br>" + issue_info
                else:
                    hover_text[display_row][display_col] = issue_info
            except Exception:
                pass
        else:
            # Column-level aggregate issue (e.g. missing_values warnings built
            # in api_server.py's _build_quality_report) - no row coordinate
            # exists to plot a cell, so track it separately for the caption.
            col_name = issue.get('column')
            if col_name is not None:
                unmapped_column_warnings[col_name] = unmapped_column_warnings.get(col_name, 0) + 1

    return {
        'issue_matrix': issue_matrix,
        'hover_text': hover_text,
        'display_rows': display_rows,
        'display_cols': display_cols,
        'row_factor': row_factor,
        'col_factor': col_factor,
        'unmapped_column_warnings': unmapped_column_warnings,
    }


def create_issue_heatmap(df: pd.DataFrame, issues: list):
    """Create an interactive heatmap with hover tooltips using Plotly"""
    try:
        rows, cols = len(df), len(df.columns)
        max_display_rows = 60
        max_display_cols = 100  # Increased for wide datasets

        built = _build_issue_matrix(df, issues, max_display_rows, max_display_cols)
        issue_matrix = built['issue_matrix']
        hover_text = built['hover_text']
        row_factor = built['row_factor']
        col_factor = built['col_factor']
        unmapped_column_warnings = built['unmapped_column_warnings']

        # Calculate aspect ratio for proper dimensions
        aspect_ratio = cols / rows
        if aspect_ratio > 1:
            # Wide dataset
            fig_width = 300
            fig_height = max(50, 300 / aspect_ratio)
        else:
            # Tall dataset
            fig_height = 300
            fig_width = max(50, 300 * aspect_ratio)

        # Create interactive Plotly heatmap. zmin/zmax are pinned explicitly so
        # the colorscale mapping (white/yellow/red at z=0/1/2) stays stable
        # regardless of which severities are actually present - without this,
        # a dataset with only warning cells (max z=1) would have Plotly
        # normalize 1 -> the top of the scale (red) instead of the middle.
        fig = go.Figure(data=go.Heatmap(
            z=issue_matrix,
            zmin=0,
            zmax=2,
            text=hover_text,
            hovertemplate='%{text}<extra></extra>',
            colorscale=[
                [0, '#ffffff'],      # White for no issue
                [0.5, '#fbbf24'],    # Yellow for warning
                [1, '#ef4444']       # Red for error
            ],
            showscale=False,
            xgap=1,
            ygap=1
        ))

        # Update layout. All color-related properties are set explicitly
        # (template + paper/plot bgcolor + font/axis colors) so the figure
        # renders identically regardless of Streamlit's active theme or the
        # app's custom dark-mode CSS overrides - without an explicit
        # template, dark-mode CSS was overriding the plot to solid orange
        # with no visible cell boundaries.
        fig.update_layout(
            template='plotly_white',
            title={
                'text': f'{rows} rows × {cols} cols' + (f' (scale {row_factor}:{col_factor})' if rows > max_display_rows or cols > max_display_cols else ''),
                'font': {'size': 10, 'color': '#111111'}
            },
            width=fig_width,
            height=fig_height,
            margin=dict(l=20, r=20, t=30, b=20),
            xaxis={'showticklabels': False, 'showgrid': False, 'color': '#111111'},
            yaxis={'showticklabels': False, 'showgrid': False, 'color': '#111111'},
            paper_bgcolor='white',
            plot_bgcolor='white',
            font_color='#111111'
        )

        # Display with Streamlit
        st.plotly_chart(fig, use_container_width=False)

        # Show summary. Cell-level counts come from the plotted matrix;
        # column-level (unmappable) warnings - e.g. missing_values issues,
        # which are column aggregates with no row coordinate - are called
        # out separately so the caption never claims "0 warnings" when
        # warnings exist but simply couldn't be drawn as cells.
        error_count = np.sum(issue_matrix == 2)
        warning_count = np.sum(issue_matrix == 1)
        unmapped_count = sum(unmapped_column_warnings.values())

        caption_parts = []
        if error_count > 0 or warning_count > 0:
            caption_parts.append(f"🔴 {int(error_count)} cells with errors, 🟡 {int(warning_count)} cells with warnings")
        if unmapped_count > 0:
            cols_str = ', '.join(list(unmapped_column_warnings.keys())[:5])
            if len(unmapped_column_warnings) > 5:
                cols_str += f", +{len(unmapped_column_warnings) - 5} more"
            caption_parts.append(
                f"⚠️ {unmapped_count} additional column-level warning(s) not shown as cells "
                f"(e.g. missing-value summaries for: {cols_str}) — see Issues Found below"
            )
        if caption_parts:
            st.caption(" · ".join(caption_parts))

    except Exception as e:
        st.caption(f"Issue map unavailable: {str(e)}")

def export_to_excel_with_highlighting(df: pd.DataFrame, issues: list) -> bytes:
    """Export data to Excel with error cells highlighted"""
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Data', index=False)

        # Create issues summary sheet
        issues_df = pd.DataFrame(issues)
        if not issues_df.empty:
            issues_df.to_excel(writer, sheet_name='Issues', index=False)

        # Get workbook and worksheet
        workbook = writer.book
        worksheet = workbook['Data']

        # Apply highlighting to cells with issues
        from openpyxl.styles import PatternFill, Font

        error_fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")
        warning_fill = PatternFill(start_color="FFE5B4", end_color="FFE5B4", fill_type="solid")

        for issue in issues:
            if 'row' in issue and 'column' in issue:
                try:
                    col_idx = df.columns.get_loc(issue['column']) + 1
                    row_idx = issue['row'] + 2  # +2 for header and 0-index
                    cell = worksheet.cell(row=row_idx, column=col_idx)

                    if issue['severity'] == 'error':
                        cell.fill = error_fill
                    elif issue['severity'] == 'warning':
                        cell.fill = warning_fill

                    # Add comment with issue details
                    from openpyxl.comments import Comment
                    cell.comment = Comment(issue['message'], "Data Analyzer")
                except:
                    pass

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    return output.getvalue()

# Initialize session state
if 'data' not in st.session_state:
    st.session_state.data = None
if 'data_source_format' not in st.session_state:
    # Tracks the original upload format ('csv', 'json', 'tsv', 'xlsx', 'xls')
    # so the analysis HTTP call can re-serialize st.session_state.data with a
    # matching content-type instead of always defaulting to CSV.
    st.session_state.data_source_format = None
if 'dictionary' not in st.session_state:
    st.session_state.dictionary = None
if 'analysis_results' not in st.session_state:
    st.session_state.analysis_results = None
if 'mcp_client' not in st.session_state:
    st.session_state.mcp_client = DataQualityAnalyzer()
if 'dict_cache' not in st.session_state:
    st.session_state.dict_cache = {}  # Cache for parsed dictionaries
if 'last_dict_file' not in st.session_state:
    st.session_state.last_dict_file = None  # Track last uploaded dictionary
if 'cache_dir' not in st.session_state:
    # Create persistent cache directory (user-specific to avoid multi-user conflicts)
    import getpass
    try:
        username = getpass.getuser()
    except:
        # Fallback if getuser() fails
        import os
        username = os.environ.get('USER', os.environ.get('USERNAME', 'default'))

    cache_dir = Path.home() / f'.data_analyzer_cache_{username}'
    cache_dir.mkdir(exist_ok=True)
    st.session_state.cache_dir = cache_dir

    print(f"\n📦 Cache directory: {cache_dir}")
    print(f"   (User-specific to avoid multi-user conflicts)\n")

if nav_choice == "Analyze":
    # Subtitle only - title is in the navbar
    st.markdown("Upload your data, optionally add validation rules, and analyze")

    # Create three columns for the main components
    col1, col2, col3 = st.columns([2, 2, 1.5])

    with col1:
        with st.container(key="card-upload-data"):
            st.markdown('<div class="app-section-title">Upload Data</div>', unsafe_allow_html=True)


            # File uploader first
            uploaded_file = st.file_uploader(
                " ",  # Empty label to avoid duplication
                type=['csv', 'json', 'txt', 'xlsx', 'xls'],
                key="data_uploader",
                label_visibility="collapsed"
            )

            if uploaded_file:
                with st.spinner(f"Processing {uploaded_file.name}..."):
                    try:
                        if uploaded_file.name.endswith('.csv'):
                            st.session_state.data = pd.read_csv(uploaded_file)
                            st.session_state.data_source_format = 'csv'
                        elif uploaded_file.name.endswith('.json'):
                            st.session_state.data = pd.read_json(uploaded_file)
                            st.session_state.data_source_format = 'json'
                        elif uploaded_file.name.endswith(('.xlsx', '.xls')):
                            # Shared loader (mcp_server.DataLoader) so MCP server, CLI,
                            # and this UI all get identical Excel handling/error semantics.
                            st.session_state.data = DataLoader.load_excel(uploaded_file.read())
                            # Remember the Excel sub-format so the analysis HTTP
                            # call (DataQualityAnalyzer._call_analyze_api) can
                            # re-upload real Excel bytes with the correct
                            # content-type instead of defaulting to CSV.
                            st.session_state.data_source_format = 'xlsx' if uploaded_file.name.endswith('.xlsx') else 'xls'
                        else:
                            st.session_state.data = pd.read_csv(uploaded_file, sep='\t')
                            st.session_state.data_source_format = 'tsv'
                        st.success(f"✅ Loaded {len(st.session_state.data)} rows × {len(st.session_state.data.columns)} columns")
                    except Exception as e:
                        st.error(f"Error loading file: {str(e)}")

            # Demo data selector below file uploader
            demo_option = st.selectbox(
                "Or load demo data:",
                ["None", "CSV - Western", "CSV - Asian", "CSV - Clinical", "JSON - Mixed", "REDCap - Clinical (synthetic)"],
                key="demo_selector",
                help="Clinical data includes matching dictionary in demo_data/clinical_dict.json"
            )

            if demo_option != "None":
                dataset_map = {
                    "CSV - Western": "western",
                    "CSV - Asian": "asian",
                    "CSV - Clinical": "clinical",
                    "JSON - Mixed": "mixed",
                    "REDCap - Clinical (synthetic)": "redcap_clinical"
                }
                if demo_option in dataset_map:
                    st.session_state.data = load_demo_data(dataset_map[demo_option])
                    # Demo datasets are always synthesized in-memory (not loaded
                    # from an Excel file), so fall back to CSV serialization for
                    # the analysis HTTP call.
                    st.session_state.data_source_format = 'csv'
                    st.success(f"✅ Loaded {demo_option} demo data")
                    if demo_option == "CSV - Clinical":
                        st.info("📖 Matching dictionary available: Upload 'demo_data/clinical_dict.json' for validation rules")
                    if demo_option == "REDCap - Clinical (synthetic)":
                        st.info("📖 Matching dictionary available: select 'REDCap - Clinical (synthetic, with branching logic)' under 'Or load demo dictionary' below")

    with col2:
        with st.container(key="card-dictionary"):
            st.markdown('<div class="app-section-title">Dictionary</div>', unsafe_allow_html=True)


            # Dictionary file uploader first - aligned with data uploader
            dict_file = st.file_uploader(
                " ",  # Empty label to avoid duplication
                type=['json', 'pdf', 'csv', 'txt', 'xlsx', 'xls'],
                key="dict_uploader",
                label_visibility="collapsed",
                help="Optional - defines validation rules for data quality checks (JSON, PDF, CSV, Excel, or TXT)"
            )

            # LLM model selector - visible whenever AI-assisted dictionary
            # parsing is available at all (not just after a file is uploaded),
            # so it's already set by the time the user uploads a file that
            # needs it. Defaults to the first configured deployment.
            if LLM_AVAILABLE:
                if "llm_deployment" not in st.session_state:
                    _available_deployments = get_available_deployments()
                    st.session_state.llm_deployment = _available_deployments[0] if _available_deployments else None

                st.session_state.llm_deployment = st.selectbox(
                    "LLM model",
                    get_available_deployments(),
                    key="llm_deployment_select",
                    help="Azure OpenAI deployment used for AI-assisted dictionary parsing.",
                )

            # Add LLM parsing option if available with auto-detection
            if LLM_AVAILABLE and dict_file:
                # DEBUG: Show filename
                st.caption(f"📎 Uploaded file: **{dict_file.name}** ({dict_file.type if hasattr(dict_file, 'type') else 'unknown type'})")

                llm_mode = st.selectbox(
                    "Dictionary parsing method:",
                    ["Auto-detect (recommended)", "Always use AI parsing", "Never use AI (manual only)"],
                    index=0,  # Default to auto-detect
                    help="Auto: PDF→AI, structured CSV→manual | Always: Force AI for all | Never: Manual parsing only"
                )

                # Determine if we should use LLM based on mode and file type
                if "Always" in llm_mode:
                    use_llm = True
                elif "Never" in llm_mode:
                    use_llm = False
                else:  # Auto-detect
                    # Check file type and structure
                    print(f"\n🔍 AUTO-DETECT: Checking file '{dict_file.name}'")
                    print(f"   Extension check: .pdf={dict_file.name.endswith('.pdf')}, .csv={dict_file.name.endswith('.csv')}")

                    if dict_file.name.endswith('.pdf'):
                        use_llm = True
                        st.info("🤖 Auto-detected: PDF requires AI parsing")
                        print(f"   ✅ Detected as PDF, will use LLM")
                    elif dict_file.name.endswith('.csv'):
                        # Peek at CSV to check if it's structured
                        dict_file.seek(0)
                        sample = dict_file.read(1024).decode('utf-8', errors='ignore')
                        dict_file.seek(0)
                        # Check for standard column names
                        if any(col in sample for col in ['Column', 'Type', 'Min', 'Max', 'Allowed_Values', 'Field Name']):
                            use_llm = False
                            st.info("📊 Auto-detected: Structured CSV, using manual parsing")
                        else:
                            use_llm = True
                            st.info("🤖 Auto-detected: Unstructured CSV, using AI parsing")
                    elif dict_file.name.endswith(('.xlsx', '.xls')):
                        # Peek at Excel column headers to check if it's structured
                        # (same convention as the CSV structured-column check above)
                        dict_file.seek(0)
                        try:
                            peek_df = DataLoader.load_excel(dict_file.read())
                            columns_str = ' '.join(str(c) for c in peek_df.columns)
                        except Exception:
                            columns_str = ""
                        dict_file.seek(0)
                        if any(col in columns_str for col in ['Column', 'Type', 'Min', 'Max', 'Allowed_Values', 'Field Name']):
                            use_llm = False
                            st.info("📊 Auto-detected: Structured Excel, using manual parsing")
                        else:
                            use_llm = True
                            st.info("🤖 Auto-detected: Unstructured Excel, using AI parsing")
                    else:
                        use_llm = True
                        st.info(f"🤖 Auto-detected: {dict_file.name.split('.')[-1].upper()} file, using AI parsing")
            else:
                use_llm = False

            if dict_file:
                try:
                    # Calculate file hash for caching (works for all file types)
                    dict_file.seek(0)
                    raw_content = dict_file.read()
                    file_hash = hashlib.md5(raw_content).hexdigest()
                    dict_file.seek(0)  # Reset for reading

                    # Create cache key with LLM flag
                    cache_key = f"{file_hash}_llm" if use_llm else file_hash
                    cache_file = st.session_state.cache_dir / f"{cache_key}.json"

                    # Check if already cached
                    if cache_key in st.session_state.dict_cache:
                        # Use in-memory cache
                        st.session_state.dictionary = st.session_state.dict_cache[cache_key]

                        # CLEAR CACHE LOGGING
                        print("\n" + "="*80)
                        print("💾 LOADING FROM MEMORY CACHE (NO LLM CALL)")
                        print(f"   Cache key: {cache_key}")
                        print(f"   File: {dict_file.name}")
                        print("="*80 + "\n")

                        st.success(f"⚡ Using cached dictionary (instant load)")
                        st.warning("🔄 **CACHE HIT**: Using previously parsed dictionary. Clear cache below if data dictionary changed.")

                        if use_llm:
                            fields_count = len(st.session_state.dictionary.get('fields', []))
                            st.info(f"📊 Contains {fields_count} AI-extracted field definitions")
                        else:
                            st.info(f"📊 Contains {len(st.session_state.dictionary.get('rules', {}))} validation rules")
                    elif cache_file.exists():
                        # Load from persistent cache file (JSON, not pickle - avoids
                        # insecure deserialization if the cache dir is ever shared/writable)
                        with open(cache_file, 'r', encoding='utf-8') as f:
                            st.session_state.dictionary = json.load(f)
                            st.session_state.dict_cache[cache_key] = st.session_state.dictionary

                        # CLEAR CACHE LOGGING
                        print("\n" + "="*80)
                        print("💾 LOADING FROM DISK CACHE (NO LLM CALL)")
                        print(f"   Cache file: {cache_file.name}")
                        print(f"   File: {dict_file.name}")
                        print("="*80 + "\n")

                        st.success(f"⚡ Loaded dictionary from disk cache (no API calls)")
                        st.warning("🔄 **CACHE HIT**: Using previously parsed dictionary. Clear cache below if data dictionary changed.")

                        if use_llm:
                            fields_count = len(st.session_state.dictionary.get('fields', []))
                            st.info(f"📊 Contains {fields_count} AI-extracted field definitions")
                        else:
                            st.info(f"📊 Contains {len(st.session_state.dictionary.get('rules', {}))} validation rules")
                    # Use LLM parsing if enabled and not cached
                    elif use_llm and LLM_AVAILABLE:
                        # CLEAR LLM MARKER
                        st.warning("🤖 **LLM ACTIVE**: Sending data to Azure OpenAI GPT-4 for intelligent dictionary parsing...")
                        print("\n" + "="*80)
                        print("🤖 LLM DICTIONARY PARSER INVOKED")
                        print(f"   File: {dict_file.name}")
                        print(f"   Size: {len(raw_content)} bytes")
                        print("="*80 + "\n")

                        # More informative spinner with warning about processing time
                        with st.spinner("🤖 Using AI to extract field definitions... This may take 30-60 seconds for large PDFs."):
                            # Read file content
                            file_content = ""
                            if dict_file.name.endswith('.pdf'):
                                pdf_reader = pypdf.PdfReader(dict_file)
                                for page in pdf_reader.pages:
                                    file_content += page.extract_text() + "\n"
                            elif dict_file.name.endswith('.csv'):
                                file_content = dict_file.read().decode('utf-8')
                            elif dict_file.name.endswith('.txt'):
                                file_content = dict_file.read().decode('utf-8')
                            elif dict_file.name.endswith('.json'):
                                # For JSON, convert to readable text
                                json_data = json.load(dict_file)
                                file_content = json.dumps(json_data, indent=2)
                            elif dict_file.name.endswith(('.xlsx', '.xls')):
                                # For Excel, convert the sheet to readable text (same
                                # approach as the JSON branch above - binary content
                                # can't be decoded as utf-8 like the CSV/TXT branches)
                                excel_df = DataLoader.load_excel(dict_file.read())
                                file_content = excel_df.to_csv(index=False)
                            else:
                                file_content = dict_file.read().decode('utf-8')

                            # Initialize LLM parser
                            llm_parser = LLMDictionaryParser()

                            # Estimate tokens for browser console log
                            import time
                            estimated_tokens = llm_parser.count_tokens(file_content)
                            start_time = time.time()
                            start_timestamp = time.strftime('%H:%M:%S')

                            # Log to browser console
                            log_to_browser_console(
                                f"🤖 LLM parsing started at {start_timestamp}",
                                {
                                    "model": llm_parser.deployment,
                                    "tokens": estimated_tokens,
                                    "file": dict_file.name,
                                    "size_bytes": len(file_content)
                                }
                            )

                            # Parse with LLM
                            # Don't truncate - the LLM parser handles chunking internally
                            # Process more fields for comprehensive extraction
                            parsed_result = llm_parser.parse_dictionary(
                                file_content,
                                max_fields=500,
                                deployment=st.session_state.get("llm_deployment"),
                            )

                            # Calculate elapsed time
                            elapsed_time = time.time() - start_time

                            # Log completion to browser console
                            log_to_browser_console(
                                f"✅ LLM parsing completed in {elapsed_time:.1f}s",
                                {
                                    "fields_extracted": len(parsed_result.get('fields', [])),
                                    "chunks_processed": parsed_result.get('metadata', {}).get('chunks_processed', 0),
                                    "mode": parsed_result.get('metadata', {}).get('mode', 'unknown')
                                }
                            )

                            # Store the parsed dictionary
                            st.session_state.dictionary = {
                                "source": "LLM Parser",
                                "filename": dict_file.name,
                                "rules": parsed_result.get("schema", {}),
                                "fields": parsed_result.get("fields", []),
                                "metadata": parsed_result.get("metadata", {})
                            }

                            # Cache the result both in memory and to disk
                            st.session_state.dict_cache[cache_key] = st.session_state.dictionary
                            with open(cache_file, 'w', encoding='utf-8') as f:
                                json.dump(st.session_state.dictionary, f)
                            st.info(f"💾 Dictionary cached - future loads will be instant (no API calls)")

                            # Add processing time to success message
                            processing_time = parsed_result.get('metadata', {}).get('processing_time_seconds', 0)
                            chunks_processed = parsed_result.get('metadata', {}).get('chunks_processed', 0)
                            st.success(f"✅ AI extracted {len(parsed_result.get('fields', []))} field definitions from {chunks_processed} chunks in {processing_time:.1f} seconds")

                            # Show extracted fields
                            if parsed_result.get('fields'):
                                # Expand by default if we got results, especially for large dictionaries
                                expand_fields = len(parsed_result['fields']) <= 20
                                with st.expander(f"📋 Extracted Fields ({len(parsed_result['fields'])})", expanded=expand_fields):
                                    for field in parsed_result['fields'][:10]:
                                        field_info = f"**{field['field_name']}** ({field['data_type']})"
                                        if field.get('required'):
                                            field_info += " *[Required]*"
                                        if field.get('description'):
                                            field_info += f"\n   {field['description']}"
                                        if field.get('min_value') or field.get('max_value'):
                                            field_info += f"\n   Range: {field.get('min_value', 'N/A')} - {field.get('max_value', 'N/A')}"
                                        if field.get('allowed_values'):
                                            field_info += f"\n   Values: {', '.join(field['allowed_values'][:5])}"
                                        st.markdown(field_info)
                                    if len(parsed_result['fields']) > 10:
                                        st.info(f"📊 Showing first 10 of {len(parsed_result['fields'])} extracted fields. Use 'View All Fields' below to see more.")

                    elif dict_file.name.endswith('.pdf'):
                        # PDF without LLM - already have hash from above
                        dict_file.seek(0)  # Reset for reading

                        # Check persistent file cache first
                        cache_file = st.session_state.cache_dir / f"{file_hash}.json"

                        if cache_file.exists():
                            # Load from persistent cache file (JSON, not pickle)
                            with open(cache_file, 'r', encoding='utf-8') as f:
                                st.session_state.dictionary = json.load(f)
                                st.session_state.dict_cache[file_hash] = st.session_state.dictionary
                            st.success(f"⚡ Loaded dictionary from cache (instant)")
                            st.info(f"📊 Contains {len(st.session_state.dictionary.get('rules', {}))} validation rules")
                        elif file_hash in st.session_state.dict_cache:
                            # Use in-memory cache
                            st.session_state.dictionary = st.session_state.dict_cache[file_hash]
                            st.success(f"⚡ Using cached dictionary '{dict_file.name}' (instant load)")
                            st.info(f"📊 Contains {len(st.session_state.dictionary.get('rules', {}))} validation rules")
                        else:
                            # Manual PDF parsing (NO LLM)
                            st.info("📄 **MANUAL PARSING**: Using basic regex patterns (limited extraction). Enable AI parsing for better results.")
                            print("\n" + "="*80)
                            print("📄 MANUAL PDF PARSER (NO LLM)")
                            print(f"   File: {dict_file.name}")
                            print("   ⚠️ WARNING: Basic regex patterns only - may miss complex field definitions")
                            print("="*80 + "\n")

                            # Parse PDF dictionary with container to prevent UI blocking
                            with st.container():
                                progress_bar = st.progress(0, text="Parsing PDF dictionary...")

                                # Read PDF content
                                pdf_reader = pypdf.PdfReader(dict_file)
                                num_pages = len(pdf_reader.pages)

                                extracted_text = ""
                                extracted_rules = {}

                                # Process pages with continuous progress updates
                                for i, page in enumerate(pdf_reader.pages):
                                    # Update progress for every page
                                    progress_bar.progress((i + 1) / num_pages, text=f"Processing page {i+1} of {num_pages}...")

                                    page_text = page.extract_text()
                                    extracted_text += page_text

                                    # Look for validation rules in the PDF (example patterns)
                                    # Look for date fields
                                    date_fields = re.findall(r'([\w_]+).*?(?:date|Date|DATE)', page_text)
                                    for field in date_fields:
                                        if field not in extracted_rules:
                                            extracted_rules[field] = {"type": "date"}

                                    # Look for numeric ranges
                                    range_patterns = re.findall(r'([\w_]+).*?(?:range|Range|between).*?(\d+).*?(?:to|and|-|–).*?(\d+)', page_text)
                                    for field, min_val, max_val in range_patterns:
                                        if field not in extracted_rules:
                                            extracted_rules[field] = {"min": int(min_val), "max": int(max_val)}

                                # Clear progress bar
                                progress_bar.empty()

                            # Store extracted dictionary
                            st.session_state.dictionary = {
                                "source": "PDF",
                                "filename": dict_file.name,
                                "rules": extracted_rules,
                                "pages": num_pages,
                                "text_length": len(extracted_text),
                                "hash": file_hash
                            }

                            # Cache the parsed dictionary both in memory and to file
                            st.session_state.dict_cache[file_hash] = st.session_state.dictionary

                            # Save to persistent cache file (JSON, not pickle)
                            cache_file = st.session_state.cache_dir / f"{file_hash}.json"
                            with open(cache_file, 'w', encoding='utf-8') as f:
                                json.dump(st.session_state.dictionary, f)

                            st.success(f"✅ Parsed {num_pages} pages from PDF dictionary")
                            st.info(f"💾 Dictionary cached to disk for permanent reuse")
                            st.caption(f"📁 Cache location: {cache_file}")

                            if extracted_rules:
                                with st.expander(f"Found {len(extracted_rules)} validation rules", expanded=False):
                                    for field, rule in list(extracted_rules.items())[:10]:  # Show first 10
                                        st.text(f"{field}: {rule}")
                                    if len(extracted_rules) > 10:
                                        st.text(f"... and {len(extracted_rules) - 10} more")
                    elif dict_file.name.endswith('.json'):
                        st.session_state.dictionary = json.load(dict_file)
                        st.success("✅ JSON dictionary loaded")
                    elif dict_file.name.endswith('.csv') and not use_llm:
                        # Parse CSV dictionary (NO LLM) - only if LLM mode not active
                        st.info("📊 **CSV PARSING**: Reading structured CSV data dictionary...")
                        print(f"\n📊 CSV DICTIONARY PARSER: {dict_file.name}")

                        import pandas as pd
                        dict_file.seek(0)
                        df = pd.read_csv(dict_file)
                        rules = {}
                        for _, row in df.iterrows():
                            if 'Column' in row or 'column' in row or 'Field' in row or 'field' in row:
                                field_name = row.get('Column') or row.get('column') or row.get('Field') or row.get('field')
                                if field_name:
                                    rule = {}
                                    if 'Type' in row or 'type' in row:
                                        rule['type'] = str(row.get('Type') or row.get('type'))
                                    if 'Min' in row or 'min' in row:
                                        rule['min'] = row.get('Min') or row.get('min')
                                    if 'Max' in row or 'max' in row:
                                        rule['max'] = row.get('Max') or row.get('max')
                                    if 'Required' in row or 'required' in row:
                                        rule['required'] = row.get('Required') or row.get('required')
                                    if 'Allowed_Values' in row or 'allowed_values' in row:
                                        allowed = row.get('Allowed_Values') or row.get('allowed_values')
                                        if allowed and not pd.isna(allowed):
                                            rule['allowed_values'] = [v.strip() for v in str(allowed).split(',')]
                                    rules[field_name] = rule
                        st.session_state.dictionary = {
                            "source": "CSV",
                            "filename": dict_file.name,
                            "rules": rules
                        }
                        st.success(f"✅ Parsed {len(rules)} field definitions from CSV")
                    elif dict_file.name.endswith(('.xlsx', '.xls')) and not use_llm:
                        # Parse Excel dictionary (NO LLM) - structured field definitions,
                        # same column-name conventions as the CSV dictionary branch above.
                        st.info("📊 **EXCEL PARSING**: Reading structured Excel data dictionary...")
                        print(f"\n📊 EXCEL DICTIONARY PARSER: {dict_file.name}")

                        dict_file.seek(0)
                        df = DataLoader.load_excel(dict_file.read())
                        rules = {}
                        for _, row in df.iterrows():
                            if 'Column' in row or 'column' in row or 'Field' in row or 'field' in row:
                                field_name = row.get('Column') or row.get('column') or row.get('Field') or row.get('field')
                                if field_name:
                                    rule = {}
                                    if 'Type' in row or 'type' in row:
                                        rule['type'] = str(row.get('Type') or row.get('type'))
                                    if 'Min' in row or 'min' in row:
                                        rule['min'] = row.get('Min') or row.get('min')
                                    if 'Max' in row or 'max' in row:
                                        rule['max'] = row.get('Max') or row.get('max')
                                    if 'Required' in row or 'required' in row:
                                        rule['required'] = row.get('Required') or row.get('required')
                                    if 'Allowed_Values' in row or 'allowed_values' in row:
                                        allowed = row.get('Allowed_Values') or row.get('allowed_values')
                                        if allowed and not pd.isna(allowed):
                                            rule['allowed_values'] = [v.strip() for v in str(allowed).split(',')]
                                    rules[field_name] = rule
                        st.session_state.dictionary = {
                            "source": "Excel",
                            "filename": dict_file.name,
                            "rules": rules
                        }
                        st.success(f"✅ Parsed {len(rules)} field definitions from Excel")
                    else:
                        st.error(f"⚠️ Unsupported dictionary format: **{dict_file.name}**")
                        st.info(f"Debug: use_llm={use_llm}, LLM_AVAILABLE={LLM_AVAILABLE}, file ends with .csv={dict_file.name.endswith('.csv')}")
                        print(f"\n⚠️ UNSUPPORTED FORMAT: {dict_file.name}")
                        print(f"   use_llm: {use_llm}")
                        print(f"   LLM_AVAILABLE: {LLM_AVAILABLE}")
                        print(f"   File extension checks: .pdf={dict_file.name.endswith('.pdf')}, .csv={dict_file.name.endswith('.csv')}, .json={dict_file.name.endswith('.json')}")
                except Exception as e:
                    st.error(f"Error loading dictionary: {str(e)}")

            # View All Fields button - accessible location near dictionary upload
            if st.session_state.dictionary and st.session_state.dictionary.get('fields'):
                fields_list = st.session_state.dictionary['fields']
                num_fields = len(fields_list)

                if num_fields > 0:
                    st.markdown("---")
                    with st.expander(f"📋 View All {num_fields} Extracted Fields", expanded=False):
                        for field in fields_list:
                            field_info = f"**{field['field_name']}** ({field.get('data_type', 'unknown')})"
                            if field.get('required'):
                                field_info += " *[Required]*"
                            if field.get('description'):
                                # Truncate very long descriptions
                                desc = field['description'][:150] + "..." if len(field['description']) > 150 else field['description']
                                field_info += f"\n   📝 {desc}"
                            if field.get('min_value') is not None or field.get('max_value') is not None:
                                field_info += f"\n   📊 Range: {field.get('min_value', 'N/A')} - {field.get('max_value', 'N/A')}"
                            if field.get('allowed_values'):
                                vals = field['allowed_values'][:8]  # Show first 8
                                vals_str = ', '.join(vals)
                                if len(field['allowed_values']) > 8:
                                    vals_str += f" ... +{len(field['allowed_values']) - 8} more"
                                field_info += f"\n   ✓ Allowed: {vals_str}"
                            st.markdown(field_info)

            # Demo dictionary selector below file uploader
            demo_dict = st.selectbox(
                "Or load demo dictionary:",
                ["None"] + list(DEMO_DICTIONARIES.keys()),
                key="demo_dict_selector"
            )

            demo_dictionary_built = None
            if demo_dict != "None":
                # Get demo dictionary CSV string and parse it
                demo_csv_string = get_demo_dictionary(demo_dict)

                # Parse CSV string into rules dictionary (same logic as CSV upload)
                import io
                df = pd.read_csv(io.StringIO(demo_csv_string))
                rules = {}
                for _, row in df.iterrows():
                    if 'Column' in row or 'column' in row or 'Field' in row or 'field' in row:
                        field_name = row.get('Column') or row.get('column') or row.get('Field') or row.get('field')
                        if field_name:
                            rule = {}
                            if 'Type' in row or 'type' in row:
                                rule['type'] = str(row.get('Type') or row.get('type'))
                            if 'Min' in row or 'min' in row:
                                rule['min'] = row.get('Min') or row.get('min')
                            if 'Max' in row or 'max' in row:
                                rule['max'] = row.get('Max') or row.get('max')
                            if 'Required' in row or 'required' in row:
                                rule['required'] = row.get('Required') or row.get('required')
                            if 'Allowed_Values' in row or 'allowed_values' in row:
                                allowed = row.get('Allowed_Values') or row.get('allowed_values')
                                if allowed and not pd.isna(allowed):
                                    rule['allowed_values'] = [v.strip() for v in str(allowed).split(',')]
                            rules[field_name] = rule

                demo_dictionary_built = {
                    "source": "Demo Dictionary",
                    "filename": demo_dict,
                    "rules": rules
                }
                st.success(f"✅ Loaded {demo_dict} ({len(rules)} field definitions)")

            # Centralized decision for what st.session_state.dictionary should be
            # this render pass (fixes stale-dictionary bug: resetting the demo
            # selector to "None" used to leave a previously-loaded demo
            # dictionary's rules active forever, since there was no `else` to
            # clear them). See resolve_effective_dictionary() docstring for the
            # full precedence rules (upload > demo selection > None).
            st.session_state.dictionary = resolve_effective_dictionary(
                demo_dict_selection=demo_dict,
                demo_dictionary_built=demo_dictionary_built,
                dict_file_uploaded=dict_file is not None,
                previous_session_dictionary=st.session_state.dictionary,
            )

            # Add cache management
            st.markdown("---")
            st.markdown("#### Cache Management")

            # Include legacy .pkl files so old pickle-based caches get cleaned up too
            cache_files = list(st.session_state.cache_dir.glob("*.json")) + list(st.session_state.cache_dir.glob("*.pkl"))
            num_cached = len(cache_files)

            if num_cached > 0:
                st.caption(f"📦 {num_cached} dictionaries cached")

                if st.button("🗑️ Clear All Cache", help="Delete all cached dictionaries to force re-parsing"):
                    try:
                        for cache_file in cache_files:
                            cache_file.unlink()
                        st.session_state.dict_cache = {}
                        st.session_state.dictionary = None
                        st.success(f"✅ Cleared {num_cached} cached dictionaries")
                        print(f"\n🗑️ CLEARED {num_cached} CACHE FILES\n")
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ Error clearing cache: {e}")
            else:
                st.caption("No cached dictionaries")

    with col3:
        # Export dropdown - moved from bottom
        if st.session_state.analysis_results:
            st.markdown('<div class="app-section-title">Export</div>', unsafe_allow_html=True)
            export_format = st.selectbox(
                "Choose format:",
                ["Select format to export", "Excel with highlighting", "JSON report"],
                key="export_format",
                on_change=lambda: None  # Trigger rerun on selection
            )

            # Show download button immediately when format is selected
            if export_format == "Excel with highlighting":
                excel_data = export_to_excel_with_highlighting(
                    st.session_state.data,
                    st.session_state.analysis_results['issues']
                )
                st.download_button(
                    label="📊 Download Excel",
                    data=excel_data,
                    file_name=f"data_quality_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    type="secondary"
                )
            elif export_format == "JSON report":
                json_str = json.dumps(st.session_state.analysis_results, indent=2, default=str)
                st.download_button(
                    label="📄 Download JSON",
                    data=json_str,
                    file_name=f"quality_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                    mime="application/json",
                    use_container_width=True,
                    type="secondary"
                )

    # Run Analysis - full width between uploads and results
    st.markdown("<div style='margin: 1.75rem 0 1rem 0; border-top: 1px solid #e2e8f0;'></div>", unsafe_allow_html=True)

    # Create centered column for button
    col_left, col_center, col_right = st.columns([1, 2, 1])
    with col_center:
        # Enable button only when data is loaded
        if st.button(
            "Analyze Data Quality",
            disabled=(st.session_state.data is None),
            use_container_width=True,
            type="primary",
            help="Load data first, then click to analyze" if st.session_state.data is None else "Run comprehensive quality checks on your data",
            key="run_analysis_main"
        ):
            if st.session_state.data is not None:
                # Log dictionary usage
                if st.session_state.dictionary:
                    dict_source = st.session_state.dictionary.get('source', 'Unknown')
                    dict_filename = st.session_state.dictionary.get('filename', 'Unknown')
                    num_rules = len(st.session_state.dictionary.get('rules', {}))
                    num_fields = len(st.session_state.dictionary.get('fields', []))

                    print("\n" + "="*80)
                    print("🔍 RUNNING DATA QUALITY ANALYSIS")
                    print(f"   Data: {len(st.session_state.data)} rows × {len(st.session_state.data.columns)} columns")
                    print(f"   Dictionary: {dict_filename} (source: {dict_source})")
                    print(f"   Rules: {num_rules}, Fields: {num_fields}")
                    print("="*80 + "\n")

                    st.info(f"📖 Using dictionary: **{dict_filename}** ({dict_source}) - {num_rules} rules, {num_fields} fields")
                else:
                    print("\n⚠️ RUNNING ANALYSIS WITHOUT DICTIONARY (auto-detection only)\n")
                    st.info("⚠️ No dictionary loaded - using auto-detection only")

                with st.spinner("🔍 Analyzing data quality... Please wait."):
                    try:
                        # Run analysis
                        results = asyncio.run(
                            st.session_state.mcp_client.analyze_data_quality(
                                st.session_state.data,
                                st.session_state.dictionary,
                                source_format=st.session_state.get('data_source_format')
                            )
                        )
                        st.session_state.analysis_results = results
                        st.success("✅ Analysis complete!")
                    except Exception as e:
                        st.error(f"❌ Analysis failed: {str(e)}")
                        import traceback
                        st.code(traceback.format_exc())

    st.markdown("<div style='margin: 1.75rem 0 1rem 0; border-top: 1px solid #e2e8f0;'></div>", unsafe_allow_html=True)

    # Display results if available
    if st.session_state.analysis_results:

        # Summary metrics
        st.subheader("Analysis Summary")
        summary = st.session_state.analysis_results['summary']

        # Create columns with space for heatmap
        col1, col2, col3, col4, col5, col6 = st.columns([1, 1, 1.2, 1, 1, 1.5])
        with col1:
            st.metric("Total Rows", f"{summary['total_rows']:,}")
        with col2:
            st.metric("Total Columns", summary['total_columns'])
        with col3:
            st.metric("Issues Found", summary['issues_found'],
                     delta=None if summary['issues_found'] == 0 else f"{summary['critical_issues']} critical")
        with col4:
            st.metric("Warnings", summary['warnings'])
        with col5:
            st.metric("Completeness", f"{summary['completeness']}%")
        with col6:
            # Create issue heatmap visualization
            st.markdown("#### Issue Map")
            create_issue_heatmap(st.session_state.data, st.session_state.analysis_results['issues'])

        # Issues details
        if st.session_state.analysis_results['issues']:
            st.subheader("Issues Found")

            # Group issues by type
            issues_by_type = {}
            for issue in st.session_state.analysis_results['issues']:
                issue_type = issue['type']
                if issue_type not in issues_by_type:
                    issues_by_type[issue_type] = []
                issues_by_type[issue_type].append(issue)

            # Display issues by type - collapsed by default for cleaner look
            for issue_type, issues in issues_by_type.items():
                # Collapse by default for Missing Values and Invalid Values
                expand_by_default = issue_type not in ['missing_values', 'invalid_value']
                with st.expander(f"{issue_type.replace('_', ' ').title()} ({len(issues)} issues)", expanded=expand_by_default):
                    for issue in issues[:10]:  # Show first 10
                        if issue['severity'] == 'error':
                            st.error(f"❌ {issue['message']}")
                        else:
                            st.warning(f"⚠️ {issue['message']}")
                    if len(issues) > 10:
                        st.info(f"... and {len(issues) - 10} more")

        # Recommendations
        if st.session_state.analysis_results['recommendations']:
            st.subheader("Recommendations")
            for rec in st.session_state.analysis_results['recommendations']:
                if rec['priority'] == 'critical':
                    st.error(f"🔴 **{rec['priority'].upper()}**: {rec['message']}")
                elif rec['priority'] == 'high':
                    st.warning(f"🟡 **{rec['priority'].upper()}**: {rec['message']}")
                else:
                    st.info(f"🔵 **{rec['priority'].upper()}**: {rec['message']}")

elif nav_choice == "About":
    st.title("About Data Quality Analyzer")

    st.markdown("""
    ### Purpose
    The Data Quality Analyzer is a tool designed to help you identify and resolve data quality issues
    in your datasets. It performs comprehensive checks to ensure your data meets quality standards.

    ### Features
    - **Multiple Format Support**: CSV, JSON, Excel (XLSX/XLS), and TXT files
    - **Automatic Issue Detection**: Missing values, invalid entries, range violations
    - **Custom Validation Rules**: Define your own business rules via data dictionaries (JSON or PDF)
    - **Visual Reporting**: Clear metrics and issue summaries with interactive heatmaps
    - **Excel Export**: Highlighted cells showing exact error locations with comments
    - **Demo Data**: Built-in datasets for testing various validation scenarios
    - **Dictionary Caching**: Fast PDF dictionary parsing with automatic caching

    ### What We Check
    1. **Missing Values**: Identifies null or empty cells
    2. **Invalid Values**: Detects entries like "invalid", "error", "n/a"
    3. **Data Type Validation**: Ensures values match expected types
    4. **Range Validation**: Checks if numeric values fall within specified ranges
    5. **Suspicious Values**: Flags test data or anomalous entries
    6. **Completeness**: Overall data completeness percentage

    ### How to Use
    1. **Upload your data** using the file uploader or select demo data
    2. **Optionally add a dictionary** (JSON or PDF) to define custom validation rules
    3. **Click Analyze** to run the quality checks
    4. **Review the results** including issues, recommendations, and visual heatmap
    5. **Export findings** to Excel (with highlighting) or JSON for further analysis

    ### Technical Details
    Built with Streamlit and powered by the Model Context Protocol (MCP) for
    advanced data analysis capabilities. Features include:
    - Interactive Plotly visualizations
    - PDF parsing with pypdf
    - Excel generation with cell highlighting and comments
    - Efficient dictionary caching system

    ### Data Flow Architecture
    """)

    # Load and render the Mermaid diagrams with selector
    try:
        # Load both diagrams
        with open('assets/data_flow_simple.mmd', 'r') as f:
            simple_diagram = f.read()
        with open('assets/data_flow_diagram.mmd', 'r') as f:
            detailed_diagram = f.read()

        st.info("Interactive flowcharts showing the data analysis pipeline:")

        # Add diagram selector within the content area
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            diagram_view = st.radio(
                "Select diagram complexity:",
                ["Simple View", "Detailed View"],
                horizontal=True,
                label_visibility="visible"
            )

        # Render the selected diagram
        if diagram_view == "Simple View":
            render_mermaid(simple_diagram, height=400)
        else:
            render_mermaid(detailed_diagram, height=700)

    except FileNotFoundError as e:
        st.info(f"Data flow diagram not found: {str(e)}")
    except Exception as e:
        st.error(f"Error rendering diagram: {str(e)}")

    with st.expander("Programmatic access (API)", expanded=False):
        st.markdown(
            "For pipelines and scripts, the REST API is the intended interface "
            "instead of this UI - it requires **(1)** the Duke VPN and **(2)** an "
            "API key sent via the `X-API-Key` header. Request the API key from "
            "the app owner, then `export DATA_ANALYZER_API_KEY=...` - the "
            "examples below work unchanged. Requests are rate-limited."
        )
        st.markdown("Key-less access via your Duke (Entra ID) sign-in is planned.")

        # Fill the example with the RUNTIME base URL when available, so this is
        # copy-paste-ready rather than a templated placeholder. The URL is not
        # secret (VPN + API key are the actual access controls), but the key
        # value itself is never read into a displayed string below - only a
        # literal `$DATA_ANALYZER_API_KEY` / `os.environ[...]` placeholder is
        # shown, so nothing here can leak the real key.
        _runtime_api_url = os.getenv("DATA_ANALYZER_API_URL", "")

        _placeholder_url = "https://<data-analyzer-api-fqdn>"

        _example_url = _runtime_api_url or _placeholder_url
        _docker_hint_comment = ""
        # "://api:" is the docker-compose service hostname (see
        # docker-compose.fullstack.yml) - it resolves inside the container
        # network only. When the running UI sees that value it's still the
        # *correct* value for server-to-server calls, but useless to a human
        # copy-pasting the example onto their own machine, so swap in
        # localhost for display purposes and explain why.
        if "://api:" in _example_url:
            _example_url = "http://localhost:8000"
            _docker_hint_comment = "\n# running via docker-compose.fullstack.yml - use localhost:8000 from your machine"

        st.markdown("**Python example**")
        st.code(
            f'''import os
import requests

# Base URL is a placeholder for local/dev use below; in production the API is
# only reachable over the Duke VPN (internal/VPN-only Azure FQDN). The API key
# is never hardcoded - it must be set in your shell environment.{_docker_hint_comment}
API_URL = os.environ.get("DATA_ANALYZER_API_URL", "{_example_url}")
API_KEY = os.environ["DATA_ANALYZER_API_KEY"]  # request this from the app owner

headers = {{"X-API-Key": API_KEY}}

# 1. Health check (no auth required)
resp = requests.get(f"{{API_URL}}/api/v1/health", timeout=10)
resp.raise_for_status()
print(resp.json())

# 2. Analyze a dataset (CSV or Excel), with optional schema/rules
with open("my_data.csv", "rb") as f:
    files = {{"data_file": ("my_data.csv", f, "text/csv")}}
    data = {{
        "schema": '{{"age": "int"}}',        # optional, JSON-encoded
        "rules": '{{"age": {{"min": 0, "max": 120}}}}',  # optional, JSON-encoded
        "min_rows": "1",                   # optional, default 1
    }}
    resp = requests.post(
        f"{{API_URL}}/api/v1/analyze",
        headers=headers,
        files=files,
        data=data,
        timeout=60,
    )
resp.raise_for_status()
result = resp.json()
print(result["summary"])
''',
            language="python",
        )

        st.markdown("**curl equivalent**")
        st.markdown(
            "Set `DATA_ANALYZER_API_KEY` in your shell first, then create a "
            "small sample file and call the API - both commands are complete, "
            "single-line commands (no `\\` line continuations, which is what "
            "broke the earlier version of this example)."
        )
        st.code(
            'printf "age,name\\n30,Jane\\n" > demo.csv',
            language="bash",
        )
        st.code(
            f'curl -sS -X POST "{_example_url}/api/v1/analyze" -H "X-API-Key: $DATA_ANALYZER_API_KEY" -F "data_file=@demo.csv;type=text/csv"',
            language="bash",
        )

        st.markdown("**Test the API**")
        st.markdown(
            f"- **Paste in a browser** (works in incognito): "
            f"[`{_example_url}/api/v1/health`]({_example_url}/api/v1/health) returns JSON; "
            f"[`{_example_url}/api/v1/docs`]({_example_url}/api/v1/docs) is an interactive "
            f"Swagger UI where you can execute `/analyze` directly from the browser.\n"
            f"- **curl one-liner** for `/analyze` - see the curl example above.\n"
            f"- Incognito/private browsing only avoids sending cookies - it does **not** "
            f"simulate being off the VPN. To test the VPN requirement itself, use a "
            f"device that is actually off-network (e.g. a phone on cellular data, wifi off)."
        )

    st.markdown(f"""
    ---
    *Version {__version__} - Enhanced UI with PDF Dictionary Support*
    """)
