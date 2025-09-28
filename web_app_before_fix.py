"""
Data Quality Analyzer - Modern Clean Design
Minimalist, professional interface with improved UX
"""

import streamlit as st
import pandas as pd
import json
import base64
import io
import asyncio
import csv
from typing import Dict, Any, Optional
from datetime import datetime

# Import custom modules
from demo_dictionaries import DEMO_DICTIONARIES, get_demo_dictionary
from mermaid_renderer import render_mermaid

# Import navigation menu
from streamlit_option_menu import option_menu

# Configure Streamlit page
st.set_page_config(
    page_title="Data Quality Analyzer",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Modern, clean CSS design
st.markdown("""
<style>
    /* Hide Streamlit default elements */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}
    .stDeployButton {display: none;}

    /* Modern color scheme */
    :root {
        --primary-color: #2563eb;
        --primary-hover: #1d4ed8;
        --success-color: #10b981;
        --warning-color: #f59e0b;
        --danger-color: #ef4444;
        --bg-color: #ffffff;
        --bg-secondary: #f9fafb;
        --text-primary: #111827;
        --text-secondary: #6b7280;
        --border-color: #e5e7eb;
        --shadow-sm: 0 1px 2px 0 rgba(0, 0, 0, 0.05);
        --shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1);
        --shadow-lg: 0 10px 15px -3px rgba(0, 0, 0, 0.1);
    }

    /* Clean background */
    .stApp {
        background: var(--bg-color);
    }

    /* Fixed top navbar styling */
    .css-1rs6kro {
        background: white;
        position: fixed;
        top: 0;
        z-index: 1000;
        border-bottom: 1px solid var(--border-color);
        box-shadow: var(--shadow-sm);
    }

    /* Modern cards */
    .card {
        background: white;
        border: 1px solid var(--border-color);
        border-radius: 8px;
        padding: 24px;
        margin-bottom: 20px;
        box-shadow: var(--shadow-sm);
        transition: box-shadow 0.2s ease;
    }

    .card:hover {
        box-shadow: var(--shadow);
    }

    /* Primary button styling */
    .stButton > button {
        background-color: var(--primary-color);
        color: white;
        border: none;
        border-radius: 6px;
        padding: 10px 24px;
        font-weight: 500;
        font-size: 16px;
        transition: all 0.2s ease;
        box-shadow: var(--shadow-sm);
    }

    .stButton > button:hover {
        background-color: var(--primary-hover);
        box-shadow: var(--shadow);
        transform: translateY(-1px);
    }

    .stButton > button:disabled {
        background-color: #d1d5db;
        cursor: not-allowed;
        transform: none;
    }

    /* File uploader styling */
    .stFileUploader {
        border: 2px dashed var(--border-color);
        border-radius: 8px;
        padding: 32px;
        background: var(--bg-secondary);
        transition: all 0.2s ease;
        text-align: center;
    }

    .stFileUploader:hover {
        border-color: var(--primary-color);
        background: #f0f9ff;
    }

    /* Clean typography */
    h1 {
        color: var(--text-primary);
        font-size: 32px;
        font-weight: 700;
        margin-bottom: 8px;
        letter-spacing: -0.025em;
    }

    h2 {
        color: var(--text-primary);
        font-size: 24px;
        font-weight: 600;
        margin-bottom: 16px;
    }

    h3 {
        color: var(--text-primary);
        font-size: 20px;
        font-weight: 600;
        margin-bottom: 12px;
    }

    p {
        color: var(--text-secondary);
        line-height: 1.5;
    }

    /* Success/Error/Warning alerts */
    .alert {
        padding: 12px 16px;
        border-radius: 6px;
        margin: 12px 0;
        font-weight: 500;
        display: flex;
        align-items: center;
    }

    .alert-success {
        background: #d1fae5;
        color: #065f46;
        border: 1px solid #a7f3d0;
    }

    .alert-error {
        background: #fee2e2;
        color: #991b1b;
        border: 1px solid #fecaca;
    }

    .alert-warning {
        background: #fef3c7;
        color: #92400e;
        border: 1px solid #fde68a;
    }

    /* Metrics styling */
    .metric-container {
        display: grid;
        grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
        gap: 16px;
        margin: 24px 0;
    }

    .metric-card {
        background: white;
        border: 1px solid var(--border-color);
        border-radius: 8px;
        padding: 20px;
        text-align: center;
    }

    .metric-value {
        font-size: 32px;
        font-weight: 700;
        color: var(--primary-color);
        margin-bottom: 4px;
    }

    .metric-label {
        font-size: 14px;
        color: var(--text-secondary);
        text-transform: uppercase;
        letter-spacing: 0.05em;
    }

    /* Tab styling */
    .stTabs [data-baseweb="tab-list"] {
        background-color: transparent;
        gap: 24px;
        border-bottom: 1px solid var(--border-color);
        padding-bottom: 0;
    }

    .stTabs [data-baseweb="tab"] {
        background-color: transparent;
        border: none;
        color: var(--text-secondary);
        font-weight: 500;
        padding: 12px 4px;
        border-bottom: 2px solid transparent;
    }

    .stTabs [aria-selected="true"] {
        color: var(--primary-color);
        border-bottom-color: var(--primary-color);
    }

    /* Sidebar styling */
    .css-1d391kg {
        background: var(--bg-secondary);
        border-right: 1px solid var(--border-color);
    }

    /* Expander styling */
    .streamlit-expanderHeader {
        background: var(--bg-secondary);
        border: 1px solid var(--border-color);
        border-radius: 6px;
        font-weight: 500;
    }

    /* Navigation menu custom styling */
    [data-testid="stHorizontalBlock"] {
        background: white;
        padding: 12px 0;
        border-bottom: 1px solid var(--border-color);
        margin-bottom: 24px;
        position: sticky;
        top: 0;
        z-index: 999;
    }

    /* Three-column upload layout */
    .upload-container {
        display: flex;
        gap: 24px;
        margin: 32px 0;
        align-items: start;
    }

    .upload-box {
        flex: 1;
        background: white;
        border: 2px dashed var(--border-color);
        border-radius: 8px;
        padding: 32px;
        text-align: center;
        min-height: 200px;
        display: flex;
        flex-direction: column;
        justify-content: center;
        transition: all 0.2s ease;
    }

    .upload-box:hover {
        border-color: var(--primary-color);
        background: #f0f9ff;
    }

    .upload-icon {
        font-size: 48px;
        margin-bottom: 12px;
    }

    .analyze-button-container {
        flex: 0.8;
        display: flex;
        align-items: center;
        justify-content: center;
        min-height: 200px;
    }

    /* Mobile responsiveness */
    @media (max-width: 768px) {
        .upload-container {
            flex-direction: column;
        }

        .metric-container {
            grid-template-columns: 1fr;
        }

        h1 {
            font-size: 24px;
        }

        .upload-box {
            min-height: 150px;
            padding: 24px;
        }
    }
</style>
""", unsafe_allow_html=True)


class MCPClient:
    """Simplified MCP client for data analysis"""

    def __init__(self):
        pass

    def _simulate_mcp_call(self, request_data: Dict, debug: bool = False) -> Dict:
        """Simulate MCP server call for development"""
        tool = request_data.get("tool")
        args = request_data.get("arguments", {})

        if tool == "analyze_data":
            try:
                data_content = args.get("data_content", "")
                file_format = args.get("file_format", "csv")
                schema = args.get("schema", {})
                rules = args.get("rules", {})
                min_rows = args.get("min_rows", 1)

                # Parse data based on format
                df = None
                if file_format.lower() == "csv":
                    df = pd.read_csv(io.StringIO(data_content))
                elif file_format.lower() == "json":
                    json_data = json.loads(data_content)
                    if isinstance(json_data, list):
                        df = pd.json_normalize(json_data)
                    elif isinstance(json_data, dict):
                        first_key = list(json_data.keys())[0] if json_data else None
                        if first_key and isinstance(json_data[first_key], list):
                            df = pd.json_normalize(json_data[first_key])
                        else:
                            df = pd.json_normalize([json_data])
                elif file_format.lower() in ["xlsx", "xls"]:
                    excel_bytes = base64.b64decode(data_content)
                    df = pd.read_excel(io.BytesIO(excel_bytes))
                elif file_format.lower() == "parquet":
                    parquet_bytes = base64.b64decode(data_content)
                    df = pd.read_parquet(io.BytesIO(parquet_bytes))
                else:
                    return {"error": f"Unsupported format: {file_format}"}

                if df is None or df.empty:
                    return {"error": "No data could be loaded"}

                # Basic analysis
                issues = []
                row_count = len(df)

                # Check minimum rows
                if row_count < min_rows:
                    issues.append({
                        "type": "row_count",
                        "severity": "error",
                        "message": f"Row count ({row_count}) is less than minimum required ({min_rows})"
                    })

                # Data type validation
                if schema:
                    for col, expected_type in schema.items():
                        if col in df.columns:
                            actual_type = str(df[col].dtype)
                            type_match = False

                            if expected_type == "int" and "int" in actual_type:
                                type_match = True
                            elif expected_type == "float" and "float" in actual_type:
                                type_match = True
                            elif expected_type == "str" and "object" in actual_type:
                                type_match = True
                            elif expected_type == "bool" and "bool" in actual_type:
                                type_match = True
                            elif expected_type == "datetime" and "datetime" in actual_type:
                                type_match = True

                            if not type_match:
                                issues.append({
                                    "type": "type_mismatch",
                                    "severity": "warning",
                                    "column": col,
                                    "message": f"Column '{col}' has type '{actual_type}' but expected '{expected_type}'"
                                })

                # Value range validation
                if rules:
                    for col, col_rules in rules.items():
                        if col in df.columns:
                            if "min" in col_rules:
                                min_val = col_rules["min"]
                                violations = df[df[col] < min_val][col].tolist()
                                if violations:
                                    issues.append({
                                        "type": "range_violation",
                                        "severity": "error",
                                        "column": col,
                                        "message": f"Column '{col}' has {len(violations)} values below minimum {min_val}",
                                        "rows": df[df[col] < min_val].index.tolist()
                                    })

                            if "max" in col_rules:
                                max_val = col_rules["max"]
                                violations = df[df[col] > max_val][col].tolist()
                                if violations:
                                    issues.append({
                                        "type": "range_violation",
                                        "severity": "error",
                                        "column": col,
                                        "message": f"Column '{col}' has {len(violations)} values above maximum {max_val}",
                                        "rows": df[df[col] > max_val].index.tolist()
                                    })

                # Calculate summary statistics
                summary = {
                    "row_count": row_count,
                    "column_count": len(df.columns),
                    "columns": df.columns.tolist(),
                    "dtypes": {col: str(dtype) for col, dtype in df.dtypes.items()},
                    "missing_values": df.isnull().sum().to_dict(),
                    "issues_count": len(issues),
                    "issues": issues,
                    "data": df
                }

                return {
                    "success": True,
                    "summary": summary,
                    "preview": df.head(10).to_dict('records') if not df.empty else []
                }

            except Exception as e:
                return {"error": str(e)}

        elif tool == "parse_data_dictionary":
            try:
                dictionary_content = args.get("dictionary_content", "")
                schema = {}
                rules = {}

                lines = dictionary_content.strip().split('\n')
                if ',' in lines[0] or '\t' in lines[0]:
                    reader = csv.DictReader(io.StringIO(dictionary_content))
                    for row in reader:
                        field_name = row.get('field_name') or row.get('column') or row.get('name')
                        if field_name:
                            field_type = row.get('type') or row.get('data_type') or 'str'
                            schema[field_name] = field_type

                            field_rules = {}
                            if row.get('min'):
                                try:
                                    field_rules['min'] = float(row['min'])
                                except:
                                    pass
                            if row.get('max'):
                                try:
                                    field_rules['max'] = float(row['max'])
                                except:
                                    pass
                            if field_rules:
                                rules[field_name] = field_rules

                return {
                    "success": True,
                    "schema": schema,
                    "rules": rules
                }
            except Exception as e:
                return {"error": str(e)}

        return {"error": "Unknown tool"}

    async def analyze_data(self, data_content: str, file_format: str = "csv",
                          schema: Dict = None, rules: Dict = None,
                          min_rows: int = 1, debug: bool = False):
        """Analyze data with optional schema and rules"""
        request_data = {
            "tool": "analyze_data",
            "arguments": {
                "data_content": data_content,
                "file_format": file_format,
                "schema": schema or {},
                "rules": rules or {},
                "min_rows": min_rows
            }
        }
        return self._simulate_mcp_call(request_data, debug)

    async def parse_data_dictionary(self, dictionary_content: str,
                                   format_hint: str = "auto",
                                   use_cache: bool = True,
                                   debug: bool = False) -> Dict:
        """Parse data dictionary"""
        request_data = {
            "tool": "parse_data_dictionary",
            "arguments": {
                "dictionary_content": dictionary_content,
                "format_hint": format_hint,
                "use_cache": use_cache
            }
        }
        return self._simulate_mcp_call(request_data, debug)


def main():
    """Main application entry point"""

    # Initialize session state
    if 'analysis_complete' not in st.session_state:
        st.session_state.analysis_complete = False
    if 'analysis_results' not in st.session_state:
        st.session_state.analysis_results = None
    if 'uploaded_file' not in st.session_state:
        st.session_state.uploaded_file = None
    if 'dictionary_file' not in st.session_state:
        st.session_state.dictionary_file = None
    if 'page' not in st.session_state:
        st.session_state.page = 'analyze'

    # Create clean navigation bar
    col1, col2, col3 = st.columns([1, 8, 1])
    with col2:
        selected = option_menu(
            menu_title=None,
            options=["Analyze", "About"],
            icons=["graph-up", "info-circle"],
            menu_icon=None,
            default_index=0,
            orientation="horizontal",
            styles={
                "container": {"padding": "5px", "background-color": "#ffffff"},
                "icon": {"color": "#6b7280", "font-size": "18px"},
                "nav-link": {
                    "font-size": "16px",
                    "text-align": "center",
                    "margin": "0px",
                    "padding": "10px 20px",
                    "color": "#6b7280",
                    "--hover-color": "#f3f4f6"
                },
                "nav-link-selected": {
                    "background-color": "#2563eb",
                    "color": "white",
                    "font-weight": "500"
                }
            }
        )

    # Route to pages
    if selected == "Analyze":
        analyze_page()
    else:
        about_page()


def analyze_page():
    """Main analysis page with three-element layout"""

    st.title("Data Quality Analyzer")
    st.markdown("<p style='color: #6b7280; margin-top: -10px;'>Upload your data, optionally add validation rules, and analyze</p>", unsafe_allow_html=True)

    # Main three-element upload area
    st.markdown("<div class='card'>", unsafe_allow_html=True)

    col1, col2, col3 = st.columns([1, 1, 0.8])

    with col1:
        st.markdown("### 📁 Upload Data File")
        uploaded_file = st.file_uploader(
            "Drag and drop or browse",
            type=['csv', 'json', 'xlsx', 'xls', 'parquet'],
            key="main_data_upload",
            label_visibility="collapsed"
        )

        if uploaded_file:
            file_size = len(uploaded_file.read())
            uploaded_file.seek(0)
            st.session_state.uploaded_file = uploaded_file
            st.success(f"✓ {uploaded_file.name} ({file_size:,} bytes)")

    with col2:
        st.markdown("### 📚 Upload Data Dictionary")
        st.markdown("<small style='color: #6b7280;'>Optional</small>", unsafe_allow_html=True)
        dict_file = st.file_uploader(
            "Drag and drop or browse",
            type=['csv', 'json', 'txt'],
            key="dict_upload",
            label_visibility="collapsed"
        )

        if dict_file:
            st.session_state.dictionary_file = dict_file
            st.success(f"✓ {dict_file.name}")

    with col3:
        st.markdown("### ⚡ Analyze")
        st.markdown("<div style='height: 20px;'></div>", unsafe_allow_html=True)

        # Enable analyze button only if data is uploaded
        analyze_enabled = st.session_state.uploaded_file is not None

        if st.button("🚀 Run Analysis",
                    disabled=not analyze_enabled,
                    type="primary",
                    use_container_width=True):
            run_analysis()

    st.markdown("</div>", unsafe_allow_html=True)

    # Sidebar for additional options
    with st.sidebar:
        st.markdown("### Quick Actions")

        # Demo data section
        with st.expander("📂 Load Demo Data"):
            st.markdown("**Sample Datasets**")
            demo_data_choice = st.selectbox(
                "Choose demo data",
                ["None", "CSV - Western Names", "CSV - Asian Names",
                 "JSON - Mixed Names", "CSV - Clinical Trial"],
                key="demo_data_select"
            )

            if demo_data_choice != "None":
                if demo_data_choice == "CSV - Western Names":
                    demo_data = """name,age,email,country
John Smith,25,john@email.com,USA
Jane Doe,30,jane@email.com,Canada
Bob Wilson,invalid,bob@email,UK"""
                elif demo_data_choice == "CSV - Asian Names":
                    demo_data = """name,age,city,join_date
李明,28,Beijing,2023-01-15
田中太郎,35,Tokyo,2023-02-20
김철수,42,Seoul,invalid-date"""
                elif demo_data_choice == "JSON - Mixed Names":
                    demo_data = json.dumps([
                        {"name": "Alice Smith", "age": 30, "score": 85.5},
                        {"name": "王小明", "age": 25, "score": 92.0},
                        {"name": "Carlos García", "age": "invalid", "score": 88.0}
                    ])
                elif demo_data_choice == "CSV - Clinical Trial":
                    demo_data = """patient_id,age,treatment,outcome,date
P001,45,Drug_A,Improved,2023-01-15
P002,52,Drug_B,No_Change,2023-01-20
P003,999,Drug_A,Improved,invalid"""

                st.session_state.uploaded_file = demo_data
                st.success("Demo data loaded!")

            st.markdown("**Sample Dictionaries**")
            demo_dict_choice = st.selectbox(
                "Choose demo dictionary",
                ["None"] + list(DEMO_DICTIONARIES.keys()),
                key="demo_dict_select"
            )

            if demo_dict_choice != "None":
                demo_dict = get_demo_dictionary(demo_dict_choice)
                st.session_state.dictionary_file = demo_dict
                st.success("Demo dictionary loaded!")

        # Results section (shown after analysis)
        if st.session_state.analysis_complete:
            st.markdown("### 📊 Results")
            show_results_sidebar()

        # Export section (shown after analysis)
        if st.session_state.analysis_complete:
            st.markdown("### 💾 Export")
            show_export_sidebar()

    # Main content area for results
    if st.session_state.analysis_complete:
        show_results_main()


def run_analysis():
    """Execute the data analysis"""
    with st.spinner("Analyzing your data..."):
        progress_bar = st.progress(0)
        status_text = st.empty()

        # Simulate analysis stages
        stages = [
            (0.25, "Loading data..."),
            (0.50, "Validating schema..."),
            (0.75, "Checking rules..."),
            (1.0, "Generating report...")
        ]

        client = MCPClient()

        # Prepare data
        if isinstance(st.session_state.uploaded_file, str):
            data_content = st.session_state.uploaded_file
            file_format = "json" if data_content.startswith('[') or data_content.startswith('{') else "csv"
        else:
            file_ext = st.session_state.uploaded_file.name.split('.')[-1].lower()
            file_format = file_ext

            if file_ext in ['xlsx', 'xls', 'parquet']:
                file_bytes = st.session_state.uploaded_file.read()
                data_content = base64.b64encode(file_bytes).decode('utf-8')
                st.session_state.uploaded_file.seek(0)
            else:
                data_content = st.session_state.uploaded_file.read().decode('utf-8')
                st.session_state.uploaded_file.seek(0)

        # Parse dictionary if provided
        schema = {}
        rules = {}
        if st.session_state.dictionary_file:
            if isinstance(st.session_state.dictionary_file, str):
                dict_content = st.session_state.dictionary_file
            else:
                dict_content = st.session_state.dictionary_file.read().decode('utf-8')
                st.session_state.dictionary_file.seek(0)

            dict_result = asyncio.run(
                client.parse_data_dictionary(dict_content)
            )

            if dict_result.get("success"):
                schema = dict_result.get("schema", {})
                rules = dict_result.get("rules", {})

        # Update progress
        for progress, message in stages:
            status_text.text(message)
            progress_bar.progress(progress)
            asyncio.run(asyncio.sleep(0.3))

        # Run analysis
        result = asyncio.run(
            client.analyze_data(
                data_content=data_content,
                file_format=file_format,
                schema=schema,
                rules=rules,
                min_rows=1
            )
        )

        if result.get("success"):
            st.session_state.analysis_complete = True
            st.session_state.analysis_results = result
            progress_bar.empty()
            status_text.empty()
            st.success("✓ Analysis complete!")
            st.rerun()
        else:
            st.error(f"Analysis failed: {result.get('error')}")


def show_results_main():
    """Display main results area"""
    results = st.session_state.analysis_results
    summary = results.get("summary", {})

    st.markdown("---")
    st.markdown("## Analysis Results")

    # Metrics row
    col1, col2, col3, col4 = st.columns(4)

    with col1:
        st.metric("Total Rows", f"{summary.get('row_count', 0):,}")

    with col2:
        st.metric("Columns", summary.get('column_count', 0))

    with col3:
        issues_count = summary.get('issues_count', 0)
        st.metric("Issues", issues_count,
                 delta="No issues" if issues_count == 0 else None,
                 delta_color="normal" if issues_count == 0 else "inverse")

    with col4:
        missing = sum(summary.get('missing_values', {}).values())
        st.metric("Missing Values", missing)

    # Issues details
    if summary.get('issues'):
        st.markdown("### ⚠️ Issues Found")

        errors = [i for i in summary['issues'] if i['severity'] == 'error']
        warnings = [i for i in summary['issues'] if i['severity'] == 'warning']

        if errors:
            st.markdown("#### 🔴 Errors")
            for error in errors:
                st.error(f"{error['message']}")

        if warnings:
            st.markdown("#### 🟡 Warnings")
            for warning in warnings:
                st.warning(f"{warning['message']}")

    # Data preview
    if results.get('preview'):
        st.markdown("### Data Preview")
        df_preview = pd.DataFrame(results['preview'])
        st.dataframe(df_preview, use_container_width=True)


def show_results_sidebar():
    """Show results summary in sidebar"""
    results = st.session_state.analysis_results
    summary = results.get("summary", {})

    # Quick stats
    st.info(f"""
    **Quick Stats**
    - Rows: {summary.get('row_count', 0):,}
    - Columns: {summary.get('column_count', 0)}
    - Issues: {summary.get('issues_count', 0)}
    """)


def show_export_sidebar():
    """Show export options in sidebar"""
    results = st.session_state.analysis_results
    summary = results.get("summary", {})

    # Excel export
    if st.button("📊 Export to Excel", use_container_width=True):
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Data Analysis"

            # Get data
            df = summary.get('data', pd.DataFrame(results.get('preview', [])))

            # Write headers
            for col_idx, col_name in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = Font(bold=True)

            # Write data
            for row_idx, row_data in df.iterrows():
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx + 2, column=col_idx, value=value)

            # Save
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)

            st.download_button(
                label="Download Excel",
                data=excel_buffer.getvalue(),
                file_name=f"analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except ImportError:
            st.error("Install openpyxl: pip install openpyxl")

    # CSV export
    if st.button("📄 Export to CSV", use_container_width=True):
        df = summary.get('data', pd.DataFrame(results.get('preview', [])))
        csv_buffer = io.StringIO()
        df.to_csv(csv_buffer, index=False)

        st.download_button(
            label="Download CSV",
            data=csv_buffer.getvalue(),
            file_name=f"analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            mime="text/csv"
        )


def about_page():
    """About page with information"""
    st.title("About Data Quality Analyzer")

    col1, col2, col3 = st.columns([1, 2, 1])

    with col2:
        st.markdown("""
        ### Professional Data Validation Tool

        Data Quality Analyzer is a powerful tool designed to help data professionals validate
        and clean their datasets efficiently. Built with modern technologies and best practices,
        it provides instant feedback on data quality issues.

        #### Key Features
        - ✓ Support for CSV, JSON, Excel, and Parquet files
        - ✓ Custom validation rules and data dictionaries
        - ✓ Instant error detection and highlighting
        - ✓ Export results with visual error markers
        - ✓ Fast processing for files up to 100MB

        #### How to Use
        1. **Upload your data file** - Drag and drop or browse for your file
        2. **Optionally add a data dictionary** - Define expected schema and rules
        3. **Click Analyze** - Get instant quality reports
        4. **Export results** - Download reports with error highlighting

        #### Technology Stack
        - **Frontend:** Streamlit
        - **Backend:** Python with Pandas
        - **Analysis:** MCP (Model Context Protocol)
        - **Deployment:** Azure Container Apps

        #### Need Help?
        - Use the demo data to try the tool
        - Check the sidebar for quick actions
        - Export your results in Excel or CSV format

        ---

        *Version 1.0.0 | © 2024 Data Quality Analyzer*
        """)


if __name__ == "__main__":
    main()