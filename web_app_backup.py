"""
Minimalist Azure Web Application for Data Analysis
Simple interface with 3 core elements: Upload Data, Upload Dictionary (optional), Run Analysis
All advanced options hidden in collapsible sections
"""

import streamlit as st
import pandas as pd
import json
import base64
import io
import asyncio
import csv
from typing import Dict, Any, Optional
from datetime import datetime

# Import custom modules
from demo_dictionaries import DEMO_DICTIONARIES, get_demo_dictionary
from mermaid_renderer import render_mermaid

# Import navigation menu
from streamlit_option_menu import option_menu

# Configure Streamlit page
st.set_page_config(
    page_title="Data Quality Analyzer",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed"  # Start with sidebar collapsed for minimal look
)

# Custom CSS for minimal, clean design
st.markdown("""
<style>
    /* Clean, minimal styling */
    .stApp {
        background: linear-gradient(135deg, #f5f7fa 0%, #c3cfe2 100%);
    }

    /* Card-like containers */
    .main-card {
        background: white;
        border-radius: 10px;
        padding: 2rem;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        margin-bottom: 2rem;
    }

    /* Primary action button styling */
    .stButton > button[kind="primary"] {
        background-color: #4CAF50;
        color: white;
        font-size: 1.2rem;
        padding: 0.75rem 2rem;
        border-radius: 5px;
        width: 100%;
        margin-top: 1rem;
    }

    /* File uploader styling */
    .stFileUploader {
        border: 2px dashed #ccc;
        border-radius: 10px;
        padding: 1rem;
    }

    /* Hide Streamlit branding */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}

    /* Center main content */
    .block-container {
        max-width: 1200px;
        padding-top: 2rem;
    }

    /* Success/Error message styling */
    .success-msg {
        background-color: #d4edda;
        border-color: #c3e6cb;
        color: #155724;
        padding: 1rem;
        border-radius: 5px;
        margin: 1rem 0;
    }

    .error-msg {
        background-color: #f8d7da;
        border-color: #f5c6cb;
        color: #721c24;
        padding: 1rem;
        border-radius: 5px;
        margin: 1rem 0;
    }

    /* Minimal header */
    h1 {
        text-align: center;
        color: #2c3e50;
        font-weight: 300;
        margin-bottom: 0.5rem;
    }

    .subtitle {
        text-align: center;
        color: #7f8c8d;
        margin-bottom: 2rem;
    }
</style>
""", unsafe_allow_html=True)


class MCPClient:
    """Simplified MCP client for data analysis"""

    def __init__(self):
        pass

    def _simulate_mcp_call(self, request_data: Dict, debug: bool = False) -> Dict:
        """Simulate MCP server call for development"""
        tool = request_data.get("tool")
        args = request_data.get("arguments", {})

        if tool == "analyze_data":
            try:
                data_content = args.get("data_content", "")
                file_format = args.get("file_format", "csv")
                schema = args.get("schema", {})
                rules = args.get("rules", {})
                min_rows = args.get("min_rows", 1)

                # Parse data based on format
                df = None
                if file_format.lower() == "csv":
                    df = pd.read_csv(io.StringIO(data_content))
                elif file_format.lower() == "json":
                    json_data = json.loads(data_content)
                    if isinstance(json_data, list):
                        df = pd.json_normalize(json_data)
                    elif isinstance(json_data, dict):
                        first_key = list(json_data.keys())[0] if json_data else None
                        if first_key and isinstance(json_data[first_key], list):
                            df = pd.json_normalize(json_data[first_key])
                        else:
                            df = pd.json_normalize([json_data])
                elif file_format.lower() in ["xlsx", "xls"]:
                    excel_bytes = base64.b64decode(data_content)
                    df = pd.read_excel(io.BytesIO(excel_bytes))
                elif file_format.lower() == "parquet":
                    parquet_bytes = base64.b64decode(data_content)
                    df = pd.read_parquet(io.BytesIO(parquet_bytes))
                else:
                    return {"error": f"Unsupported format: {file_format}"}

                if df is None or df.empty:
                    return {"error": "No data could be loaded"}

                # Basic analysis
                issues = []
                row_count = len(df)

                # Check minimum rows
                if row_count < min_rows:
                    issues.append({
                        "type": "row_count",
                        "severity": "error",
                        "message": f"Row count ({row_count}) is less than minimum required ({min_rows})"
                    })

                # Data type validation
                if schema:
                    for col, expected_type in schema.items():
                        if col in df.columns:
                            actual_type = str(df[col].dtype)
                            type_match = False

                            if expected_type == "int" and "int" in actual_type:
                                type_match = True
                            elif expected_type == "float" and "float" in actual_type:
                                type_match = True
                            elif expected_type == "str" and "object" in actual_type:
                                type_match = True
                            elif expected_type == "bool" and "bool" in actual_type:
                                type_match = True
                            elif expected_type == "datetime" and "datetime" in actual_type:
                                type_match = True

                            if not type_match:
                                issues.append({
                                    "type": "type_mismatch",
                                    "severity": "warning",
                                    "column": col,
                                    "message": f"Column '{col}' has type '{actual_type}' but expected '{expected_type}'"
                                })

                # Value range validation
                if rules:
                    for col, col_rules in rules.items():
                        if col in df.columns:
                            if "min" in col_rules:
                                min_val = col_rules["min"]
                                violations = df[df[col] < min_val][col].tolist()
                                if violations:
                                    issues.append({
                                        "type": "range_violation",
                                        "severity": "error",
                                        "column": col,
                                        "message": f"Column '{col}' has {len(violations)} values below minimum {min_val}"
                                    })

                            if "max" in col_rules:
                                max_val = col_rules["max"]
                                violations = df[df[col] > max_val][col].tolist()
                                if violations:
                                    issues.append({
                                        "type": "range_violation",
                                        "severity": "error",
                                        "column": col,
                                        "message": f"Column '{col}' has {len(violations)} values above maximum {max_val}"
                                    })

                # Calculate summary statistics
                summary = {
                    "row_count": row_count,
                    "column_count": len(df.columns),
                    "columns": df.columns.tolist(),
                    "dtypes": {col: str(dtype) for col, dtype in df.dtypes.items()},
                    "missing_values": df.isnull().sum().to_dict(),
                    "issues_count": len(issues),
                    "issues": issues
                }

                return {
                    "success": True,
                    "summary": summary,
                    "preview": df.head(10).to_dict('records') if not df.empty else []
                }

            except Exception as e:
                return {"error": str(e)}

        elif tool == "parse_data_dictionary":
            # Simulate dictionary parsing with actual logic
            try:
                dictionary_content = args.get("dictionary_content", "")
                format_hint = args.get("format_hint", "auto")

                # Parse dictionary based on format
                schema = {}
                rules = {}

                # Try to detect format
                lines = dictionary_content.strip().split('\n')
                if not lines:
                    return {"error": "Empty dictionary"}

                # Check if it's CSV format (common data dictionary format)
                if ',' in lines[0] or '\t' in lines[0]:
                    # Parse as CSV-like dictionary
                    import csv
                    reader = csv.DictReader(io.StringIO(dictionary_content))

                    for row in reader:
                        # Common dictionary formats have columns like:
                        # field_name, field_type, min, max, allowed_values, description
                        field_name = (row.get('field_name') or row.get('column') or
                                    row.get('variable') or row.get('name') or '')

                        if field_name:
                            # Extract type
                            field_type = (row.get('type') or row.get('field_type') or
                                        row.get('data_type') or 'str')

                            # Map common type names
                            type_map = {
                                'integer': 'int', 'number': 'float', 'string': 'str',
                                'text': 'str', 'boolean': 'bool', 'date': 'datetime',
                                'timestamp': 'datetime', 'numeric': 'float'
                            }
                            field_type = type_map.get(field_type.lower(), field_type.lower())
                            schema[field_name] = field_type

                            # Extract rules
                            field_rules = {}
                            if row.get('min'):
                                try:
                                    field_rules['min'] = float(row['min'])
                                except:
                                    pass
                            if row.get('max'):
                                try:
                                    field_rules['max'] = float(row['max'])
                                except:
                                    pass
                            if row.get('allowed_values') or row.get('values'):
                                allowed = row.get('allowed_values') or row.get('values')
                                field_rules['allowed_values'] = [v.strip() for v in allowed.split('|')]

                            if field_rules:
                                rules[field_name] = field_rules

                elif lines[0].startswith('{') or lines[0].startswith('['):
                    # Parse as JSON dictionary
                    dict_data = json.loads(dictionary_content)

                    if isinstance(dict_data, list):
                        # Array of field definitions
                        for field in dict_data:
                            if isinstance(field, dict):
                                name = field.get('name') or field.get('field_name')
                                if name:
                                    dtype = field.get('type', 'str')
                                    schema[name] = dtype

                                    field_rules = {}
                                    if 'min' in field:
                                        field_rules['min'] = field['min']
                                    if 'max' in field:
                                        field_rules['max'] = field['max']
                                    if 'allowed_values' in field:
                                        field_rules['allowed_values'] = field['allowed_values']

                                    if field_rules:
                                        rules[name] = field_rules

                    elif isinstance(dict_data, dict):
                        # Object with schema/rules keys or field definitions
                        if 'schema' in dict_data:
                            schema = dict_data['schema']
                        if 'rules' in dict_data:
                            rules = dict_data['rules']

                        # Or it might be a direct field mapping
                        if not schema and not rules:
                            for key, value in dict_data.items():
                                if isinstance(value, dict):
                                    schema[key] = value.get('type', 'str')
                                    if 'min' in value or 'max' in value:
                                        rules[key] = {}
                                        if 'min' in value:
                                            rules[key]['min'] = value['min']
                                        if 'max' in value:
                                            rules[key]['max'] = value['max']
                                else:
                                    # Simple type mapping
                                    schema[key] = value

                return {
                    "success": True,
                    "schema": schema,
                    "rules": rules,
                    "metadata": {
                        "source": "parsed",
                        "fields_count": len(schema),
                        "rules_count": len(rules)
                    }
                }

            except Exception as e:
                return {"error": f"Dictionary parsing failed: {str(e)}"}

        return {"error": "Unknown tool"}

    async def analyze_data(self, data_content: str, file_format: str = "csv",
                          schema: Dict = None, rules: Dict = None,
                          min_rows: int = 1, debug: bool = False):
        """Analyze data with optional schema and rules"""
        request_data = {
            "tool": "analyze_data",
            "arguments": {
                "data_content": data_content,
                "file_format": file_format,
                "schema": schema or {},
                "rules": rules or {},
                "min_rows": min_rows
            }
        }
        return self._simulate_mcp_call(request_data, debug)

    async def parse_data_dictionary(self, dictionary_content: str,
                                   format_hint: str = "auto",
                                   use_cache: bool = True,
                                   debug: bool = False) -> Dict:
        """Parse data dictionary"""
        request_data = {
            "tool": "parse_data_dictionary",
            "arguments": {
                "dictionary_content": dictionary_content,
                "format_hint": format_hint,
                "use_cache": use_cache
            }
        }
        return self._simulate_mcp_call(request_data, debug)


def main():
    """Main application with minimalist interface"""

    # Initialize session state
    if 'analysis_complete' not in st.session_state:
        st.session_state.analysis_complete = False
    if 'analysis_results' not in st.session_state:
        st.session_state.analysis_results = None
    if 'uploaded_file' not in st.session_state:
        st.session_state.uploaded_file = None
    if 'dictionary_file' not in st.session_state:
        st.session_state.dictionary_file = None

    # Create horizontal navigation menu
    selected = option_menu(
        menu_title=None,  # No title for horizontal menu
        options=["Home", "About", "Help"],
        icons=["house", "info-circle", "question-circle"],
        menu_icon="cast",
        default_index=0,
        orientation="horizontal",
        styles={
            "container": {"padding": "0!important", "background-color": "#fafafa"},
            "icon": {"color": "#31333f", "font-size": "18px"},
            "nav-link": {
                "font-size": "16px",
                "text-align": "center",
                "margin": "0px",
                "--hover-color": "#eee",
            },
            "nav-link-selected": {"background-color": "#4CAF50"},
        }
    )

    # Route to different pages based on selection
    if selected == "Home":
        home_page()
    elif selected == "About":
        about_page()
    elif selected == "Help":
        help_page()


def home_page():
    """Main home page with data analysis functionality"""

    # Minimal header
    st.markdown("<h1>📊 Data Quality Analyzer</h1>", unsafe_allow_html=True)
    st.markdown("<p class='subtitle'>Simple, Fast, Accurate</p>", unsafe_allow_html=True)

    # Create three columns for the main interface
    col1, col2, col3 = st.columns([2, 2, 1])

    with col1:
        st.markdown("<div class='main-card'>", unsafe_allow_html=True)
        st.markdown("### 1️⃣ Upload Data")

        # File uploader
        uploaded_file = st.file_uploader(
            "Choose a file",
            type=['csv', 'json', 'xlsx', 'xls', 'parquet'],
            help="Upload your data file (CSV, JSON, Excel, or Parquet)",
            key="data_uploader"
        )

        if uploaded_file:
            st.session_state.uploaded_file = uploaded_file
            file_size = len(uploaded_file.read())
            uploaded_file.seek(0)  # Reset file pointer
            st.success(f"✅ {uploaded_file.name} ({file_size:,} bytes)")

        # Demo data in expander
        with st.expander("📁 Or use demo data"):
            demo_choice = st.selectbox(
                "Select demo dataset",
                ["None", "CSV - Western Names", "CSV - Asian Names",
                 "JSON - Mixed Names", "CSV - Clinical Trial"],
                key="demo_selector"
            )

            if demo_choice != "None":
                # Load demo data
                if demo_choice == "CSV - Western Names":
                    demo_data = """name,age,email,country
John Smith,25,john@email.com,USA
Jane Doe,30,jane@email.com,Canada
Bob Wilson,invalid,bob@email,UK"""
                    st.session_state.uploaded_file = demo_data
                    st.success("✅ Demo data loaded")
                elif demo_choice == "CSV - Asian Names":
                    demo_data = """name,age,city,join_date
李明,28,Beijing,2023-01-15
田中太郎,35,Tokyo,2023-02-20
김철수,42,Seoul,invalid-date"""
                    st.session_state.uploaded_file = demo_data
                    st.success("✅ Demo data loaded")
                elif demo_choice == "JSON - Mixed Names":
                    demo_data = json.dumps([
                        {"name": "Alice Smith", "age": 30, "score": 85.5},
                        {"name": "王小明", "age": 25, "score": 92.0},
                        {"name": "Carlos García", "age": "invalid", "score": 88.0}
                    ])
                    st.session_state.uploaded_file = demo_data
                    st.success("✅ Demo data loaded")
                elif demo_choice == "CSV - Clinical Trial":
                    demo_data = """patient_id,age,treatment,outcome,date
P001,45,Drug_A,Improved,2023-01-15
P002,52,Drug_B,No_Change,2023-01-20
P003,999,Drug_A,Improved,invalid"""
                    st.session_state.uploaded_file = demo_data
                    st.success("✅ Demo data loaded")

        st.markdown("</div>", unsafe_allow_html=True)

    with col2:
        st.markdown("<div class='main-card'>", unsafe_allow_html=True)
        st.markdown("### 2️⃣ Upload Data Dictionary")
        st.markdown("*Optional - defines expected data structure*")

        dictionary_file = st.file_uploader(
            "Choose a dictionary file",
            type=['csv', 'json', 'txt'],
            help="Upload your data dictionary (optional)",
            key="dict_uploader"
        )

        if dictionary_file:
            st.session_state.dictionary_file = dictionary_file
            st.success(f"✅ {dictionary_file.name}")

        # Demo dictionaries in expander
        with st.expander("📚 Or use demo dictionary"):
            dict_choice = st.selectbox(
                "Select demo dictionary",
                ["None"] + list(DEMO_DICTIONARIES.keys()),
                key="dict_demo_selector"
            )

            if dict_choice != "None":
                st.session_state.dictionary_file = get_demo_dictionary(dict_choice)
                st.success(f"✅ {dict_choice} dictionary loaded")

        st.markdown("</div>", unsafe_allow_html=True)

    with col3:
        st.markdown("<div class='main-card'>", unsafe_allow_html=True)
        st.markdown("### 3️⃣ Run Analysis")
        st.markdown("&nbsp;")

        # Big green run button
        run_button = st.button(
            "🚀 Analyze",
            type="primary",
            use_container_width=True,
            disabled=not st.session_state.uploaded_file
        )

        if not st.session_state.uploaded_file:
            st.info("Upload data first")

        st.markdown("</div>", unsafe_allow_html=True)

    # Advanced options in sidebar (collapsed by default)
    with st.sidebar:
        st.markdown("### ⚙️ Advanced Options")

        with st.expander("🔧 Analysis Settings"):
            min_rows = st.number_input("Minimum rows required", min_value=0, value=1)
            debug_mode = st.checkbox("Debug mode", value=False)

        with st.expander("📋 Manual Schema Definition"):
            st.markdown("Define expected data types for columns")
            schema = {}

            if st.session_state.uploaded_file:
                # Try to get column names
                try:
                    if isinstance(st.session_state.uploaded_file, str):
                        if st.session_state.uploaded_file.startswith('[') or st.session_state.uploaded_file.startswith('{'):
                            # JSON
                            df_preview = pd.json_normalize(json.loads(st.session_state.uploaded_file))
                        else:
                            # CSV
                            df_preview = pd.read_csv(io.StringIO(st.session_state.uploaded_file))
                    else:
                        file_ext = st.session_state.uploaded_file.name.split('.')[-1].lower()
                        if file_ext == 'csv':
                            df_preview = pd.read_csv(st.session_state.uploaded_file)
                        elif file_ext == 'json':
                            df_preview = pd.read_json(st.session_state.uploaded_file)
                        elif file_ext in ['xlsx', 'xls']:
                            df_preview = pd.read_excel(st.session_state.uploaded_file)
                        elif file_ext == 'parquet':
                            df_preview = pd.read_parquet(st.session_state.uploaded_file)
                        st.session_state.uploaded_file.seek(0)

                    for col in df_preview.columns:
                        col_type = st.selectbox(
                            f"{col}",
                            ["auto", "str", "int", "float", "bool", "datetime"],
                            key=f"schema_{col}"
                        )
                        if col_type != "auto":
                            schema[col] = col_type
                except:
                    st.info("Upload data to define schema")

        with st.expander("📏 Validation Rules"):
            st.markdown("Define validation rules for columns")
            rules = {}

            if st.session_state.uploaded_file:
                try:
                    # Add rule inputs based on columns
                    st.markdown("*Numeric columns can have min/max values*")
                    # Simplified rules interface
                    rule_col = st.selectbox("Select column", ["None"] + (df_preview.columns.tolist() if 'df_preview' in locals() else []))
                    if rule_col != "None":
                        min_val = st.number_input(f"Min value for {rule_col}", value=0.0)
                        max_val = st.number_input(f"Max value for {rule_col}", value=100.0)
                        if st.button(f"Add rule for {rule_col}"):
                            if rule_col not in rules:
                                rules[rule_col] = {}
                            rules[rule_col]["min"] = min_val
                            rules[rule_col]["max"] = max_val
                            st.success(f"Rule added for {rule_col}")
                except:
                    st.info("Upload data to define rules")

        with st.expander("📊 Display Options"):
            show_preview = st.checkbox("Show data preview", value=True)
            show_statistics = st.checkbox("Show statistics", value=True)
            max_preview_rows = st.slider("Preview rows", 5, 50, 10)

    # Divider
    st.markdown("---")

    # Run analysis when button clicked
    if run_button and st.session_state.uploaded_file:
        with st.spinner("🔍 Analyzing data..."):
            # Initialize MCP client
            client = MCPClient()

            # Parse data dictionary if provided
            dict_schema = {}
            dict_rules = {}

            if st.session_state.dictionary_file:
                with st.spinner("📖 Parsing data dictionary..."):
                    # Get dictionary content
                    if isinstance(st.session_state.dictionary_file, str):
                        # Demo dictionary
                        dict_content = st.session_state.dictionary_file
                    else:
                        # Uploaded file
                        dict_content = st.session_state.dictionary_file.read().decode('utf-8')
                        st.session_state.dictionary_file.seek(0)

                    # Parse dictionary
                    dict_result = asyncio.run(
                        client.parse_data_dictionary(
                            dictionary_content=dict_content,
                            format_hint="auto",
                            use_cache=True,
                            debug=debug_mode if 'debug_mode' in locals() else False
                        )
                    )

                    if dict_result.get("success"):
                        dict_schema = dict_result.get("schema", {})
                        dict_rules = dict_result.get("rules", {})
                        st.success(f"✅ Dictionary parsed: {len(dict_schema)} fields, {len(dict_rules)} rules")
                    elif "error" in dict_result:
                        st.warning(f"⚠️ Dictionary parsing failed: {dict_result['error']}")

            # Merge manual schema/rules with dictionary
            final_schema = {**dict_schema, **(schema if 'schema' in locals() else {})}
            final_rules = {**dict_rules, **(rules if 'rules' in locals() else {})}

            # Prepare data content
            if isinstance(st.session_state.uploaded_file, str):
                # Demo data (already a string)
                data_content = st.session_state.uploaded_file
                if data_content.startswith('[') or data_content.startswith('{'):
                    file_format = "json"
                else:
                    file_format = "csv"
            else:
                # Uploaded file
                file_ext = st.session_state.uploaded_file.name.split('.')[-1].lower()
                file_format = file_ext

                if file_ext in ['xlsx', 'xls', 'parquet']:
                    # Binary formats need base64 encoding
                    file_bytes = st.session_state.uploaded_file.read()
                    data_content = base64.b64encode(file_bytes).decode('utf-8')
                    st.session_state.uploaded_file.seek(0)
                else:
                    # Text formats
                    data_content = st.session_state.uploaded_file.read().decode('utf-8')
                    st.session_state.uploaded_file.seek(0)

            # Run analysis
            result = asyncio.run(
                client.analyze_data(
                    data_content=data_content,
                    file_format=file_format,
                    schema=final_schema if final_schema else None,
                    rules=final_rules if final_rules else None,
                    min_rows=min_rows if 'min_rows' in locals() else 1,
                    debug=debug_mode if 'debug_mode' in locals() else False
                )
            )

            st.session_state.analysis_results = result
            st.session_state.analysis_complete = True

    # Display results
    if st.session_state.analysis_complete and st.session_state.analysis_results:
        result = st.session_state.analysis_results

        if "error" in result:
            st.error(f"❌ Analysis Error: {result['error']}")
        else:
            # Success - show results in a clean layout
            st.markdown("## 📈 Analysis Results")

            # Summary metrics in columns
            summary = result.get("summary", {})

            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Rows", f"{summary.get('row_count', 0):,}")
            with col2:
                st.metric("Columns", summary.get('column_count', 0))
            with col3:
                issues_count = summary.get('issues_count', 0)
                st.metric("Issues Found", issues_count,
                         delta=None if issues_count == 0 else f"{issues_count} issues",
                         delta_color="inverse")
            with col4:
                missing_total = sum(summary.get('missing_values', {}).values())
                st.metric("Missing Values", f"{missing_total:,}")

            # Issues section
            if summary.get('issues'):
                st.markdown("### 🚨 Data Quality Issues")

                # Group issues by severity
                errors = [i for i in summary['issues'] if i.get('severity') == 'error']
                warnings = [i for i in summary['issues'] if i.get('severity') == 'warning']

                if errors:
                    st.markdown("#### ❌ Errors")
                    for issue in errors:
                        col = issue.get('column', '')
                        if col:
                            st.error(f"**{col}**: {issue['message']}")
                        else:
                            st.error(issue['message'])

                if warnings:
                    st.markdown("#### ⚠️ Warnings")
                    for issue in warnings:
                        col = issue.get('column', '')
                        if col:
                            st.warning(f"**{col}**: {issue['message']}")
                        else:
                            st.warning(issue['message'])
            else:
                st.success("✅ No data quality issues found!")

            # Data preview in expander
            if show_preview if 'show_preview' in locals() else True:
                with st.expander("📋 Data Preview", expanded=True):
                    preview_data = result.get("preview", [])
                    if preview_data:
                        df_preview = pd.DataFrame(preview_data)
                        st.dataframe(
                            df_preview,
                            use_container_width=True,
                            hide_index=True
                        )

            # Statistics in expander
            if show_statistics if 'show_statistics' in locals() else True:
                with st.expander("📊 Column Statistics"):
                    col1, col2 = st.columns(2)

                    with col1:
                        st.markdown("**Data Types**")
                        dtypes_df = pd.DataFrame(
                            list(summary.get('dtypes', {}).items()),
                            columns=['Column', 'Type']
                        )
                        st.dataframe(dtypes_df, use_container_width=True, hide_index=True)

                    with col2:
                        st.markdown("**Missing Values**")
                        missing_df = pd.DataFrame(
                            list(summary.get('missing_values', {}).items()),
                            columns=['Column', 'Missing Count']
                        )
                        missing_df = missing_df[missing_df['Missing Count'] > 0]
                        if not missing_df.empty:
                            st.dataframe(missing_df, use_container_width=True, hide_index=True)
                        else:
                            st.info("No missing values")

            # Download section in expander
            with st.expander("💾 Download Results"):
                col1, col2, col3 = st.columns(3)

                with col1:
                    # Create downloadable report
                    report = {
                        "analysis_date": datetime.now().isoformat(),
                        "summary": summary,
                        "issues": summary.get('issues', [])
                    }
                    report_json = json.dumps(report, indent=2)

                    st.download_button(
                        label="📄 Download Report (JSON)",
                        data=report_json,
                        file_name=f"data_quality_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                        mime="application/json"
                    )

                with col2:
                    if summary.get('issues'):
                        # Create CSV of issues
                        issues_df = pd.DataFrame(summary['issues'])
                        csv = issues_df.to_csv(index=False)
                        st.download_button(
                            label="📊 Download Issues (CSV)",
                            data=csv,
                            file_name=f"data_quality_issues_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                            mime="text/csv"
                        )

                with col3:
                    # Excel with error highlighting
                    if result.get("preview"):
                        try:
                            import openpyxl
                            from openpyxl.styles import PatternFill, Font, Comment
                            from openpyxl.workbook import Workbook

                            # Create Excel workbook
                            wb = Workbook()
                            ws = wb.active
                            ws.title = "Data with Errors"

                            # Get data
                            df_export = pd.DataFrame(result["preview"])

                            # Write headers
                            for col_idx, col_name in enumerate(df_export.columns, 1):
                                ws.cell(row=1, column=col_idx, value=col_name)
                                ws.cell(row=1, column=col_idx).font = Font(bold=True)

                            # Write data and highlight errors
                            red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
                            yellow_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")

                            # Track which cells have issues
                            error_cells = {}
                            for issue in summary.get('issues', []):
                                if 'column' in issue:
                                    col = issue['column']
                                    if col not in error_cells:
                                        error_cells[col] = []
                                    error_cells[col].append({
                                        'severity': issue.get('severity', 'error'),
                                        'message': issue.get('message', '')
                                    })

                            # Write data rows
                            for row_idx, row in df_export.iterrows():
                                for col_idx, (col_name, value) in enumerate(row.items(), 1):
                                    cell = ws.cell(row=row_idx + 2, column=col_idx, value=value)

                                    # Highlight errors
                                    if col_name in error_cells:
                                        for error_info in error_cells[col_name]:
                                            if error_info['severity'] == 'error':
                                                cell.fill = red_fill
                                            else:
                                                cell.fill = yellow_fill

                                            # Add comment with error message
                                            if cell.comment is None:
                                                cell.comment = Comment(error_info['message'], "Data Analyzer")

                            # Add summary sheet
                            ws2 = wb.create_sheet("Analysis Summary")
                            ws2['A1'] = "Data Quality Analysis Summary"
                            ws2['A1'].font = Font(bold=True, size=14)

                            ws2['A3'] = "Metrics"
                            ws2['A3'].font = Font(bold=True)
                            ws2['A4'] = "Total Rows:"
                            ws2['B4'] = summary.get('row_count', 0)
                            ws2['A5'] = "Total Columns:"
                            ws2['B5'] = summary.get('column_count', 0)
                            ws2['A6'] = "Issues Found:"
                            ws2['B6'] = summary.get('issues_count', 0)

                            if summary.get('issues'):
                                ws2['A8'] = "Issues Detail"
                                ws2['A8'].font = Font(bold=True)
                                row = 9
                                for issue in summary['issues']:
                                    ws2[f'A{row}'] = issue.get('severity', 'info').upper()
                                    ws2[f'B{row}'] = issue.get('column', 'General')
                                    ws2[f'C{row}'] = issue.get('message', '')
                                    row += 1

                            # Save to bytes
                            excel_buffer = io.BytesIO()
                            wb.save(excel_buffer)
                            excel_buffer.seek(0)

                            st.download_button(
                                label="📊 Download Excel with Errors",
                                data=excel_buffer.getvalue(),
                                file_name=f"data_with_errors_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                help="Download Excel file with errors highlighted in red"
                            )

                        except ImportError:
                            st.info("Install openpyxl for Excel export")


if __name__ == "__main__":
    main()